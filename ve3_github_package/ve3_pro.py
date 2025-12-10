#!/usr/bin/env python3
"""
VE3 Tool Pro v2.0
=================
Beautiful, Smart, Powerful
1 Click: Voice → Images
"""

import os
import sys
import json
import shutil
import threading
import webbrowser
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Setup
ROOT_DIR = Path(__file__).parent
sys.path.insert(0, str(ROOT_DIR))

# Support external config/projects directories (for auto-update setup)
CONFIG_DIR = Path(os.environ.get('VE3_CONFIG_DIR', ROOT_DIR / "config"))
PROJECTS_DIR = Path(os.environ.get('VE3_PROJECTS_DIR', ROOT_DIR / "PROJECTS"))

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except:
    HAS_PIL = False


class VE3ToolPro:
    """VE3 Tool Pro - Beautiful GUI."""
    
    VERSION = "2.0"
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("VE3 Tool Pro")
        self.root.geometry("1280x800")
        self.root.minsize(1100, 700)
        
        # Config background
        self.root.configure(bg='#f0f0f0')
        
        # Variables
        self.input_mode = tk.StringVar(value="file")
        self.input_path = tk.StringVar()
        
        # Data
        self.profiles: List[str] = []
        self.groq_keys: List[str] = []
        self.gemini_keys: List[str] = []
        self.chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        
        # State
        self._running = False
        self._stop = False
        self._engine = None
        
        # Current project data
        self.current_project_dir: Optional[Path] = None
        self.characters: List[Dict] = []
        self.scenes: List[Dict] = []
        
        # Image cache
        self.image_cache = {}
        
        # Load config
        self.load_config()
        
        # Create UI
        self.setup_styles()
        self.create_ui()
        
        # Initial state
        self.update_resource_display()
    
    def setup_styles(self):
        """Setup ttk styles."""
        style = ttk.Style()
        
        # Use clam theme
        style.theme_use('clam')
        
        # Custom fonts
        style.configure('.', font=('Segoe UI', 10))
        style.configure('Title.TLabel', font=('Segoe UI', 20, 'bold'), foreground='#2c3e50')
        style.configure('Subtitle.TLabel', font=('Segoe UI', 11), foreground='#7f8c8d')
        style.configure('Section.TLabelframe.Label', font=('Segoe UI', 10, 'bold'))
        style.configure('Big.TButton', font=('Segoe UI', 14, 'bold'), padding=(20, 15))
        style.configure('Status.TLabel', font=('Segoe UI', 9))
        style.configure('Card.TFrame', background='white')
        
        # Map colors
        style.map('Big.TButton',
            background=[('active', '#27ae60'), ('!active', '#2ecc71')],
            foreground=[('active', 'white'), ('!active', 'white')])
    
    def create_ui(self):
        """Create main UI layout."""
        # Main container
        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # === LEFT PANEL (Controls) ===
        left_panel = ttk.Frame(main, width=350)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 15))
        left_panel.pack_propagate(False)
        
        self.create_controls(left_panel)
        
        # === RIGHT PANEL (Preview & Log) ===
        right_panel = ttk.Frame(main)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.create_preview(right_panel)
    
    def create_controls(self, parent):
        """Create left control panel."""
        
        # === HEADER ===
        header = ttk.Frame(parent)
        header.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(header, text="🎨 VE3 Tool", style='Title.TLabel').pack(anchor=tk.W)
        ttk.Label(header, text="Voice → Images (1 Click)", style='Subtitle.TLabel').pack(anchor=tk.W)
        
        # === 1. INPUT ===
        input_frame = ttk.LabelFrame(parent, text=" 📁 Đầu vào ", padding=10)
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Mode selection
        mode_row = ttk.Frame(input_frame)
        mode_row.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Radiobutton(mode_row, text="📄 File đơn", variable=self.input_mode, 
                        value="file", command=self.on_mode_change).pack(side=tk.LEFT)
        ttk.Radiobutton(mode_row, text="📂 Thư mục", variable=self.input_mode,
                        value="folder", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)
        
        # Path row
        path_row = ttk.Frame(input_frame)
        path_row.pack(fill=tk.X)
        
        self.path_entry = ttk.Entry(path_row, textvariable=self.input_path, font=('Consolas', 9))
        self.path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        ttk.Button(path_row, text="Chọn...", width=8, command=self.browse_input).pack(side=tk.LEFT)
        
        # Input info
        self.input_info_label = ttk.Label(input_frame, text="Hỗ trợ: .mp3, .wav, .xlsx", 
                                          foreground='gray', font=('Segoe UI', 9))
        self.input_info_label.pack(anchor=tk.W, pady=(5, 0))
        
        # === 2. START BUTTON ===
        self.start_btn = tk.Button(
            parent, text="▶  BẮT ĐẦU", 
            font=('Segoe UI', 14, 'bold'),
            bg='#2ecc71', fg='white', activebackground='#27ae60',
            relief=tk.FLAT, cursor='hand2',
            command=self.start_processing
        )
        self.start_btn.pack(fill=tk.X, pady=10, ipady=12)
        
        # Stop button
        self.stop_btn = tk.Button(
            parent, text="⏹  Dừng",
            font=('Segoe UI', 10),
            bg='#e74c3c', fg='white', activebackground='#c0392b',
            relief=tk.FLAT, state=tk.DISABLED,
            command=self.stop_processing
        )
        self.stop_btn.pack(fill=tk.X, pady=(0, 10))
        
        # === 3. PROGRESS ===
        progress_frame = ttk.LabelFrame(parent, text=" 📊 Tiến độ ", padding=10)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))
        
        self.progress_label = ttk.Label(progress_frame, text="Sẵn sàng", font=('Segoe UI', 10, 'bold'))
        self.progress_label.pack(anchor=tk.W)
        
        self.progress_detail = ttk.Label(progress_frame, text="", foreground='gray')
        self.progress_detail.pack(anchor=tk.W)
        
        # === 4. RESOURCES ===
        res_frame = ttk.LabelFrame(parent, text=" 🔧 Tài nguyên ", padding=10)
        res_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.res_profiles = ttk.Label(res_frame, text="👤 Profiles: 0")
        self.res_profiles.pack(anchor=tk.W)
        
        self.res_groq = ttk.Label(res_frame, text="🔑 Groq: 0")
        self.res_groq.pack(anchor=tk.W)
        
        self.res_gemini = ttk.Label(res_frame, text="🔑 Gemini: 0")
        self.res_gemini.pack(anchor=tk.W)
        
        # Config buttons
        btn_row = ttk.Frame(res_frame)
        btn_row.pack(fill=tk.X, pady=(8, 0))
        
        ttk.Button(btn_row, text="⚙️ Cài đặt", command=self.open_settings).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="🔄 Reload", command=self.reload_config).pack(side=tk.LEFT)
        
        # === 5. QUICK ACTIONS ===
        actions_frame = ttk.LabelFrame(parent, text=" ⚡ Thao tác nhanh ", padding=10)
        actions_frame.pack(fill=tk.X)
        
        ttk.Button(actions_frame, text="📂 Mở Output", command=self.open_output_folder).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(actions_frame, text="🔑 Lấy Token thủ công", command=self.get_token_manual).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(actions_frame, text="📋 Mở Config", command=self.open_config_file).pack(fill=tk.X)
    
    def create_preview(self, parent):
        """Create right preview panel."""
        
        # Notebook for tabs
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # === TAB 1: PREVIEW ===
        preview_tab = ttk.Frame(notebook, padding=10)
        notebook.add(preview_tab, text="  🖼️ Preview  ")
        
        # Top: Character & Scene selection
        select_row = ttk.Frame(preview_tab)
        select_row.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(select_row, text="Nhân vật:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        self.char_combo = ttk.Combobox(select_row, state='readonly', width=20)
        self.char_combo.pack(side=tk.LEFT, padx=(5, 20))
        self.char_combo.bind('<<ComboboxSelected>>', self.on_char_selected)
        
        ttk.Label(select_row, text="Scene:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        self.scene_combo = ttk.Combobox(select_row, state='readonly', width=20)
        self.scene_combo.pack(side=tk.LEFT, padx=(5, 20))
        self.scene_combo.bind('<<ComboboxSelected>>', self.on_scene_selected)
        
        ttk.Button(select_row, text="🔄 Refresh", command=self.refresh_preview).pack(side=tk.RIGHT)
        
        # Main preview area (3 columns)
        preview_main = ttk.Frame(preview_tab)
        preview_main.pack(fill=tk.BOTH, expand=True)
        
        # Column 1: Character
        char_col = ttk.LabelFrame(preview_main, text=" 👤 Nhân vật ", padding=5)
        char_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        self.char_image_label = ttk.Label(char_col, text="Chưa có", anchor=tk.CENTER, background='#ecf0f1')
        self.char_image_label.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        self.char_prompt_text = tk.Text(char_col, height=4, wrap=tk.WORD, font=('Segoe UI', 9), bg='#f9f9f9')
        self.char_prompt_text.pack(fill=tk.X)
        
        # Column 2: Reference (optional)
        ref_col = ttk.LabelFrame(preview_main, text=" 🎨 Tham chiếu ", padding=5)
        ref_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        self.ref_image_label = ttk.Label(ref_col, text="(Tùy chọn)", anchor=tk.CENTER, background='#ecf0f1')
        self.ref_image_label.pack(fill=tk.BOTH, expand=True)
        
        # Column 3: Result
        result_col = ttk.LabelFrame(preview_main, text=" ✨ Kết quả ", padding=5)
        result_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        self.result_image_label = ttk.Label(result_col, text="Chưa có", anchor=tk.CENTER, background='#ecf0f1')
        self.result_image_label.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        self.result_prompt_text = tk.Text(result_col, height=4, wrap=tk.WORD, font=('Segoe UI', 9), bg='#f9f9f9')
        self.result_prompt_text.pack(fill=tk.X)
        
        # Bottom: Scene thumbnails
        thumb_frame = ttk.LabelFrame(preview_tab, text=" 📋 Tất cả Scenes ", padding=5)
        thumb_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Scrollable canvas for thumbnails
        self.thumb_canvas = tk.Canvas(thumb_frame, height=90, bg='white')
        self.thumb_canvas.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        thumb_scroll = ttk.Scrollbar(thumb_frame, orient=tk.HORIZONTAL, command=self.thumb_canvas.xview)
        thumb_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.thumb_canvas.configure(xscrollcommand=thumb_scroll.set)
        
        # === TAB 2: LOG ===
        log_tab = ttk.Frame(notebook, padding=10)
        notebook.add(log_tab, text="  📝 Log  ")
        
        self.log_text = tk.Text(log_tab, wrap=tk.WORD, font=('Consolas', 9), bg='#1e1e1e', fg='#d4d4d4')
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Log scrollbar
        log_scroll = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        
        # Log buttons
        log_btn_row = ttk.Frame(log_tab)
        log_btn_row.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(log_btn_row, text="🗑️ Xóa", command=self.clear_log).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(log_btn_row, text="💾 Lưu", command=self.save_log).pack(side=tk.LEFT)
        
        # === TAB 3: PROMPTS ===
        prompts_tab = ttk.Frame(notebook, padding=10)
        notebook.add(prompts_tab, text="  📄 Prompts  ")
        
        # Characters table
        char_lf = ttk.LabelFrame(prompts_tab, text=" 👥 Nhân vật ", padding=5)
        char_lf.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        cols = ('id', 'prompt', 'status')
        self.char_tree = ttk.Treeview(char_lf, columns=cols, show='headings', height=5)
        self.char_tree.heading('id', text='ID')
        self.char_tree.heading('prompt', text='Prompt')
        self.char_tree.heading('status', text='Status')
        self.char_tree.column('id', width=80)
        self.char_tree.column('prompt', width=500)
        self.char_tree.column('status', width=80)
        self.char_tree.pack(fill=tk.BOTH, expand=True)
        
        # Scenes table
        scene_lf = ttk.LabelFrame(prompts_tab, text=" 🎬 Scenes ", padding=5)
        scene_lf.pack(fill=tk.BOTH, expand=True)
        
        cols2 = ('id', 'time', 'prompt', 'status')
        self.scene_tree = ttk.Treeview(scene_lf, columns=cols2, show='headings', height=10)
        self.scene_tree.heading('id', text='ID')
        self.scene_tree.heading('time', text='Thời gian')
        self.scene_tree.heading('prompt', text='Prompt')
        self.scene_tree.heading('status', text='Status')
        self.scene_tree.column('id', width=60)
        self.scene_tree.column('time', width=100)
        self.scene_tree.column('prompt', width=420)
        self.scene_tree.column('status', width=80)
        self.scene_tree.pack(fill=tk.BOTH, expand=True)
    
    # ========== ACTIONS ==========
    
    def on_mode_change(self):
        """Handle mode change."""
        self.input_path.set("")
        if self.input_mode.get() == "folder":
            self.input_info_label.config(text="Mỗi file voice trong thư mục = 1 dự án")
        else:
            self.input_info_label.config(text="Hỗ trợ: .mp3, .wav, .xlsx")
    
    def browse_input(self):
        """Browse for input."""
        if self.input_mode.get() == "folder":
            path = filedialog.askdirectory(title="Chọn thư mục chứa voice")
        else:
            path = filedialog.askopenfilename(
                title="Chọn file",
                filetypes=[
                    ("Supported", "*.mp3 *.wav *.xlsx"),
                    ("Audio", "*.mp3 *.wav"),
                    ("Excel", "*.xlsx"),
                    ("All", "*.*")
                ]
            )
        
        if path:
            self.input_path.set(path)
            self.update_input_info()
    
    def update_input_info(self):
        """Update input info display."""
        path = self.input_path.get()
        if not path:
            return
        
        p = Path(path)
        if self.input_mode.get() == "folder":
            voices = list(p.glob("*.mp3")) + list(p.glob("*.wav"))
            self.input_info_label.config(text=f"📁 {len(voices)} file voice")
        else:
            size = p.stat().st_size / 1024 if p.exists() else 0
            self.input_info_label.config(text=f"📄 {p.name} ({size:.1f} KB)")
    
    def log(self, msg: str, level: str = "INFO"):
        """Add log message."""
        ts = datetime.now().strftime("%H:%M:%S")
        
        # Color tags
        colors = {
            "OK": "#2ecc71",
            "ERROR": "#e74c3c", 
            "WARN": "#f39c12",
            "INFO": "#3498db"
        }
        
        prefix = {"OK": "✅", "ERROR": "❌", "WARN": "⚠️", "INFO": "ℹ️"}.get(level, "•")
        
        self.log_text.insert(tk.END, f"[{ts}] {prefix} {msg}\n")
        self.log_text.see(tk.END)
    
    def clear_log(self):
        self.log_text.delete(1.0, tk.END)
    
    def save_log(self):
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text", "*.txt")])
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(self.log_text.get(1.0, tk.END))
            self.log(f"Saved log to {path}", "OK")
    
    def update_progress(self, percent: float, text: str = "", detail: str = ""):
        """Update progress display."""
        self.progress_var.set(percent)
        if text:
            self.progress_label.config(text=text)
        if detail:
            self.progress_detail.config(text=detail)
    
    def update_resource_display(self):
        """Update resource display."""
        self.res_profiles.config(text=f"👤 Profiles: {len(self.profiles)}" + 
                                 (" ✅" if self.profiles else " ⚠️"))
        self.res_groq.config(text=f"🔑 Groq: {len(self.groq_keys)}" +
                            (" ✅" if self.groq_keys else ""))
        self.res_gemini.config(text=f"🔑 Gemini: {len(self.gemini_keys)}")
    
    # ========== CONFIG ==========
    
    def load_config(self):
        """Load config from accounts.json."""
        accounts_file = CONFIG_DIR / "accounts.json"
        
        if not accounts_file.exists():
            self.create_default_config()
            return
        
        try:
            with open(accounts_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.chrome_path = data.get('chrome_path', self.chrome_path)
            
            # Profiles
            self.profiles = []
            for p in data.get('chrome_profiles', []):
                path = p if isinstance(p, str) else p.get('path', '')
                if path and not path.startswith('THAY_BANG') and Path(path).exists():
                    self.profiles.append(path)
            
            # API keys
            api = data.get('api_keys', {})
            self.groq_keys = [k for k in api.get('groq', []) 
                            if k and not k.startswith('THAY_BANG')]
            self.gemini_keys = [k for k in api.get('gemini', [])
                              if k and not k.startswith('THAY_BANG')]
            
        except Exception as e:
            print(f"Load config error: {e}")
    
    def create_default_config(self):
        """Create default config file."""
        CONFIG_DIR.mkdir(exist_ok=True)
        
        default = {
            "_README": [
                "=== VE3 TOOL CONFIG ===",
                "Điền thông tin Chrome profiles và API keys vào đây",
                "Groq API FREE: https://console.groq.com/keys"
            ],
            "chrome_path": r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            "chrome_profiles": [
                "THAY_BANG_DUONG_DAN_PROFILE_1",
                "THAY_BANG_DUONG_DAN_PROFILE_2"
            ],
            "api_keys": {
                "groq": ["THAY_BANG_GROQ_KEY"],
                "gemini": []
            },
            "settings": {
                "parallel": 2,
                "delay_between_images": 2
            }
        }
        
        with open(CONFIG_DIR / "accounts.json", 'w', encoding='utf-8') as f:
            json.dump(default, f, indent=4, ensure_ascii=False)
    
    def reload_config(self):
        """Reload config."""
        self.load_config()
        self.update_resource_display()
        self.log("Đã reload config", "OK")
    
    def open_settings(self):
        """Open settings dialog."""
        win = tk.Toplevel(self.root)
        win.title("⚙️ Cài đặt")
        win.geometry("650x500")
        win.transient(self.root)
        win.grab_set()
        
        notebook = ttk.Notebook(win)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tab: Profiles
        prof_tab = ttk.Frame(notebook, padding=15)
        notebook.add(prof_tab, text="  👤 Chrome Profiles  ")
        
        ttk.Label(prof_tab, text="Danh sách Chrome Profiles:", 
                  font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        
        prof_list = tk.Listbox(prof_tab, height=10, font=('Consolas', 9))
        prof_list.pack(fill=tk.BOTH, expand=True)
        
        for p in self.profiles:
            prof_list.insert(tk.END, p)
        
        if not self.profiles:
            prof_list.insert(tk.END, "(Chưa có profile nào)")
        
        ttk.Label(prof_tab, text="💡 Mở config để thêm/sửa profiles:", foreground='gray').pack(anchor=tk.W, pady=(10, 5))
        ttk.Button(prof_tab, text="📂 Mở accounts.json", command=self.open_config_file).pack(anchor=tk.W)
        
        # Tab: API Keys
        api_tab = ttk.Frame(notebook, padding=15)
        notebook.add(api_tab, text="  🔑 API Keys  ")
        
        # Groq
        ttk.Label(api_tab, text="Groq Keys (FREE):", font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W)
        
        groq_link = ttk.Label(api_tab, text="🔗 Lấy tại: console.groq.com/keys", 
                             foreground='blue', cursor='hand2')
        groq_link.pack(anchor=tk.W)
        groq_link.bind('<Button-1>', lambda e: webbrowser.open("https://console.groq.com/keys"))
        
        groq_list = tk.Listbox(api_tab, height=4, font=('Consolas', 9))
        groq_list.pack(fill=tk.X, pady=(5, 15))
        for k in self.groq_keys:
            groq_list.insert(tk.END, k[:30] + "..." if len(k) > 30 else k)
        
        # Gemini
        ttk.Label(api_tab, text="Gemini Keys:", font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W)
        
        gem_list = tk.Listbox(api_tab, height=4, font=('Consolas', 9))
        gem_list.pack(fill=tk.X, pady=(5, 10))
        for k in self.gemini_keys:
            gem_list.insert(tk.END, k[:30] + "..." if len(k) > 30 else k)
        
        # Tab: Help
        help_tab = ttk.Frame(notebook, padding=15)
        notebook.add(help_tab, text="  ❓ Hướng dẫn  ")
        
        help_text = """
🎯 CÁCH SỬ DỤNG:

1️⃣ Thêm Chrome Profile:
   • Mở chrome://version trong Chrome
   • Tìm "Profile Path" và copy đường dẫn
   • Dán vào accounts.json

2️⃣ Thêm Groq API Key (FREE):
   • Vào console.groq.com/keys
   • Tạo API key mới
   • Copy và dán vào accounts.json

3️⃣ Chạy Tool:
   • Chọn file voice hoặc thư mục
   • Nhấn BẮT ĐẦU
   • Tool tự động: Voice → SRT → Prompts → Images

⚠️ LƯU Ý:
   • Đóng Chrome trước khi chạy
   • Không di chuột khi đang lấy token
   • Profiles phải đã đăng nhập Google
"""
        
        help_label = ttk.Label(help_tab, text=help_text, justify=tk.LEFT, font=('Segoe UI', 10))
        help_label.pack(anchor=tk.W)
        
        # Close button
        ttk.Button(win, text="Đóng", command=win.destroy).pack(pady=10)
    
    def open_config_file(self):
        """Open config file in editor."""
        config_file = CONFIG_DIR / "accounts.json"
        
        if not config_file.exists():
            self.create_default_config()
        
        if sys.platform == "win32":
            os.startfile(str(config_file))
        else:
            import subprocess
            subprocess.Popen(["xdg-open", str(config_file)])
    
    def open_output_folder(self):
        """Open output folder."""
        PROJECTS_DIR.mkdir(exist_ok=True)
        
        if sys.platform == "win32":
            os.startfile(str(PROJECTS_DIR))
        else:
            import subprocess
            subprocess.Popen(["xdg-open", str(PROJECTS_DIR)])
    
    # ========== MAIN PROCESSING ==========
    
    def start_processing(self):
        """Start main processing."""
        path = self.input_path.get()
        
        if not path:
            messagebox.showerror("Lỗi", "Vui lòng chọn file hoặc thư mục đầu vào!")
            return
        
        if not Path(path).exists():
            messagebox.showerror("Lỗi", f"Không tìm thấy:\n{path}")
            return
        
        # Reload config
        self.load_config()
        
        if not self.profiles:
            result = messagebox.askyesno(
                "Thiếu Chrome Profile",
                "Chưa có Chrome profile nào!\n\n"
                "Bạn cần thêm profile vào file config.\n\n"
                "Mở file config ngay?"
            )
            if result:
                self.open_config_file()
            return
        
        # Check AI keys for voice
        ext = Path(path).suffix.lower() if Path(path).is_file() else ""
        if ext in ['.mp3', '.wav'] and not self.groq_keys and not self.gemini_keys:
            result = messagebox.askyesno(
                "Thiếu AI API Key",
                "Cần Groq hoặc Gemini API key để xử lý voice!\n\n"
                "Groq API hoàn toàn FREE.\n"
                "Mở trang Groq để lấy key?"
            )
            if result:
                webbrowser.open("https://console.groq.com/keys")
            return
        
        # Start
        self._running = True
        self._stop = False
        self.start_btn.config(state=tk.DISABLED, bg='#95a5a6')
        self.stop_btn.config(state=tk.NORMAL)
        
        self.clear_log()
        self.log("=" * 50)
        self.log("🚀 BẮT ĐẦU XỬ LÝ")
        self.log("=" * 50)
        
        if self.input_mode.get() == "folder":
            threading.Thread(target=self._process_folder, daemon=True).start()
        else:
            threading.Thread(target=self._process_single, daemon=True).start()
    
    def stop_processing(self):
        """Stop processing."""
        self._stop = True
        if self._engine:
            self._engine.stop()
        self.log("⏹️ Đang dừng...", "WARN")
    
    def _process_single(self):
        """Process single file in background thread."""
        try:
            from modules.smart_engine import SmartEngine
            
            path = self.input_path.get()
            
            engine = SmartEngine()
            self._engine = engine
            
            def log_cb(msg):
                # Parse level from message
                level = "INFO"
                if "[OK]" in msg or "OK!" in msg:
                    level = "OK"
                elif "[ERROR]" in msg or "ERROR" in msg:
                    level = "ERROR"
                elif "[WARN]" in msg:
                    level = "WARN"
                
                self.root.after(0, lambda: self.log(msg, level))
            
            results = engine.run(path, callback=log_cb)
            
            # Update UI
            self.root.after(0, lambda: self._on_complete(results))
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda: self.log(f"Lỗi: {e}", "ERROR"))
            self.root.after(0, lambda: messagebox.showerror("Lỗi", str(e)))
        finally:
            self._running = False
            self.root.after(0, self._reset_ui)
    
    def _process_folder(self):
        """Process folder with multiple voice files."""
        try:
            from modules.smart_engine import SmartEngine
            
            folder = Path(self.input_path.get())
            voices = list(folder.glob("*.mp3")) + list(folder.glob("*.wav"))
            
            if not voices:
                self.root.after(0, lambda: messagebox.showerror("Lỗi", "Không tìm thấy file voice nào!"))
                return
            
            self.log(f"📁 Tìm thấy {len(voices)} file voice")
            
            total_success = 0
            total_failed = 0
            
            for i, voice in enumerate(voices):
                if self._stop:
                    break
                
                self.root.after(0, lambda v=voice.name, i=i, t=len(voices): 
                    self.log(f"\n[{i+1}/{t}] 📄 {v}"))
                
                self.root.after(0, lambda p=(i/len(voices))*100, t=f"Xử lý {i+1}/{len(voices)}": 
                    self.update_progress(p, t, voice.name))
                
                engine = SmartEngine()
                self._engine = engine
                
                def log_cb(msg):
                    self.root.after(0, lambda: self.log(f"    {msg}"))
                
                results = engine.run(str(voice), callback=log_cb)
                
                if 'error' not in results:
                    total_success += results.get('success', 0)
                    total_failed += results.get('failed', 0)
            
            # Summary
            self.root.after(0, lambda: self.update_progress(100, "Hoàn tất!"))
            self.root.after(0, lambda: self.log(f"\n📊 TỔNG KẾT: {total_success} ✅ | {total_failed} ❌", "OK"))
            
            if total_failed > 0:
                self.root.after(0, lambda: messagebox.showwarning(
                    "Chưa hoàn thành",
                    f"✅ Thành công: {total_success}\n❌ Thất bại: {total_failed}"
                ))
            else:
                self.root.after(0, lambda: messagebox.showinfo(
                    "Hoàn tất!",
                    f"✅ Đã tạo {total_success} ảnh!"
                ))
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda: self.log(f"Lỗi: {e}", "ERROR"))
        finally:
            self._running = False
            self.root.after(0, self._reset_ui)
    
    def _on_complete(self, results):
        """Handle completion."""
        # Set current project for preview
        path = self.input_path.get()
        if path:
            name = Path(path).stem
            self.current_project_dir = PROJECTS_DIR / name
        
        if 'error' in results:
            err = results['error']
            if err == 'missing_requirements':
                missing = results.get('missing', [])
                messagebox.showerror("Thiếu yêu cầu",
                    "Cần bổ sung:\n\n" + "\n".join(f"• {m}" for m in missing))
            else:
                messagebox.showerror("Lỗi", str(err))
        else:
            success = results.get('success', 0)
            failed = results.get('failed', 0)
            
            self.update_progress(100, "Hoàn tất!")
            
            # Auto refresh preview
            self.refresh_preview()
            
            if failed > 0:
                messagebox.showwarning("Chưa hoàn thành",
                    f"✅ Thành công: {success}\n❌ Thất bại: {failed}\n\nXem log để biết chi tiết.")
            else:
                messagebox.showinfo("Hoàn tất!", f"✅ Đã tạo {success} ảnh!")
    
    def _reset_ui(self):
        """Reset UI after processing."""
        self.start_btn.config(state=tk.NORMAL, bg='#2ecc71')
        self.stop_btn.config(state=tk.DISABLED)
    
    # ========== PREVIEW ==========
    
    def on_char_selected(self, event=None):
        """Handle character selection."""
        sel = self.char_combo.get()
        if not sel or not self.current_project_dir:
            return
        
        # Load character image
        img_path = self.current_project_dir / "nv" / f"{sel}.png"
        self.load_image_to_label(img_path, self.char_image_label, (200, 200))
        
        # Load prompt from Excel
        self.char_prompt_text.delete(1.0, tk.END)
        prompt = self.get_prompt_for_id(sel)
        if prompt:
            self.char_prompt_text.insert(tk.END, prompt)
    
    def on_scene_selected(self, event=None):
        """Handle scene selection."""
        sel = self.scene_combo.get()
        if not sel or not self.current_project_dir:
            return
        
        # Load scene image
        img_path = self.current_project_dir / "img" / f"{sel}.png"
        self.load_image_to_label(img_path, self.result_image_label, (300, 200))
        
        # Load prompt
        self.result_prompt_text.delete(1.0, tk.END)
        prompt = self.get_prompt_for_id(sel)
        if prompt:
            self.result_prompt_text.insert(tk.END, prompt)
    
    def load_image_to_label(self, img_path: Path, label: ttk.Label, size: tuple):
        """Load image and display on label."""
        if not HAS_PIL:
            label.config(text="Cần cài PIL:\npip install Pillow")
            return
        
        if not img_path.exists():
            label.config(text=f"Chưa có ảnh\n{img_path.name}")
            return
        
        try:
            img = Image.open(img_path)
            img.thumbnail(size, Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            
            # Keep reference
            label._photo = photo
            label.config(image=photo, text="")
        except Exception as e:
            label.config(text=f"Lỗi: {e}")
    
    def get_prompt_for_id(self, pid: str) -> str:
        """Get prompt for an ID from Excel."""
        if not self.current_project_dir:
            return ""
        
        # Find Excel file
        prompts_dir = self.current_project_dir / "prompts"
        excel_files = list(prompts_dir.glob("*_prompts.xlsx"))
        
        if not excel_files:
            return ""
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_files[0], read_only=True)
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [c.value for c in ws[1]]
                
                # Find columns
                id_col = prompt_col = None
                for i, h in enumerate(headers or []):
                    if h is None:
                        continue
                    h_lower = str(h).lower()
                    if 'id' in h_lower and id_col is None:
                        id_col = i
                    if 'english' in h_lower and 'prompt' in h_lower:
                        prompt_col = i
                    elif h_lower == 'img_prompt' and prompt_col is None:
                        prompt_col = i
                    elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower and 'viet' not in h_lower:
                        prompt_col = i
                
                if id_col is None or prompt_col is None:
                    continue
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > max(id_col, prompt_col):
                        if str(row[id_col]).strip() == pid:
                            return str(row[prompt_col] or "")
            
            wb.close()
        except:
            pass
        
        return ""
    
    def refresh_preview(self):
        """Refresh preview from current project."""
        # Find project dir from input
        path = self.input_path.get()
        if path:
            name = Path(path).stem
            self.current_project_dir = PROJECTS_DIR / name
        
        if not self.current_project_dir or not self.current_project_dir.exists():
            self.log("Chưa có project để preview")
            return
        
        self.log(f"Loading preview from: {self.current_project_dir}")
        
        # Load characters
        nv_dir = self.current_project_dir / "nv"
        chars = []
        if nv_dir.exists():
            chars = sorted([c.stem for c in nv_dir.glob("*.png")])
        
        self.char_combo['values'] = chars
        if chars:
            self.char_combo.set(chars[0])
            self.on_char_selected()
        
        # Load scenes
        img_dir = self.current_project_dir / "img"
        scenes = []
        if img_dir.exists():
            scenes = sorted([s.stem for s in img_dir.glob("*.png")])
        
        self.scene_combo['values'] = scenes
        if scenes:
            self.scene_combo.set(scenes[0])
            self.on_scene_selected()
        
        # Update thumbnails
        self.update_thumbnails(scenes)
        
        # Update prompts tab
        self.update_prompts_tab()
        
        self.log(f"Preview: {len(chars)} nhân vật, {len(scenes)} scenes", "OK")
    
    def update_thumbnails(self, scene_ids: List[str]):
        """Update scene thumbnails."""
        self.thumb_canvas.delete("all")
        
        if not HAS_PIL or not self.current_project_dir:
            return
        
        img_dir = self.current_project_dir / "img"
        x = 5
        
        self._thumb_photos = []  # Keep references
        
        for sid in scene_ids[:20]:  # Max 20 thumbnails
            img_path = img_dir / f"{sid}.png"
            if not img_path.exists():
                continue
            
            try:
                img = Image.open(img_path)
                img.thumbnail((80, 80), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                self._thumb_photos.append(photo)
                
                self.thumb_canvas.create_image(x, 5, anchor=tk.NW, image=photo)
                self.thumb_canvas.create_text(x + 40, 85, text=sid, font=('Segoe UI', 7))
                x += 90
            except:
                pass
        
        self.thumb_canvas.configure(scrollregion=(0, 0, x, 100))
    
    def update_prompts_tab(self):
        """Update prompts treeviews."""
        # Clear existing
        for item in self.char_tree.get_children():
            self.char_tree.delete(item)
        for item in self.scene_tree.get_children():
            self.scene_tree.delete(item)
        
        if not self.current_project_dir:
            return
        
        # Find Excel
        prompts_dir = self.current_project_dir / "prompts"
        excel_files = list(prompts_dir.glob("*_prompts.xlsx"))
        
        if not excel_files:
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_files[0], read_only=True)
            
            nv_dir = self.current_project_dir / "nv"
            img_dir = self.current_project_dir / "img"
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [c.value for c in ws[1]]
                
                # Find columns
                id_col = prompt_col = time_col = None
                for i, h in enumerate(headers or []):
                    if h is None:
                        continue
                    h_lower = str(h).lower()
                    if 'id' in h_lower and id_col is None:
                        id_col = i
                    if 'english' in h_lower and 'prompt' in h_lower:
                        prompt_col = i
                    elif h_lower == 'img_prompt' and prompt_col is None:
                        prompt_col = i
                    elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower and 'viet' not in h_lower:
                        prompt_col = i
                    if 'time' in h_lower or 'start' in h_lower:
                        time_col = i
                
                if id_col is None or prompt_col is None:
                    continue
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= max(id_col, prompt_col):
                        continue
                    
                    pid = str(row[id_col] or "").strip()
                    prompt = str(row[prompt_col] or "")[:60] + "..."
                    time_str = str(row[time_col] or "") if time_col else ""
                    
                    if not pid:
                        continue
                    
                    # Check status
                    if pid.startswith('nv'):
                        img_path = nv_dir / f"{pid}.png"
                        status = "✅" if img_path.exists() else "⏳"
                        self.char_tree.insert('', tk.END, values=(pid, prompt, status))
                    else:
                        img_path = img_dir / f"{pid}.png"
                        status = "✅" if img_path.exists() else "⏳"
                        self.scene_tree.insert('', tk.END, values=(pid, time_str, prompt, status))
            
            wb.close()
        except Exception as e:
            self.log(f"Error loading prompts: {e}", "ERROR")
    
    def get_token_manual(self):
        """Get token manually."""
        if not self.profiles:
            messagebox.showerror("Lỗi", "Chưa có Chrome profile!\n\nThêm vào config/accounts.json")
            return
        
        self.log("🔑 Đang lấy token thủ công...")
        
        def worker():
            try:
                from modules.auto_token import ChromeAutoToken
                
                extractor = ChromeAutoToken(
                    chrome_path=self.chrome_path,
                    profile_path=self.profiles[0]
                )
                
                def log_cb(msg):
                    self.root.after(0, lambda: self.log(msg))
                
                token, proj_id, error = extractor.extract_token(callback=log_cb)
                
                if token:
                    self.root.after(0, lambda: self.log(f"✅ Token: {token[:40]}...", "OK"))
                    self.root.after(0, lambda: messagebox.showinfo("OK", "Đã lấy được token!"))
                else:
                    self.root.after(0, lambda: self.log(f"❌ {error}", "ERROR"))
                    
            except Exception as e:
                self.root.after(0, lambda: self.log(f"Lỗi: {e}", "ERROR"))
        
        threading.Thread(target=worker, daemon=True).start()
    
    # ========== RUN ==========
    
    def run(self):
        """Start application."""
        self.root.mainloop()


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Entry point."""
    app = VE3ToolPro()
    app.run()


if __name__ == "__main__":
    main()
