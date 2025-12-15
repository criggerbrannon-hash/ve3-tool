"""
VE3 Tool - Image to Video Converter
====================================
Chuyển đổi ảnh thành video với các tính năng:
- Kết hợp nhiều ảnh theo thứ tự từ SRT/Excel
- Thêm audio từ file voice gốc
- Thêm phụ đề từ file SRT
- Hỗ trợ hiệu ứng chuyển cảnh (fade, crossfade)
- Cấu hình độ phân giải, FPS, codec

Yêu cầu:
- FFmpeg cài đặt trên hệ thống
- MoviePy (pip install moviepy)
"""

import os
import subprocess
import shutil
from pathlib import Path
from typing import List, Dict, Optional, Callable, Tuple, Any
from dataclasses import dataclass
from datetime import timedelta

from modules.utils import get_logger, parse_srt_file, SrtEntry


# ============================================================================
# AVAILABILITY CHECK
# ============================================================================

MOVIEPY_AVAILABLE = False
FFMPEG_AVAILABLE = False

try:
    from moviepy.editor import (
        ImageClip, AudioFileClip, CompositeVideoClip,
        concatenate_videoclips, TextClip
    )
    MOVIEPY_AVAILABLE = True
except ImportError:
    pass

# Check FFmpeg
try:
    result = subprocess.run(
        ['ffmpeg', '-version'],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=True
    )
    FFMPEG_AVAILABLE = True
except (subprocess.CalledProcessError, FileNotFoundError):
    pass


class FFmpegNotFoundError(Exception):
    """Exception khi FFmpeg không được cài đặt."""

    def __init__(self):
        message = """
FFmpeg không được cài đặt. Vui lòng cài đặt FFmpeg:

- Windows:
    1. Download từ https://ffmpeg.org/download.html
    2. Giải nén và thêm vào PATH
    HOẶC: choco install ffmpeg

- macOS: brew install ffmpeg

- Linux: sudo apt install ffmpeg
        """
        super().__init__(message)


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class VideoSettings:
    """Cấu hình cho video output."""
    width: int = 1920
    height: int = 1080
    fps: int = 30
    codec: str = "libx264"  # libx264, libx265, mpeg4
    audio_codec: str = "aac"
    bitrate: str = "5M"
    audio_bitrate: str = "192k"
    format: str = "mp4"

    # Timing
    default_image_duration: float = 5.0  # Seconds per image if no SRT

    # Effects
    fade_duration: float = 0.5  # Fade in/out duration
    crossfade: bool = True  # Use crossfade between images

    # Subtitles
    subtitle_font: str = "Arial"
    subtitle_fontsize: int = 40
    subtitle_color: str = "white"
    subtitle_bg_color: str = "black"
    subtitle_position: str = "bottom"  # bottom, center, top


@dataclass
class SceneInfo:
    """Thông tin về một scene trong video."""
    scene_id: str
    image_path: Path
    start_time: float  # seconds
    end_time: float  # seconds
    duration: float  # seconds
    text: str = ""  # Subtitle text

    @property
    def exists(self) -> bool:
        return self.image_path.exists()


# ============================================================================
# IMAGE TO VIDEO CONVERTER
# ============================================================================

class ImageToVideo:
    """
    Class chuyển đổi ảnh thành video.

    Hỗ trợ 2 backends:
    1. FFmpeg (nhanh, ổn định)
    2. MoviePy (nhiều tùy chỉnh hơn)
    """

    def __init__(
        self,
        settings: Optional[VideoSettings] = None,
        use_moviepy: bool = False
    ):
        """
        Khởi tạo ImageToVideo converter.

        Args:
            settings: Cấu hình video (dùng default nếu None)
            use_moviepy: True để dùng MoviePy, False để dùng FFmpeg
        """
        self.settings = settings or VideoSettings()
        self.use_moviepy = use_moviepy and MOVIEPY_AVAILABLE
        self.logger = get_logger("image_to_video")

        # Validate
        if not FFMPEG_AVAILABLE:
            raise FFmpegNotFoundError()

        if use_moviepy and not MOVIEPY_AVAILABLE:
            self.logger.warning("MoviePy không có sẵn, sẽ dùng FFmpeg")
            self.use_moviepy = False

    def convert(
        self,
        project_dir: Path,
        output_path: Optional[Path] = None,
        audio_path: Optional[Path] = None,
        srt_path: Optional[Path] = None,
        callback: Optional[Callable[[str], None]] = None
    ) -> Tuple[bool, Path, str]:
        """
        Chuyển đổi ảnh trong project thành video.

        Args:
            project_dir: Thư mục project chứa img/, srt/, ...
            output_path: Path output video (None = auto)
            audio_path: Path file audio (None = tự tìm)
            srt_path: Path file SRT (None = tự tìm)
            callback: Callback để report progress

        Returns:
            Tuple (success, output_path, error_message)
        """
        project_dir = Path(project_dir)

        def log(msg: str):
            self.logger.info(msg)
            if callback:
                callback(msg)

        log(f"Bắt đầu tạo video từ: {project_dir}")

        # 1. Tìm các file cần thiết
        img_dir = project_dir / "img"
        if not img_dir.exists():
            return False, None, f"Thư mục img không tồn tại: {img_dir}"

        # Tìm audio
        if audio_path is None:
            audio_path = self._find_audio(project_dir)

        # Tìm SRT
        if srt_path is None:
            srt_path = self._find_srt(project_dir)

        # Output path
        if output_path is None:
            vid_dir = project_dir / "vid"
            vid_dir.mkdir(exist_ok=True)
            output_path = vid_dir / f"{project_dir.name}.{self.settings.format}"

        # 2. Load scenes
        scenes = self._load_scenes(project_dir, img_dir, srt_path)

        if not scenes:
            return False, None, "Không tìm thấy ảnh nào để tạo video"

        # Filter only existing images
        existing_scenes = [s for s in scenes if s.exists]
        log(f"Tìm thấy {len(existing_scenes)}/{len(scenes)} ảnh")

        if not existing_scenes:
            return False, None, "Không có ảnh nào tồn tại"

        # 3. Tạo video
        try:
            if self.use_moviepy:
                success = self._create_video_moviepy(
                    existing_scenes, output_path, audio_path, log
                )
            else:
                success = self._create_video_ffmpeg(
                    existing_scenes, output_path, audio_path, srt_path, log
                )

            if success:
                log(f"Video đã được tạo: {output_path}")
                return True, output_path, ""
            else:
                return False, None, "Tạo video thất bại"

        except Exception as e:
            self.logger.error(f"Lỗi tạo video: {e}")
            return False, None, str(e)

    def _find_audio(self, project_dir: Path) -> Optional[Path]:
        """Tìm file audio trong project."""
        for ext in ['.mp3', '.wav', '.m4a', '.ogg', '.aac']:
            # Tìm trong root
            for f in project_dir.glob(f"*{ext}"):
                return f
            # Tìm trong thư mục con
            for f in project_dir.glob(f"**/*{ext}"):
                return f
        return None

    def _find_srt(self, project_dir: Path) -> Optional[Path]:
        """Tìm file SRT trong project."""
        srt_dir = project_dir / "srt"
        if srt_dir.exists():
            for f in srt_dir.glob("*.srt"):
                return f
        # Tìm trong root
        for f in project_dir.glob("*.srt"):
            return f
        return None

    def _load_scenes(
        self,
        project_dir: Path,
        img_dir: Path,
        srt_path: Optional[Path]
    ) -> List[SceneInfo]:
        """
        Load thông tin scenes từ SRT hoặc Excel.

        Args:
            project_dir: Thư mục project
            img_dir: Thư mục chứa ảnh
            srt_path: Path đến file SRT

        Returns:
            List các SceneInfo
        """
        scenes = []

        # 1. Thử load từ Excel prompts
        excel_scenes = self._load_scenes_from_excel(project_dir, img_dir)
        if excel_scenes:
            self.logger.info(f"Loaded {len(excel_scenes)} scenes từ Excel")
            return excel_scenes

        # 2. Load từ SRT nếu có
        if srt_path and srt_path.exists():
            srt_scenes = self._load_scenes_from_srt(img_dir, srt_path)
            if srt_scenes:
                self.logger.info(f"Loaded {len(srt_scenes)} scenes từ SRT")
                return srt_scenes

        # 3. Fallback: list tất cả ảnh với timing đều
        images = sorted(img_dir.glob("*.png")) + sorted(img_dir.glob("*.jpg"))
        duration = self.settings.default_image_duration

        for i, img_path in enumerate(images):
            scenes.append(SceneInfo(
                scene_id=img_path.stem,
                image_path=img_path,
                start_time=i * duration,
                end_time=(i + 1) * duration,
                duration=duration
            ))

        self.logger.info(f"Loaded {len(scenes)} scenes (fallback mode)")
        return scenes

    def _load_scenes_from_excel(
        self,
        project_dir: Path,
        img_dir: Path
    ) -> List[SceneInfo]:
        """Load scenes từ Excel prompts file."""
        prompts_dir = project_dir / "prompts"
        if not prompts_dir.exists():
            return []

        excel_files = list(prompts_dir.glob("*_prompts.xlsx"))
        if not excel_files:
            return []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_files[0], read_only=True)
        except Exception as e:
            self.logger.warning(f"Không thể đọc Excel: {e}")
            return []

        scenes = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1]]

            if not headers:
                continue

            # Tìm các cột
            id_col = time_start_col = time_end_col = text_col = None

            for i, h in enumerate(headers):
                if h is None:
                    continue
                h_lower = str(h).lower()

                if 'id' in h_lower and id_col is None:
                    id_col = i
                if ('start' in h_lower or h_lower == 'time') and time_start_col is None:
                    time_start_col = i
                if 'end' in h_lower and time_end_col is None:
                    time_end_col = i
                if h_lower in ['text', 'content', 'noi_dung', 'subtitle'] and text_col is None:
                    text_col = i

            if id_col is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= id_col:
                    continue

                scene_id = str(row[id_col] or "").strip()
                if not scene_id or scene_id.startswith('nv'):
                    continue  # Skip character images

                img_path = img_dir / f"{scene_id}.png"
                if not img_path.exists():
                    img_path = img_dir / f"{scene_id}.jpg"

                # Parse timing
                start_time = 0.0
                end_time = self.settings.default_image_duration

                if time_start_col is not None and len(row) > time_start_col:
                    start_time = self._parse_time(row[time_start_col])

                if time_end_col is not None and len(row) > time_end_col:
                    end_time = self._parse_time(row[time_end_col])
                elif start_time > 0:
                    # Estimate end time
                    end_time = start_time + self.settings.default_image_duration

                text = ""
                if text_col is not None and len(row) > text_col:
                    text = str(row[text_col] or "")

                scenes.append(SceneInfo(
                    scene_id=scene_id,
                    image_path=img_path,
                    start_time=start_time,
                    end_time=end_time,
                    duration=end_time - start_time,
                    text=text
                ))

        wb.close()

        # Sort by start time
        scenes.sort(key=lambda s: (s.start_time, s.scene_id))

        return scenes

    def _load_scenes_from_srt(
        self,
        img_dir: Path,
        srt_path: Path
    ) -> List[SceneInfo]:
        """Load scenes từ SRT file."""
        try:
            entries = parse_srt_file(srt_path)
        except Exception as e:
            self.logger.warning(f"Không thể parse SRT: {e}")
            return []

        if not entries:
            return []

        scenes = []
        images = sorted(img_dir.glob("*.png")) + sorted(img_dir.glob("*.jpg"))

        # Map images theo index hoặc scene_id
        image_map = {}
        for img in images:
            stem = img.stem
            # Try parse as number
            try:
                idx = int(stem.replace('scene_', '').replace('s', ''))
                image_map[idx] = img
            except ValueError:
                image_map[stem] = img

        # Group entries thành scenes (mỗi scene ~15-25s)
        current_scene_entries = []
        scene_start = entries[0].start_time
        scene_num = 1

        for entry in entries:
            current_duration = (entry.end_time - scene_start).total_seconds()

            if current_duration > 20 and current_scene_entries:
                # Tạo scene mới
                last_entry = current_scene_entries[-1]
                text = " ".join(e.text for e in current_scene_entries)

                img_path = image_map.get(scene_num)
                if img_path is None:
                    img_path = image_map.get(f"scene_{scene_num}")
                if img_path is None:
                    img_path = image_map.get(f"s{scene_num}")
                if img_path is None and images:
                    idx = min(scene_num - 1, len(images) - 1)
                    img_path = images[idx]

                if img_path:
                    scenes.append(SceneInfo(
                        scene_id=f"s{scene_num}",
                        image_path=img_path,
                        start_time=scene_start.total_seconds(),
                        end_time=last_entry.end_time.total_seconds(),
                        duration=current_duration,
                        text=text
                    ))

                scene_num += 1
                scene_start = entry.start_time
                current_scene_entries = []

            current_scene_entries.append(entry)

        # Scene cuối cùng
        if current_scene_entries:
            last_entry = current_scene_entries[-1]
            text = " ".join(e.text for e in current_scene_entries)

            img_path = image_map.get(scene_num)
            if img_path is None and images:
                idx = min(scene_num - 1, len(images) - 1)
                img_path = images[idx]

            if img_path:
                scenes.append(SceneInfo(
                    scene_id=f"s{scene_num}",
                    image_path=img_path,
                    start_time=scene_start.total_seconds(),
                    end_time=last_entry.end_time.total_seconds(),
                    duration=(last_entry.end_time - scene_start).total_seconds(),
                    text=text
                ))

        return scenes

    def _parse_time(self, value: Any) -> float:
        """Parse time value từ nhiều định dạng."""
        if value is None:
            return 0.0

        if isinstance(value, (int, float)):
            return float(value)

        # Try parse string
        s = str(value).strip()

        # Format: MM:SS hoặc HH:MM:SS hoặc HH:MM:SS,mmm
        if ':' in s:
            s = s.replace(',', '.')
            parts = s.split(':')

            if len(parts) == 2:
                return int(parts[0]) * 60 + float(parts[1])
            elif len(parts) == 3:
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])

        try:
            return float(s)
        except ValueError:
            return 0.0

    # ========== FFMPEG BACKEND ==========

    def _create_video_ffmpeg(
        self,
        scenes: List[SceneInfo],
        output_path: Path,
        audio_path: Optional[Path],
        srt_path: Optional[Path],
        log: Callable[[str], None]
    ) -> bool:
        """Tạo video sử dụng FFmpeg."""

        # 1. Tạo file concat list
        temp_dir = output_path.parent / ".temp_video"
        temp_dir.mkdir(exist_ok=True)

        try:
            # Create concat file
            concat_file = temp_dir / "concat.txt"

            with open(concat_file, 'w', encoding='utf-8') as f:
                for scene in scenes:
                    # Escape single quotes in path
                    img_path = str(scene.image_path).replace("'", "'\\''")
                    f.write(f"file '{img_path}'\n")
                    f.write(f"duration {scene.duration}\n")

                # Duplicate last frame (FFmpeg requirement)
                if scenes:
                    last_path = str(scenes[-1].image_path).replace("'", "'\\''")
                    f.write(f"file '{last_path}'\n")

            log(f"Tạo video từ {len(scenes)} ảnh...")

            # 2. Build FFmpeg command
            cmd = [
                'ffmpeg', '-y',  # Overwrite output
                '-f', 'concat',
                '-safe', '0',
                '-i', str(concat_file),
            ]

            # Add audio if available
            if audio_path and audio_path.exists():
                cmd.extend(['-i', str(audio_path)])
                log(f"Thêm audio: {audio_path.name}")

            # Video settings
            cmd.extend([
                '-vf', f'scale={self.settings.width}:{self.settings.height}:force_original_aspect_ratio=decrease,pad={self.settings.width}:{self.settings.height}:(ow-iw)/2:(oh-ih)/2:black',
                '-c:v', self.settings.codec,
                '-preset', 'medium',
                '-b:v', self.settings.bitrate,
                '-r', str(self.settings.fps),
                '-pix_fmt', 'yuv420p',
            ])

            # Audio settings
            if audio_path and audio_path.exists():
                cmd.extend([
                    '-c:a', self.settings.audio_codec,
                    '-b:a', self.settings.audio_bitrate,
                    '-shortest',  # Match video length to shortest stream
                ])

            cmd.append(str(output_path))

            log("Đang render video...")

            # 3. Run FFmpeg
            result = subprocess.run(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )

            if result.returncode != 0:
                self.logger.error(f"FFmpeg error: {result.stderr}")
                return False

            # 4. Add subtitles if available
            if srt_path and srt_path.exists():
                log("Thêm phụ đề...")
                self._add_subtitles_ffmpeg(output_path, srt_path)

            return output_path.exists()

        finally:
            # Cleanup
            if temp_dir.exists():
                shutil.rmtree(temp_dir, ignore_errors=True)

    def _add_subtitles_ffmpeg(self, video_path: Path, srt_path: Path) -> bool:
        """Thêm phụ đề vào video sử dụng FFmpeg."""
        temp_output = video_path.with_suffix('.temp.mp4')

        # Escape path for subtitles filter
        srt_escaped = str(srt_path).replace('\\', '/').replace(':', '\\:')

        cmd = [
            'ffmpeg', '-y',
            '-i', str(video_path),
            '-vf', f"subtitles='{srt_escaped}':force_style='FontSize={self.settings.subtitle_fontsize},PrimaryColour=&Hffffff&,OutlineColour=&H000000&,Outline=2'",
            '-c:a', 'copy',
            str(temp_output)
        ]

        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        if result.returncode == 0 and temp_output.exists():
            video_path.unlink()
            temp_output.rename(video_path)
            return True

        # Cleanup on failure
        if temp_output.exists():
            temp_output.unlink()

        return False

    # ========== MOVIEPY BACKEND ==========

    def _create_video_moviepy(
        self,
        scenes: List[SceneInfo],
        output_path: Path,
        audio_path: Optional[Path],
        log: Callable[[str], None]
    ) -> bool:
        """Tạo video sử dụng MoviePy."""
        if not MOVIEPY_AVAILABLE:
            return False

        from moviepy.editor import (
            ImageClip, AudioFileClip, CompositeVideoClip,
            concatenate_videoclips
        )

        clips = []

        try:
            log(f"Tạo video từ {len(scenes)} ảnh (MoviePy)...")

            for i, scene in enumerate(scenes):
                # Create image clip
                clip = ImageClip(str(scene.image_path), duration=scene.duration)
                clip = clip.resize((self.settings.width, self.settings.height))

                # Add fade effects
                if self.settings.fade_duration > 0:
                    clip = clip.fadein(self.settings.fade_duration)
                    clip = clip.fadeout(self.settings.fade_duration)

                clips.append(clip)

                if (i + 1) % 10 == 0:
                    log(f"Đã xử lý {i + 1}/{len(scenes)} ảnh")

            # Concatenate
            if self.settings.crossfade:
                # Crossfade between clips
                final_clip = clips[0]
                for clip in clips[1:]:
                    final_clip = concatenate_videoclips(
                        [final_clip, clip],
                        method="crossfade",
                        crossfade_duration=self.settings.fade_duration
                    )
            else:
                final_clip = concatenate_videoclips(clips, method="compose")

            # Add audio
            if audio_path and audio_path.exists():
                log(f"Thêm audio: {audio_path.name}")
                audio = AudioFileClip(str(audio_path))

                # Match video duration to audio
                if audio.duration < final_clip.duration:
                    final_clip = final_clip.subclip(0, audio.duration)

                final_clip = final_clip.set_audio(audio)

            # Write
            log("Đang render video...")

            final_clip.write_videofile(
                str(output_path),
                fps=self.settings.fps,
                codec=self.settings.codec,
                audio_codec=self.settings.audio_codec,
                bitrate=self.settings.bitrate,
                logger=None  # Suppress moviepy logs
            )

            return output_path.exists()

        except Exception as e:
            self.logger.error(f"MoviePy error: {e}")
            return False

        finally:
            # Cleanup clips
            for clip in clips:
                try:
                    clip.close()
                except:
                    pass


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def convert_images_to_video(
    project_dir: Path,
    output_path: Optional[Path] = None,
    settings: Optional[VideoSettings] = None,
    callback: Optional[Callable[[str], None]] = None
) -> Tuple[bool, Path, str]:
    """
    Hàm tiện ích để chuyển đổi ảnh thành video.

    Args:
        project_dir: Thư mục project chứa img/, srt/...
        output_path: Path output video (None = auto)
        settings: Cấu hình video
        callback: Callback để report progress

    Returns:
        Tuple (success, output_path, error_message)
    """
    converter = ImageToVideo(settings=settings)
    return converter.convert(project_dir, output_path, callback=callback)


def check_video_requirements() -> Dict[str, bool]:
    """
    Kiểm tra các yêu cầu để tạo video.

    Returns:
        Dict với các key: ffmpeg, moviepy
    """
    return {
        "ffmpeg": FFMPEG_AVAILABLE,
        "moviepy": MOVIEPY_AVAILABLE
    }


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        project_path = Path(sys.argv[1])

        def log_callback(msg):
            print(msg)

        success, output, error = convert_images_to_video(
            project_path,
            callback=log_callback
        )

        if success:
            print(f"\n✅ Video created: {output}")
        else:
            print(f"\n❌ Error: {error}")
    else:
        print("Usage: python image_to_video.py <project_dir>")
        print("\nRequirements:")
        reqs = check_video_requirements()
        for name, available in reqs.items():
            status = "✅" if available else "❌"
            print(f"  {status} {name}")
