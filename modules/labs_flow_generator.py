"""
VE3 Tool - Labs Flow Generator Module
=====================================
Tich hop Labs Google API (session token + CAPTCHA) voi Excel workflow.

Workflow:
1. Doc Excel prompts (characters + scenes sheets)
2. Giai CAPTCHA bang Capsolver/2Captcha
3. Goi API labs.google voi session token
4. Luu anh va cap nhat Excel

Uu diem so voi cac mode khac:
- Khong can browser automation (nhanh hon)
- Khong can bearer token (de setup hon)
- Chi can cookie tu Cookie Editor
"""

import os
import time
import yaml
import base64
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

from .labs_google_api import LabsGoogleAPI


class LabsFlowGenerator:
    """
    Generator anh su dung Labs Google API (session token + CAPTCHA).
    Doc prompts tu Excel va tao anh tu dong.
    """

    def __init__(
        self,
        project_path: Path,
        session_token: str,
        captcha_api_key: str = None,
        captcha_service: str = "capsolver",
        aspect_ratio: str = "landscape",
        delay_between_requests: float = 2.0,
        verbose: bool = True
    ):
        """
        Khoi tao Labs Flow Generator.

        Args:
            project_path: Duong dan den thu muc project (PROJECTS/{CODE}/)
            session_token: __Secure-next-auth.session-token tu Cookie Editor
            captcha_api_key: API key cua dich vu CAPTCHA solver
            captcha_service: Ten dich vu ("capsolver", "2captcha")
            aspect_ratio: Ty le khung hinh (landscape/portrait/square)
            delay_between_requests: Thoi gian cho giua cac request (giay)
            verbose: In log chi tiet
        """
        self.project_path = Path(project_path)
        self.delay = delay_between_requests
        self.verbose = verbose
        self.aspect_ratio = aspect_ratio.lower()

        # Tao Labs API client
        self.labs_client = LabsGoogleAPI(
            session_token=session_token,
            captcha_api_key=captcha_api_key,
            captcha_service=captcha_service,
            verbose=verbose
        )

        # Paths
        self.nv_path = self.project_path / "nv"
        self.img_path = self.project_path / "img"
        self.prompts_path = self.project_path / "prompts"

        # Tao thu muc neu chua co
        self.nv_path.mkdir(parents=True, exist_ok=True)
        self.img_path.mkdir(parents=True, exist_ok=True)

        # Stats
        self.stats = {
            "characters_total": 0,
            "characters_success": 0,
            "characters_failed": 0,
            "scenes_total": 0,
            "scenes_success": 0,
            "scenes_failed": 0,
        }

    def _log(self, message: str) -> None:
        """Print log message."""
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] [LabsGen] {message}")

    def _find_excel_file(self) -> Optional[Path]:
        """Tim file Excel prompts trong thu muc project."""
        # Tim trong thu muc prompts/
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.prompts_path.glob(pattern))
            if files:
                return files[0]

        # Tim truc tiep trong project
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.project_path.glob(pattern))
            if files:
                return files[0]

        return None

    def generate_single_image(
        self,
        prompt: str,
        output_path: Path,
        count: int = 1
    ) -> Tuple[bool, str]:
        """
        Tao mot anh tu prompt.

        Args:
            prompt: Text mo ta anh
            output_path: Duong dan file output
            count: So luong anh (chi lay anh dau tien)

        Returns:
            Tuple[success, error_message]
        """
        success, images, error = self.labs_client.generate_image(
            prompt=prompt,
            count=count,
            aspect_ratio=self.aspect_ratio
        )

        if success and images:
            # Luu anh dau tien
            saved = self.labs_client.save_image(images[0], output_path)
            if saved:
                return True, ""
            return False, "Failed to save image"

        return False, error

    def generate_character_images(
        self,
        excel_path: Optional[Path] = None,
        overwrite: bool = False,
        callback=None
    ) -> Tuple[int, int, List[str]]:
        """
        Tao anh cho tat ca nhan vat trong sheet "characters".

        Args:
            excel_path: Duong dan file Excel (tu tim neu khong chi dinh)
            overwrite: Ghi de anh da co
            callback: Function callback(char_id, status, message)

        Returns:
            Tuple[success_count, failed_count, error_messages]
        """
        self._log("=" * 60)
        self._log("GENERATING CHARACTER IMAGES (Labs API)")
        self._log("=" * 60)

        # Tim file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()

        if excel_path is None or not excel_path.exists():
            return 0, 0, ["Excel file not found"]

        self._log(f"Excel file: {excel_path}")

        errors = []
        success_count = 0
        failed_count = 0

        try:
            # Load workbook
            wb = load_workbook(excel_path)

            if "characters" not in wb.sheetnames:
                return 0, 0, ["Sheet 'characters' not found in Excel"]

            ws = wb["characters"]

            # Get header row
            headers = [cell.value for cell in ws[1]]

            # Find column indices
            col_idx = {
                "id": headers.index("id") if "id" in headers else -1,
                "english_prompt": headers.index("english_prompt") if "english_prompt" in headers else -1,
                "image_file": headers.index("image_file") if "image_file" in headers else -1,
                "status": headers.index("status") if "status" in headers else -1,
            }

            if col_idx["english_prompt"] == -1:
                return 0, 0, ["Column 'english_prompt' not found"]

            # Process each character
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                char_id = row[col_idx["id"]].value if col_idx["id"] >= 0 else f"char_{row_num}"
                prompt = row[col_idx["english_prompt"]].value
                image_file = row[col_idx["image_file"]].value if col_idx["image_file"] >= 0 else f"{char_id}.png"
                status = row[col_idx["status"]].value if col_idx["status"] >= 0 else "pending"

                if not prompt:
                    continue

                self.stats["characters_total"] += 1

                # Check if already done
                output_file = self.nv_path / image_file
                if output_file.exists() and not overwrite:
                    if status == "done":
                        self._log(f"  [SKIP] {char_id}: Already done")
                        success_count += 1
                        self.stats["characters_success"] += 1
                        continue

                self._log(f"\n[GEN] Character: {char_id}")
                self._log(f"   Prompt: {prompt[:80]}...")

                if callback:
                    callback(char_id, "generating", f"Dang tao anh...")

                # Generate image
                success, error = self.generate_single_image(prompt, output_file)

                if success:
                    self._log(f"   [OK] Saved to: {output_file}")
                    success_count += 1
                    self.stats["characters_success"] += 1

                    # Update status in Excel
                    if col_idx["status"] >= 0:
                        row[col_idx["status"]].value = "done"

                    if callback:
                        callback(char_id, "done", str(output_file))
                else:
                    self._log(f"   [FAIL] {error}")
                    failed_count += 1
                    self.stats["characters_failed"] += 1
                    errors.append(f"{char_id}: {error}")

                    if callback:
                        callback(char_id, "failed", error)

                # Delay between requests
                if self.delay > 0:
                    time.sleep(self.delay)

            # Save workbook
            wb.save(excel_path)
            self._log(f"\n[SAVE] Excel updated: {excel_path}")

        except Exception as e:
            errors.append(f"Excel error: {str(e)}")
            self._log(f"[ERROR] {e}")

        self._log(f"\n[SUMMARY] Characters: {success_count} success, {failed_count} failed")
        return success_count, failed_count, errors

    def generate_scene_images(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False,
        callback=None
    ) -> Tuple[int, int, List[str]]:
        """
        Tao anh cho cac scenes trong sheet "scenes".

        Args:
            excel_path: Duong dan file Excel
            start_scene: Scene bat dau (1-indexed)
            end_scene: Scene ket thuc (None = tat ca)
            overwrite: Ghi de anh da co
            callback: Function callback(scene_id, status, message)

        Returns:
            Tuple[success_count, failed_count, error_messages]
        """
        self._log("=" * 60)
        self._log("GENERATING SCENE IMAGES (Labs API)")
        self._log("=" * 60)

        # Tim file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()

        if excel_path is None or not excel_path.exists():
            return 0, 0, ["Excel file not found"]

        self._log(f"Excel file: {excel_path}")

        errors = []
        success_count = 0
        failed_count = 0

        try:
            # Load workbook
            wb = load_workbook(excel_path)

            if "scenes" not in wb.sheetnames:
                return 0, 0, ["Sheet 'scenes' not found in Excel"]

            ws = wb["scenes"]

            # Get header row
            headers = [cell.value for cell in ws[1]]

            # Find column indices
            col_idx = {
                "scene_id": headers.index("scene_id") if "scene_id" in headers else -1,
                "img_prompt": headers.index("img_prompt") if "img_prompt" in headers else -1,
                "img_path": headers.index("img_path") if "img_path" in headers else -1,
                "status_img": headers.index("status_img") if "status_img" in headers else -1,
            }

            if col_idx["img_prompt"] == -1:
                return 0, 0, ["Column 'img_prompt' not found"]

            # Process each scene
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                scene_id = row[col_idx["scene_id"]].value if col_idx["scene_id"] >= 0 else row_num - 1

                # Filter by scene range
                if isinstance(scene_id, int):
                    if scene_id < start_scene:
                        continue
                    if end_scene is not None and scene_id > end_scene:
                        break

                prompt = row[col_idx["img_prompt"]].value
                img_path_val = row[col_idx["img_path"]].value if col_idx["img_path"] >= 0 else None
                status = row[col_idx["status_img"]].value if col_idx["status_img"] >= 0 else "pending"

                if not prompt:
                    continue

                self.stats["scenes_total"] += 1

                # Generate filename
                filename = f"scene_{scene_id:03d}"
                output_file = self.img_path / f"{filename}.png"

                # Check if already done
                if output_file.exists() and not overwrite:
                    if status == "done":
                        self._log(f"  [SKIP] Scene {scene_id}: Already done")
                        success_count += 1
                        self.stats["scenes_success"] += 1
                        continue

                self._log(f"\n[GEN] Scene {scene_id}")
                self._log(f"   Prompt: {prompt[:100]}...")

                if callback:
                    callback(scene_id, "generating", f"Dang tao anh...")

                # Generate image
                success, error = self.generate_single_image(prompt, output_file)

                if success:
                    self._log(f"   [OK] Saved to: {output_file}")
                    success_count += 1
                    self.stats["scenes_success"] += 1

                    # Update Excel
                    if col_idx["img_path"] >= 0:
                        row[col_idx["img_path"]].value = str(output_file.relative_to(self.project_path))
                    if col_idx["status_img"] >= 0:
                        row[col_idx["status_img"]].value = "done"

                    if callback:
                        callback(scene_id, "done", str(output_file))
                else:
                    self._log(f"   [FAIL] {error}")
                    failed_count += 1
                    self.stats["scenes_failed"] += 1
                    errors.append(f"Scene {scene_id}: {error}")

                    if callback:
                        callback(scene_id, "failed", error)

                # Delay between requests
                if self.delay > 0:
                    time.sleep(self.delay)

            # Save workbook
            wb.save(excel_path)
            self._log(f"\n[SAVE] Excel updated: {excel_path}")

        except Exception as e:
            errors.append(f"Excel error: {str(e)}")
            self._log(f"[ERROR] {e}")

        self._log(f"\n[SUMMARY] Scenes: {success_count} success, {failed_count} failed")
        return success_count, failed_count, errors

    def generate_all(
        self,
        excel_path: Optional[Path] = None,
        characters: bool = True,
        scenes: bool = True,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False,
        callback=None
    ) -> Dict[str, Any]:
        """
        Tao tat ca anh (characters + scenes).

        Args:
            excel_path: Duong dan file Excel
            characters: Tao anh characters
            scenes: Tao anh scenes
            start_scene: Scene bat dau
            end_scene: Scene ket thuc
            overwrite: Ghi de
            callback: Callback function

        Returns:
            Dict voi ket qua
        """
        results = {
            "characters": {"success": 0, "failed": 0, "errors": []},
            "scenes": {"success": 0, "failed": 0, "errors": []},
        }

        if characters:
            s, f, e = self.generate_character_images(excel_path, overwrite, callback)
            results["characters"] = {"success": s, "failed": f, "errors": e}

        if scenes:
            s, f, e = self.generate_scene_images(excel_path, start_scene, end_scene, overwrite, callback)
            results["scenes"] = {"success": s, "failed": f, "errors": e}

        # Print summary
        self._log("\n" + "=" * 60)
        self._log("SUMMARY")
        self._log("=" * 60)
        self._log(f"Characters: {results['characters']['success']} success, {results['characters']['failed']} failed")
        self._log(f"Scenes: {results['scenes']['success']} success, {results['scenes']['failed']} failed")

        return results

    def get_stats(self) -> Dict[str, int]:
        """Lay thong ke."""
        return self.stats.copy()

    def test_connection(self) -> Tuple[bool, str]:
        """Test ket noi API."""
        return self.labs_client.test_connection()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def load_config(config_path: str = "config/settings.yaml") -> Dict[str, Any]:
    """Load config tu file YAML."""
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def create_labs_generator_from_config(
    project_path: str,
    config_path: str = "config/settings.yaml",
    verbose: bool = True
) -> LabsFlowGenerator:
    """
    Tao LabsFlowGenerator tu config file.

    Args:
        project_path: Duong dan thu muc project
        config_path: Duong dan file config
        verbose: In log

    Returns:
        LabsFlowGenerator instance
    """
    config = load_config(config_path)

    session_token = config.get("labs_session_token", "")
    if not session_token:
        raise ValueError(
            "Chua cau hinh labs_session_token!\n"
            "Cach lay:\n"
            "1. Cai Cookie Editor (Chrome/Firefox)\n"
            "2. Vao labs.google, dang nhap Google\n"
            "3. Mo Cookie Editor, tim __Secure-next-auth.session-token\n"
            "4. Copy value vao config/settings.yaml"
        )

    captcha_api_key = config.get("captcha_api_key", "")
    if not captcha_api_key:
        print("[WARNING] Chua cau hinh captcha_api_key - co the bi loi khi goi API!")

    return LabsFlowGenerator(
        project_path=Path(project_path),
        session_token=session_token,
        captcha_api_key=captcha_api_key,
        captcha_service=config.get("captcha_service", "capsolver"),
        aspect_ratio=config.get("flow_aspect_ratio", "landscape"),
        delay_between_requests=config.get("labs_delay", 2.0),
        verbose=verbose
    )


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    import sys

    print("""
+==============================================================+
|           LABS FLOW GENERATOR - VE3 TOOL                     |
+==============================================================+
|  Usage:                                                      |
|    python labs_flow_generator.py <project_path> [options]    |
|                                                              |
|  Options:                                                    |
|    --characters    Generate character images only            |
|    --scenes        Generate scene images only                |
|    --all           Generate all (default)                    |
|    --start N       Start from scene N                        |
|    --end N         End at scene N                            |
|    --overwrite     Overwrite existing images                 |
|    --test          Test connection only                      |
+==============================================================+
""")

    if len(sys.argv) < 2:
        print("Error: Please provide project path")
        print("Example: python labs_flow_generator.py ./PROJECTS/KA1-0001")
        sys.exit(1)

    project_path = sys.argv[1]

    # Test mode
    if "--test" in sys.argv:
        print("[*] Testing Labs API connection...")
        try:
            generator = create_labs_generator_from_config(project_path)
            success, message = generator.test_connection()
            if success:
                print(f"[OK] {message}")
            else:
                print(f"[FAIL] {message}")
            sys.exit(0 if success else 1)
        except Exception as e:
            print(f"[ERROR] {e}")
            sys.exit(1)

    # Parse options
    do_characters = "--all" in sys.argv or "--characters" in sys.argv or (
        "--scenes" not in sys.argv and "--characters" not in sys.argv
    )
    do_scenes = "--all" in sys.argv or "--scenes" in sys.argv or (
        "--scenes" not in sys.argv and "--characters" not in sys.argv
    )
    overwrite = "--overwrite" in sys.argv

    start_scene = 1
    end_scene = None

    for i, arg in enumerate(sys.argv):
        if arg == "--start" and i + 1 < len(sys.argv):
            start_scene = int(sys.argv[i + 1])
        if arg == "--end" and i + 1 < len(sys.argv):
            end_scene = int(sys.argv[i + 1])

    # Create generator
    try:
        generator = create_labs_generator_from_config(project_path)

        # Run
        results = generator.generate_all(
            characters=do_characters,
            scenes=do_scenes,
            start_scene=start_scene,
            end_scene=end_scene,
            overwrite=overwrite
        )

        # Exit code
        total_failed = results["characters"]["failed"] + results["scenes"]["failed"]
        sys.exit(0 if total_failed == 0 else 1)

    except FileNotFoundError as e:
        print(f"[ERROR] Config file not found: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"[ERROR] Config error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
