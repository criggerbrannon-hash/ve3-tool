"""
VE3 Tool - Excel Manager Module
===============================
Quản lý file Excel chứa prompts và thông tin nhân vật.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from modules.utils import get_logger


# ============================================================================
# CONSTANTS
# ============================================================================

# Cột cho sheet Characters
CHARACTERS_COLUMNS = [
    "id",               # ID nhân vật (nvc, nvp1, nvp2, ...)
    "role",             # Vai trò (main/supporting)
    "name",             # Tên nhân vật trong truyện
    "english_prompt",   # Prompt tiếng Anh mô tả ngoại hình
    "vietnamese_prompt", # Prompt tiếng Việt (nếu cần)
    "image_file",       # Tên file ảnh tham chiếu (nvc.png, nvp1.png, ...)
    "status",           # Trạng thái (pending/done/error)
    "media_id",         # Media ID từ Google Flow API (dùng cho reference)
]

# Cột cho sheet Scenes
# QUAN TRONG: srt_start/srt_end la timestamp chinh (khong dung start_time/end_time nua)
SCENES_COLUMNS = [
    "scene_id",         # ID scene (1, 2, 3, ...)
    "srt_start",        # Thời gian bắt đầu (HH:MM:SS,mmm) - từ SRT hoặc do đạo diễn set
    "srt_end",          # Thời gian kết thúc (HH:MM:SS,mmm) - từ SRT hoặc do đạo diễn set
    "duration",         # Độ dài tính từ srt_start/srt_end (giây) - auto-calculate
    "planned_duration", # Thời lượng đạo diễn lên kế hoạch cho ảnh này (giây) - dùng khi edit video
    "srt_text",         # Nội dung text của scene
    "img_prompt",       # Prompt tạo ảnh (text)
    "prompt_json",      # JSON prompt đầy đủ (có seed, imageInputs) - dùng khi tạo ảnh
    "video_prompt",     # Prompt tạo video
    "img_path",         # Path đến ảnh đã tạo
    "video_path",       # Path đến video đã tạo
    "status_img",       # Trạng thái ảnh (pending/done/error)
    "status_vid",       # Trạng thái video (pending/done/error)
    # Thông tin reference
    "characters_used",  # JSON list nhân vật trong scene
    "location_used",    # Location ID
    "reference_files",  # JSON list reference files cho image generation
    "media_id",         # Media ID từ Google Flow API (dùng cho I2V - Image to Video)
]


# ============================================================================
# CHARACTER DATA CLASS
# ============================================================================

class Character:
    """Đại diện cho một nhân vật trong truyện."""

    def __init__(
        self,
        id: str,
        role: str = "supporting",
        name: str = "",
        english_prompt: str = "",
        vietnamese_prompt: str = "",
        character_lock: str = "",
        image_file: str = "",
        status: str = "pending",
        is_child: bool = False,
        media_id: str = ""
    ):
        self.id = id
        self.role = role
        self.name = name
        self.english_prompt = english_prompt
        self.vietnamese_prompt = vietnamese_prompt
        self.character_lock = character_lock
        self.image_file = image_file
        self.status = status
        self.is_child = is_child
        self.media_id = media_id

    def to_dict(self) -> Dict[str, Any]:
        """Chuyển đổi thành dictionary."""
        return {
            "id": self.id,
            "role": self.role,
            "name": self.name,
            "english_prompt": self.english_prompt,
            "vietnamese_prompt": self.vietnamese_prompt,
            "character_lock": self.character_lock,
            "image_file": self.image_file,
            "status": self.status,
            "is_child": self.is_child,
            "media_id": self.media_id,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Character":
        """Tạo Character từ dictionary."""
        return cls(
            id=str(data.get("id", "")),
            role=str(data.get("role", "supporting")),
            name=str(data.get("name", "")),
            english_prompt=str(data.get("english_prompt", "")),
            vietnamese_prompt=str(data.get("vietnamese_prompt", "")),
            character_lock=str(data.get("character_lock", "")),
            image_file=str(data.get("image_file", "")),
            status=str(data.get("status", "pending")),
            is_child=bool(data.get("is_child", False)),
            media_id=str(data.get("media_id", "")),
        )


# ============================================================================
# LOCATION DATA CLASS
# ============================================================================

class Location:
    """Dai dien cho mot dia diem trong truyen."""

    def __init__(
        self,
        id: str,
        name: str = "",
        english_prompt: str = "",
        location_lock: str = "",
        lighting_default: str = "",
        image_file: str = "",
        status: str = "pending",
        media_id: str = ""
    ):
        self.id = id
        self.name = name
        self.english_prompt = english_prompt
        self.location_lock = location_lock
        self.lighting_default = lighting_default
        self.image_file = image_file
        self.status = status
        self.media_id = media_id

    def to_dict(self) -> Dict[str, Any]:
        """Chuyen doi thanh dictionary."""
        return {
            "id": self.id,
            "name": self.name,
            "english_prompt": self.english_prompt,
            "location_lock": self.location_lock,
            "lighting_default": self.lighting_default,
            "image_file": self.image_file,
            "status": self.status,
            "media_id": self.media_id,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Location":
        """Tao Location tu dictionary."""
        return cls(
            id=str(data.get("id", "")),
            name=str(data.get("name", "")),
            english_prompt=str(data.get("english_prompt", "")),
            location_lock=str(data.get("location_lock", "")),
            lighting_default=str(data.get("lighting_default", "")),
            image_file=str(data.get("image_file", "")),
            status=str(data.get("status", "pending")),
            media_id=str(data.get("media_id", "")),
        )


# ============================================================================
# SCENE DATA CLASS
# ============================================================================

class Scene:
    """Đại diện cho một scene trong video."""

    def __init__(
        self,
        scene_id: int,
        srt_start: str = "",            # Timestamp bắt đầu (HH:MM:SS,mmm)
        srt_end: str = "",              # Timestamp kết thúc (HH:MM:SS,mmm)
        duration: float = 0.0,          # Độ dài tính từ timestamps (giây) - auto
        planned_duration: float = 0.0,  # Thời lượng đạo diễn lên kế hoạch (giây) - dùng khi edit
        srt_text: str = "",
        img_prompt: str = "",
        prompt_json: str = "",          # JSON prompt đầy đủ (seed, imageInputs)
        video_prompt: str = "",
        img_path: str = "",
        video_path: str = "",
        status_img: str = "pending",
        status_vid: str = "pending",
        # Reference info
        characters_used: str = "",      # JSON list of character IDs used
        location_used: str = "",        # Location ID used
        reference_files: str = "",      # JSON list of reference files
        media_id: str = "",             # Media ID từ Google Flow API (dùng cho I2V)
        # DEPRECATED - giữ để backward compatible, sẽ map sang srt_start/srt_end
        start_time: str = "",
        end_time: str = ""
    ):
        self.scene_id = scene_id
        # Ưu tiên srt_start/srt_end, fallback sang start_time/end_time (backward compatible)
        self.srt_start = str(srt_start) if srt_start else (str(start_time) if start_time else "")
        self.srt_end = str(srt_end) if srt_end else (str(end_time) if end_time else "")
        self.duration = duration
        self.planned_duration = planned_duration  # Thời lượng đạo diễn lên kế hoạch
        self.srt_text = srt_text
        self.img_prompt = img_prompt
        self.prompt_json = prompt_json
        self.video_prompt = video_prompt
        self.img_path = img_path
        self.video_path = video_path
        self.status_img = status_img
        self.status_vid = status_vid
        # References
        self.characters_used = characters_used
        self.location_used = location_used
        self.reference_files = reference_files
        self.media_id = media_id  # Media ID cho I2V

        # DEPRECATED aliases (để code cũ không bị lỗi)
        self.start_time = self.srt_start
        self.end_time = self.srt_end
    
    def to_dict(self) -> Dict[str, Any]:
        """Chuyển đổi thành dictionary."""
        return {
            "scene_id": self.scene_id,
            "srt_start": self.srt_start,
            "srt_end": self.srt_end,
            "duration": self.duration,
            "planned_duration": self.planned_duration,  # Thời lượng đạo diễn
            "srt_text": self.srt_text,
            "img_prompt": self.img_prompt,
            "prompt_json": self.prompt_json,
            "video_prompt": self.video_prompt,
            "img_path": self.img_path,
            "video_path": self.video_path,
            "status_img": self.status_img,
            "status_vid": self.status_vid,
            "characters_used": self.characters_used,
            "location_used": self.location_used,
            "reference_files": self.reference_files,
            "media_id": self.media_id,  # Media ID cho I2V
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Scene":
        """Tao Scene tu dictionary."""
        def safe_int(val, default=0):
            """Convert to int safely, handling time strings like '00:00'."""
            if val is None or val == "":
                return default
            if isinstance(val, int):
                return val
            if isinstance(val, float):
                return int(val)
            if isinstance(val, str):
                # Handle time format "HH:MM" or "MM:SS"
                if ":" in val:
                    return default  # Skip time strings
                try:
                    return int(val)
                except ValueError:
                    return default
            return default

        def safe_float(val, default=0.0):
            """Convert to float safely."""
            if val is None or val == "":
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        return cls(
            scene_id=safe_int(data.get("scene_id", 0)),
            srt_start=str(data.get("srt_start", "") or ""),  # Timestamp string
            srt_end=str(data.get("srt_end", "") or ""),      # Timestamp string
            duration=safe_float(data.get("duration", 0.0)),
            planned_duration=safe_float(data.get("planned_duration", 0.0)),  # Thời lượng đạo diễn
            srt_text=str(data.get("srt_text", "") or ""),
            img_prompt=str(data.get("img_prompt", "") or ""),
            prompt_json=str(data.get("prompt_json", "") or ""),
            video_prompt=str(data.get("video_prompt", "") or ""),
            img_path=str(data.get("img_path", "") or ""),
            video_path=str(data.get("video_path", "") or ""),
            status_img=str(data.get("status_img", "pending") or "pending"),
            status_vid=str(data.get("status_vid", "pending") or "pending"),
            # DEPRECATED: backward compatible - sẽ được map vào srt_start/srt_end
            start_time=str(data.get("start_time", "") or ""),
            end_time=str(data.get("end_time", "") or ""),
            characters_used=str(data.get("characters_used", "") or ""),
            location_used=str(data.get("location_used", "") or ""),
            reference_files=str(data.get("reference_files", "") or ""),
            media_id=str(data.get("media_id", "") or ""),  # Media ID cho I2V
        )


# ============================================================================
# PROMPT WORKBOOK CLASS
# ============================================================================

class PromptWorkbook:
    """
    Class quản lý file Excel chứa prompts.
    
    Attributes:
        path: Path đến file Excel
        workbook: Workbook object
        characters_sheet: Sheet chứa thông tin nhân vật
        scenes_sheet: Sheet chứa thông tin các scene
    """
    
    CHARACTERS_SHEET = "characters"
    SCENES_SHEET = "scenes"
    
    def __init__(self, path: Path):
        """
        Khởi tạo PromptWorkbook.
        
        Args:
            path: Path đến file Excel
        """
        self.path = path
        self.workbook: Optional[Workbook] = None
        self.logger = get_logger("excel_manager")
    
    def load_or_create(self) -> "PromptWorkbook":
        """
        Load file Excel nếu tồn tại, hoặc tạo mới nếu chưa có.
        
        Returns:
            self để hỗ trợ method chaining
        """
        if self.path.exists():
            self.logger.info(f"Loading existing Excel file: {self.path}")
            self.workbook = load_workbook(self.path)
        else:
            self.logger.info(f"Creating new Excel file: {self.path}")
            self._create_new_workbook()
        
        return self
    
    def _create_new_workbook(self) -> None:
        """Tạo workbook mới với cấu trúc chuẩn."""
        self.workbook = Workbook()
        
        # Xóa sheet mặc định
        default_sheet = self.workbook.active
        
        # Tạo sheet Characters
        self._create_characters_sheet()
        
        # Tạo sheet Scenes
        self._create_scenes_sheet()
        
        # Xóa sheet mặc định
        if default_sheet and default_sheet.title == "Sheet":
            self.workbook.remove(default_sheet)
        
        # Lưu file
        self.save()
    
    def _create_characters_sheet(self) -> None:
        """Tạo sheet Characters với header."""
        ws = self.workbook.create_sheet(self.CHARACTERS_SHEET)
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Thêm header
        for col, column_name in enumerate(CHARACTERS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Điều chỉnh độ rộng cột
        column_widths = {
            "id": 10,
            "role": 12,
            "name": 20,
            "english_prompt": 60,
            "vietnamese_prompt": 40,
            "image_file": 15,
            "status": 10,
        }
        
        for col, column_name in enumerate(CHARACTERS_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)
    
    def _create_scenes_sheet(self) -> None:
        """Tạo sheet Scenes với header."""
        ws = self.workbook.create_sheet(self.SCENES_SHEET)
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Thêm header
        for col, column_name in enumerate(SCENES_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Điều chỉnh độ rộng cột
        column_widths = {
            "scene_id": 10,
            "srt_start": 10,
            "srt_end": 10,
            "srt_text": 50,
            "img_prompt": 70,
            "video_prompt": 70,
            "img_path": 30,
            "video_path": 30,
            "status_img": 12,
            "status_vid": 12,
        }
        
        for col, column_name in enumerate(SCENES_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)
    
    def save(self) -> None:
        """Lưu workbook ra file."""
        if self.workbook is None:
            raise RuntimeError("Workbook chưa được load hoặc tạo")
        
        # Đảm bảo thư mục tồn tại
        self.path.parent.mkdir(parents=True, exist_ok=True)
        
        self.workbook.save(self.path)
        self.logger.debug(f"Saved Excel file: {self.path}")
    
    # ========================================================================
    # CHARACTERS METHODS
    # ========================================================================
    
    def get_characters(self) -> List[Character]:
        """
        Lấy danh sách tất cả nhân vật.
        
        Returns:
            List các Character objects
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.CHARACTERS_SHEET]
        characters = []
        
        # Đọc từ dòng 2 (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            
            data = dict(zip(CHARACTERS_COLUMNS, row))
            characters.append(Character.from_dict(data))
        
        return characters
    
    def add_character(self, character: Character) -> None:
        """
        Thêm nhân vật mới vào sheet.
        
        Args:
            character: Character object
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.CHARACTERS_SHEET]
        
        # Tìm dòng trống tiếp theo
        next_row = ws.max_row + 1
        
        # Thêm dữ liệu
        data = character.to_dict()
        for col, column_name in enumerate(CHARACTERS_COLUMNS, start=1):
            ws.cell(row=next_row, column=col, value=data.get(column_name, ""))
        
        self.logger.debug(f"Added character: {character.id}")
    
    def update_character(self, character_id: str, **kwargs) -> bool:
        """
        Cập nhật thông tin nhân vật.
        
        Args:
            character_id: ID của nhân vật cần cập nhật
            **kwargs: Các field cần cập nhật
            
        Returns:
            True nếu cập nhật thành công, False nếu không tìm thấy
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.CHARACTERS_SHEET]
        
        # Tìm dòng có character_id
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == character_id:
                # Cập nhật các field
                for key, value in kwargs.items():
                    if key in CHARACTERS_COLUMNS:
                        col_idx = CHARACTERS_COLUMNS.index(key) + 1
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.debug(f"Updated character: {character_id}")
                return True
        
        self.logger.warning(f"Character not found: {character_id}")
        return False
    
    def clear_characters(self) -> None:
        """Xóa tất cả nhân vật (giữ lại header)."""
        if self.workbook is None:
            self.load_or_create()

        ws = self.workbook[self.CHARACTERS_SHEET]

        # Xóa tất cả dòng trừ header
        ws.delete_rows(2, ws.max_row)
        self.logger.debug("Cleared all characters")

    def get_media_ids(self) -> Dict[str, str]:
        """
        Lấy tất cả media_id từ characters sheet.

        Returns:
            Dict mapping character_id -> media_id
            VD: {"nvc": "CAMSJDZiYzQ2...", "nv1": "CAMSJDZiYzQ1..."}
        """
        if self.workbook is None:
            self.load_or_create()

        result = {}
        ws = self.workbook[self.CHARACTERS_SHEET]

        # Tìm cột media_id
        media_id_col = None
        for col_idx, col_name in enumerate(CHARACTERS_COLUMNS, start=1):
            if col_name == "media_id":
                media_id_col = col_idx
                break

        if media_id_col is None:
            return result

        # Đọc từ dòng 2 (skip header)
        for row_idx in range(2, ws.max_row + 1):
            char_id = ws.cell(row=row_idx, column=1).value
            media_id = ws.cell(row=row_idx, column=media_id_col).value

            if char_id and media_id:
                result[str(char_id)] = str(media_id)

        self.logger.debug(f"Loaded {len(result)} media_ids from Excel")
        return result

    def get_scene_media_ids(self) -> Dict[str, str]:
        """
        Lấy tất cả media_id từ scenes sheet (cho I2V).

        Returns:
            Dict mapping scene_id (string) -> media_id
            VD: {"1": "CAMSJDZiYzQ2...", "2": "CAMSJDZiYzQ1..."}
        """
        if self.workbook is None:
            self.load_or_create()

        result = {}
        ws = self.workbook[self.SCENES_SHEET]

        # Tìm cột media_id
        media_id_col = None
        for col_idx, col_name in enumerate(SCENES_COLUMNS, start=1):
            if col_name == "media_id":
                media_id_col = col_idx
                break

        if media_id_col is None:
            return result

        # Đọc từ dòng 2 (skip header)
        for row_idx in range(2, ws.max_row + 1):
            scene_id = ws.cell(row=row_idx, column=1).value
            media_id = ws.cell(row=row_idx, column=media_id_col).value

            if scene_id and media_id:
                result[str(scene_id)] = str(media_id)

        self.logger.debug(f"Loaded {len(result)} scene media_ids from Excel")
        return result

    # ========================================================================
    # SCENES METHODS
    # ========================================================================
    
    def get_scenes(self) -> List[Scene]:
        """
        Lấy danh sách tất cả scenes.
        
        Returns:
            List các Scene objects
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        scenes = []
        
        # Đọc từ dòng 2 (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            
            data = dict(zip(SCENES_COLUMNS, row))
            scenes.append(Scene.from_dict(data))
        
        return scenes
    
    def add_scene(self, scene: Scene) -> None:
        """
        Thêm scene mới vào sheet.
        
        Args:
            scene: Scene object
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        
        # Tìm dòng trống tiếp theo
        next_row = ws.max_row + 1
        
        # Thêm dữ liệu
        data = scene.to_dict()
        for col, column_name in enumerate(SCENES_COLUMNS, start=1):
            ws.cell(row=next_row, column=col, value=data.get(column_name, ""))
        
        self.logger.debug(f"Added scene: {scene.scene_id}")
    
    def update_scene(self, scene_id: int, **kwargs) -> bool:
        """
        Cập nhật thông tin scene.
        
        Args:
            scene_id: ID của scene cần cập nhật
            **kwargs: Các field cần cập nhật
            
        Returns:
            True nếu cập nhật thành công, False nếu không tìm thấy
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        
        # Tìm dòng có scene_id
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value is not None and int(cell_value) == scene_id:
                # Cập nhật các field
                for key, value in kwargs.items():
                    if key in SCENES_COLUMNS:
                        col_idx = SCENES_COLUMNS.index(key) + 1
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.debug(f"Updated scene: {scene_id}")
                return True
        
        self.logger.warning(f"Scene not found: {scene_id}")
        return False
    
    def clear_scenes(self) -> None:
        """Xóa tất cả scenes (giữ lại header)."""
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        
        # Xóa tất cả dòng trừ header
        ws.delete_rows(2, ws.max_row)
        self.logger.debug("Cleared all scenes")
    
    def get_pending_image_scenes(self) -> List[Scene]:
        """Lấy danh sách scenes chưa tạo ảnh."""
        scenes = self.get_scenes()
        return [s for s in scenes if s.status_img != "done" and s.img_prompt]
    
    def get_pending_video_scenes(self) -> List[Scene]:
        """Lấy danh sách scenes chưa tạo video (nhưng đã có ảnh)."""
        scenes = self.get_scenes()
        return [s for s in scenes if s.status_vid != "done" and s.img_path and s.video_prompt]
    
    # ========================================================================
    # UTILITY METHODS
    # ========================================================================
    
    def has_prompts(self) -> bool:
        """Kiểm tra xem đã có prompt nào chưa."""
        scenes = self.get_scenes()
        return any(s.img_prompt for s in scenes)
    
    def get_stats(self) -> Dict[str, Any]:
        """
        Lấy thống kê tổng quan.
        
        Returns:
            Dictionary chứa các thống kê
        """
        characters = self.get_characters()
        scenes = self.get_scenes()
        
        return {
            "total_characters": len(characters),
            "total_scenes": len(scenes),
            "scenes_with_prompts": sum(1 for s in scenes if s.img_prompt),
            "images_done": sum(1 for s in scenes if s.status_img == "done"),
            "images_error": sum(1 for s in scenes if s.status_img == "error"),
            "videos_done": sum(1 for s in scenes if s.status_vid == "done"),
            "videos_error": sum(1 for s in scenes if s.status_vid == "error"),
        }
