"""
VE3 Tool - Flow Image Generator Module
======================================
Tích hợp Google Flow API vào pipeline để tự động tạo ảnh từ Excel prompts.

Workflow:
1. Đọc Excel prompts (characters + scenes sheets)
2. Tạo ảnh NV (nhân vật) trước - lưu vào thư mục nv/
3. Tạo ảnh scenes sau - lưu vào thư mục img/ với REFERENCE IMAGES từ nv/
4. Cập nhật Excel với đường dẫn ảnh và status
"""

import os
import time
import json
import base64
import yaml
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

from .google_flow_api import GoogleFlowAPI, AspectRatio, ImageInput, ImageInputType, GeneratedImage


class FlowImageGenerator:
    """
    Generator ảnh sử dụng Google Flow API.
    Đọc prompts từ Excel và tạo ảnh tự động.
    """

    def __init__(
        self,
        project_path: Path = None,
        bearer_token: str = None,
        project_id: Optional[str] = None,
        aspect_ratio: str = "landscape",
        delay_between_requests: float = 3.0,
        verbose: bool = True
    ):
        """
        Khởi tạo Flow Image Generator.

        Args:
            project_path: Đường dẫn đến thư mục project (PROJECTS/{CODE}/)
            bearer_token: Google Flow Bearer token (có thể để None nếu dùng generate_and_save với token riêng)
            project_id: Flow Project ID (optional)
            aspect_ratio: Tỷ lệ khung hình (landscape/portrait/square)
            delay_between_requests: Thời gian chờ giữa các request (giây)
            verbose: In log chi tiết
        """
        self.project_path = Path(project_path) if project_path else None
        self.bearer_token = bearer_token
        self.project_id = project_id
        self.delay = delay_between_requests
        self.verbose = verbose

        # Map aspect ratio
        ar_map = {
            "landscape": AspectRatio.LANDSCAPE,
            "portrait": AspectRatio.PORTRAIT,
            "square": AspectRatio.SQUARE,
            "16:9": AspectRatio.LANDSCAPE,
            "9:16": AspectRatio.PORTRAIT,
            "1:1": AspectRatio.SQUARE,
        }
        self.aspect_ratio = ar_map.get(aspect_ratio.lower(), AspectRatio.LANDSCAPE)

        # Tạo Flow API client (nếu có bearer_token)
        self.flow_client = None
        if bearer_token:
            self.flow_client = GoogleFlowAPI(
                bearer_token=bearer_token,
                project_id=project_id,
                verbose=verbose
            )

        # Paths (chỉ khởi tạo nếu có project_path)
        self.nv_path = None
        self.img_path = None
        self.prompts_path = None

        if self.project_path:
            self.nv_path = self.project_path / "nv"
            self.img_path = self.project_path / "img"
            self.prompts_path = self.project_path / "prompts"

            # Tạo thư mục nếu chưa có
            self.nv_path.mkdir(parents=True, exist_ok=True)
            self.img_path.mkdir(parents=True, exist_ok=True)
        
        # Stats
        self.stats = {
            "characters_total": 0,
            "characters_success": 0,
            "characters_failed": 0,
            "scenes_total": 0,
            "scenes_success": 0,
            "scenes_failed": 0,
        }

        # Character references cache - map character_id -> GeneratedImage
        # Dùng để reference khi tạo scene images
        self.character_references: Dict[str, GeneratedImage] = {}

        # Reference image settings
        self.use_character_references = True  # Enable/disable reference feature
    
    def _log(self, message: str) -> None:
        """Print log message."""
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] {message}")

    def upload_character_as_reference(
        self,
        char_id: str,
        image_path: Path
    ) -> bool:
        """
        Upload ảnh character có sẵn lên làm reference.

        Sử dụng khi bạn đã có ảnh nhân vật và muốn dùng nó làm reference
        mà không cần generate lại.

        Args:
            char_id: ID của nhân vật (vd: "nv01")
            image_path: Đường dẫn đến file ảnh

        Returns:
            True nếu upload thành công
        """
        self._log(f"📤 Uploading {char_id} reference: {image_path}")

        success, img_input, error = self.flow_client.upload_image(image_path)

        if success and img_input:
            # Tạo GeneratedImage object để lưu vào cache
            ref_image = GeneratedImage(
                media_name=img_input.name,
                local_path=image_path
            )
            self.character_references[char_id] = ref_image
            self._log(f"✅ Uploaded {char_id} reference successfully")
            return True
        else:
            self._log(f"❌ Upload failed: {error}")
            return False

    def upload_all_existing_characters(self) -> int:
        """
        Upload tất cả ảnh character có sẵn trong thư mục nv/ làm reference.

        Returns:
            Số lượng references đã upload thành công
        """
        self._log("📤 Uploading all existing character images as references...")

        uploaded = 0
        for img_path in self.nv_path.glob("*.png"):
            char_id = img_path.stem  # nv01.png -> nv01
            if self.upload_character_as_reference(char_id, img_path):
                uploaded += 1

        self._log(f"✅ Uploaded {uploaded} character references")
        return uploaded

    def _auto_upload_existing_characters(self) -> int:
        """
        Tự động upload ảnh character có sẵn khi chưa có trong cache.
        Được gọi tự động trước khi generate scenes.

        Returns:
            Số lượng references đã upload thành công
        """
        if not self.nv_path.exists():
            return 0

        uploaded = 0
        for img_path in self.nv_path.glob("*.png"):
            char_id = img_path.stem  # nv01.png -> nv01

            # Skip nếu đã có trong cache
            if char_id in self.character_references:
                continue

            self._log(f"  📤 Uploading {char_id}: {img_path.name}...")

            success, img_input, error = self.flow_client.upload_image(img_path)

            if success and img_input:
                # Tạo GeneratedImage object để lưu vào cache
                ref_image = GeneratedImage(
                    media_name=img_input.name,
                    local_path=img_path
                )
                self.character_references[char_id] = ref_image
                self._log(f"  ✅ {char_id}: Got media_name")
                uploaded += 1
            else:
                self._log(f"  ❌ {char_id}: Upload failed - {error}")

        return uploaded
    
    def _find_excel_file(self) -> Optional[Path]:
        """Tìm file Excel prompts trong thư mục project."""
        # Tìm trong thư mục prompts/
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.prompts_path.glob(pattern))
            if files:
                return files[0]
        
        # Tìm trực tiếp trong project
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.project_path.glob(pattern))
            if files:
                return files[0]
        
        return None

    def _load_image_as_base64(self, image_path: Path) -> Optional[str]:
        """
        Load ảnh từ file và convert sang base64.

        Args:
            image_path: Path đến file ảnh

        Returns:
            Base64 encoded string hoặc None nếu lỗi
        """
        try:
            if not image_path.exists():
                self._log(f"  ⚠️  Reference image not found: {image_path}")
                return None

            with open(image_path, "rb") as f:
                img_data = f.read()

            b64_data = base64.b64encode(img_data).decode("utf-8")
            self._log(f"  ✓ Loaded reference: {image_path.name} ({len(img_data) / 1024:.1f}KB)")
            return b64_data

        except Exception as e:
            self._log(f"  ❌ Error loading {image_path}: {e}")
            return None

    def _get_reference_images(self, reference_files: str) -> List[str]:
        """
        Load tất cả reference images từ danh sách files.

        Args:
            reference_files: JSON string hoặc comma-separated string của file names
                           Ví dụ: '["nv1.png", "loc1.png"]' hoặc 'nv1.png, loc1.png'

        Returns:
            List of base64 encoded images
        """
        if not reference_files:
            return []

        # Parse reference files
        file_list = []
        try:
            # Try JSON format first
            parsed = json.loads(reference_files)
            if isinstance(parsed, list):
                file_list = parsed
            elif isinstance(parsed, str):
                file_list = [parsed]
        except (json.JSONDecodeError, TypeError):
            # Fallback to comma-separated
            file_list = [f.strip() for f in str(reference_files).split(",") if f.strip()]

        if not file_list:
            return []

        # Check if nv_path is available
        if not self.nv_path:
            self._log(f"  ⚠️  Cannot load reference images: nv_path not set")
            return []

        self._log(f"  📎 Loading {len(file_list)} reference images: {file_list}")

        # Load each reference image from nv/ folder (all reference images are in nv/)
        base64_images = []
        for filename in file_list:
            # Reference images are always in nv/ folder
            img_path = self.nv_path / filename
            b64 = self._load_image_as_base64(img_path)
            if b64:
                base64_images.append(b64)

        return base64_images

    def generate_character_images(
        self,
        excel_path: Optional[Path] = None,
        overwrite: bool = False
    ) -> Tuple[int, int, List[str]]:
        """
        Tạo ảnh cho tất cả nhân vật trong sheet "characters".
        
        Args:
            excel_path: Đường dẫn file Excel (tự tìm nếu không chỉ định)
            overwrite: Ghi đè ảnh đã có
            
        Returns:
            Tuple[success_count, failed_count, error_messages]
        """
        self._log("=" * 60)
        self._log("GENERATING CHARACTER IMAGES")
        self._log("=" * 60)
        
        # Tìm file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()
        
        if excel_path is None or not excel_path.exists():
            return 0, 0, ["Excel file not found"]
        
        self._log(f"Excel file: {excel_path}")
        
        errors = []
        success_count = 0
        failed_count = 0
        
        try:
            # Load workbook
            wb = load_workbook(excel_path)
            
            if "characters" not in wb.sheetnames:
                return 0, 0, ["Sheet 'characters' not found in Excel"]
            
            ws = wb["characters"]
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Find column indices
            col_idx = {
                "id": headers.index("id") if "id" in headers else -1,
                "english_prompt": headers.index("english_prompt") if "english_prompt" in headers else -1,
                "image_file": headers.index("image_file") if "image_file" in headers else -1,
                "status": headers.index("status") if "status" in headers else -1,
            }
            
            if col_idx["english_prompt"] == -1:
                return 0, 0, ["Column 'english_prompt' not found"]
            
            # Process each character
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                char_id = row[col_idx["id"]].value if col_idx["id"] >= 0 else f"char_{row_num}"
                prompt = row[col_idx["english_prompt"]].value
                image_file = row[col_idx["image_file"]].value if col_idx["image_file"] >= 0 else f"{char_id}.png"
                status = row[col_idx["status"]].value if col_idx["status"] >= 0 else "pending"
                
                if not prompt:
                    continue
                
                self.stats["characters_total"] += 1
                
                # Check if already done
                output_file = self.nv_path / image_file
                if output_file.exists() and not overwrite:
                    if status == "done":
                        self._log(f"  ⏭️  {char_id}: Already done, skipping")
                        success_count += 1
                        self.stats["characters_success"] += 1
                        continue
                
                self._log(f"\n🎨 Generating image for character: {char_id}")
                self._log(f"   Prompt: {prompt[:80]}...")

                # Generate image
                success, images, error = self.flow_client.generate_images(
                    prompt=prompt,
                    count=1,
                    aspect_ratio=self.aspect_ratio
                )

                if success and images:
                    # Download image
                    filename = image_file.replace(".png", "")
                    downloaded = self.flow_client.download_image(
                        images[0],
                        self.nv_path,
                        filename
                    )

                    if downloaded:
                        self._log(f"   ✅ Saved to: {downloaded}")
                        success_count += 1
                        self.stats["characters_success"] += 1

                        # QUAN TRỌNG: Lưu reference cho character này
                        # để dùng khi generate scenes sau
                        if images[0].media_name:
                            self.character_references[char_id] = images[0]
                            self._log(f"   📌 Saved reference for {char_id}")

                        # Update status in Excel
                        if col_idx["status"] >= 0:
                            row[col_idx["status"]].value = "done"
                    else:
                        self._log(f"   ❌ Download failed")
                        failed_count += 1
                        self.stats["characters_failed"] += 1
                        errors.append(f"{char_id}: Download failed")
                else:
                    self._log(f"   ❌ Generation failed: {error}")
                    failed_count += 1
                    self.stats["characters_failed"] += 1
                    errors.append(f"{char_id}: {error}")
                
                # Delay between requests
                if self.delay > 0:
                    time.sleep(self.delay)
            
            # Save workbook
            wb.save(excel_path)
            self._log(f"\n💾 Excel updated: {excel_path}")
            
        except Exception as e:
            errors.append(f"Excel error: {str(e)}")
            self._log(f"❌ Error: {e}")
        
        self._log(f"\n📊 Characters: {success_count} success, {failed_count} failed")
        return success_count, failed_count, errors
    
    def generate_scene_images(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False
    ) -> Tuple[int, int, List[str]]:
        """
        Tạo ảnh cho các scenes trong sheet "scenes".
        
        Args:
            excel_path: Đường dẫn file Excel
            start_scene: Scene bắt đầu (1-indexed)
            end_scene: Scene kết thúc (None = tất cả)
            overwrite: Ghi đè ảnh đã có
            
        Returns:
            Tuple[success_count, failed_count, error_messages]
        """
        self._log("=" * 60)
        self._log("GENERATING SCENE IMAGES")
        self._log("=" * 60)
        
        # Tìm file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()
        
        if excel_path is None or not excel_path.exists():
            return 0, 0, ["Excel file not found"]
        
        self._log(f"Excel file: {excel_path}")
        
        errors = []
        success_count = 0
        failed_count = 0
        
        try:
            # Load workbook
            wb = load_workbook(excel_path)
            
            if "scenes" not in wb.sheetnames:
                return 0, 0, ["Sheet 'scenes' not found in Excel"]
            
            ws = wb["scenes"]
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Find column indices
            col_idx = {
                "scene_id": headers.index("scene_id") if "scene_id" in headers else -1,
                "img_prompt": headers.index("img_prompt") if "img_prompt" in headers else -1,
                "img_path": headers.index("img_path") if "img_path" in headers else -1,
                "status_img": headers.index("status_img") if "status_img" in headers else -1,
                "reference_files": headers.index("reference_files") if "reference_files" in headers else -1,
                "characters_used": headers.index("characters_used") if "characters_used" in headers else -1,
                "location_used": headers.index("location_used") if "location_used" in headers else -1,
            }

            # Tìm cột characters (có thể có nhiều tên khác nhau)
            char_col_idx = -1
            for i, h in enumerate(headers):
                if h and str(h).lower() in ["characters", "character_ids", "nv_ids", "nhan_vat", "char_ids"]:
                    char_col_idx = i
                    break

            if col_idx["img_prompt"] == -1:
                return 0, 0, ["Column 'img_prompt' not found"]

            # Log reference status
            if self.use_character_references and self.character_references:
                self._log(f"📌 Using {len(self.character_references)} character references")
            elif self.use_character_references:
                # Nếu chưa có references trong cache, tự động upload ảnh có sẵn
                self._log("📌 No references in cache, checking for existing character images...")
                self._auto_upload_existing_characters()
                if self.character_references:
                    self._log(f"📌 Uploaded {len(self.character_references)} character references")

            # Process each scene
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                scene_id = row[col_idx["scene_id"]].value if col_idx["scene_id"] >= 0 else row_num - 1

                # Filter by scene range
                if isinstance(scene_id, int):
                    if scene_id < start_scene:
                        continue
                    if end_scene is not None and scene_id > end_scene:
                        break

                prompt = row[col_idx["img_prompt"]].value
                img_path_val = row[col_idx["img_path"]].value if col_idx["img_path"] >= 0 else None
                status = row[col_idx["status_img"]].value if col_idx["status_img"] >= 0 else "pending"

                # Lấy danh sách character IDs cho scene này
                scene_characters = []
                if char_col_idx >= 0 and char_col_idx < len(row):
                    char_value = row[char_col_idx].value
                    if char_value:
                        # Parse character IDs (có thể là "nv01, nv02" hoặc "nv01;nv02")
                        char_str = str(char_value)
                        for sep in [",", ";", "|"]:
                            if sep in char_str:
                                scene_characters = [c.strip() for c in char_str.split(sep)]
                                break
                        if not scene_characters:
                            scene_characters = [char_str.strip()]

                # Also check reference_files column
                reference_files = ""
                if col_idx["reference_files"] >= 0:
                    reference_files = row[col_idx["reference_files"]].value or ""

                # Fallback: build reference from characters_used and location_used
                if not reference_files and not scene_characters:
                    ref_list = []
                    if col_idx["characters_used"] >= 0:
                        chars_used = row[col_idx["characters_used"]].value or ""
                        try:
                            chars = json.loads(chars_used) if chars_used.startswith("[") else [c.strip() for c in chars_used.split(",") if c.strip()]
                            scene_characters.extend(chars)
                        except:
                            pass
                    if col_idx["location_used"] >= 0:
                        loc_used = row[col_idx["location_used"]].value or ""
                        if loc_used:
                            scene_characters.append(loc_used)

                if not prompt:
                    continue

                self.stats["scenes_total"] += 1

                # Generate filename
                filename = f"scene_{scene_id:03d}"
                output_file = self.img_path / f"{filename}.png"

                # Check if already done
                if output_file.exists() and not overwrite:
                    if status == "done":
                        self._log(f"  ⏭️  Scene {scene_id}: Already done, skipping")
                        success_count += 1
                        self.stats["scenes_success"] += 1
                        continue

                self._log(f"\n🎬 Generating image for Scene {scene_id}")
                self._log(f"   Prompt: {prompt[:100]}...")

                # Collect reference images for this scene
                # QUAN TRONG: Dùng media_name từ cache, KHÔNG dùng base64!
                reference_images = []
                if self.use_character_references and scene_characters:
                    for char_id in scene_characters:
                        if char_id in self.character_references:
                            ref_img = self.character_references[char_id]
                            if ref_img.media_name:
                                reference_images.append(ref_img)
                                # Debug: Show actual media_name being used
                                self._log(f"   📌 Using reference: {char_id} -> media_name={ref_img.media_name[:50]}...")
                            else:
                                self._log(f"   ⚠️ {char_id}: No media_name available!")
                        else:
                            self._log(f"   ⚠️ {char_id}: Not found in character_references cache")

                if reference_images:
                    self._log(f"   🖼️  Using {len(reference_images)} reference images for consistency")

                # Generate image (with or without references)
                success, images, error = self.flow_client.generate_images(
                    prompt=prompt,
                    count=1,
                    aspect_ratio=self.aspect_ratio,
                    reference_images=reference_images if reference_images else None
                )

                if success and images:
                    # Download image
                    downloaded = self.flow_client.download_image(
                        images[0],
                        self.img_path,
                        filename
                    )
                    
                    if downloaded:
                        self._log(f"   ✅ Saved to: {downloaded}")
                        success_count += 1
                        self.stats["scenes_success"] += 1
                        
                        # Update Excel
                        if col_idx["img_path"] >= 0:
                            row[col_idx["img_path"]].value = str(downloaded.relative_to(self.project_path))
                        if col_idx["status_img"] >= 0:
                            row[col_idx["status_img"]].value = "done"
                    else:
                        self._log(f"   ❌ Download failed")
                        failed_count += 1
                        self.stats["scenes_failed"] += 1
                        errors.append(f"Scene {scene_id}: Download failed")
                else:
                    self._log(f"   ❌ Generation failed: {error}")
                    failed_count += 1
                    self.stats["scenes_failed"] += 1
                    errors.append(f"Scene {scene_id}: {error}")
                
                # Delay between requests
                if self.delay > 0:
                    time.sleep(self.delay)
            
            # Save workbook
            wb.save(excel_path)
            self._log(f"\n💾 Excel updated: {excel_path}")
            
        except Exception as e:
            errors.append(f"Excel error: {str(e)}")
            self._log(f"❌ Error: {e}")
        
        self._log(f"\n📊 Scenes: {success_count} success, {failed_count} failed")
        return success_count, failed_count, errors
    
    def generate_all(
        self,
        excel_path: Optional[Path] = None,
        characters: bool = True,
        scenes: bool = True,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False
    ) -> Dict[str, Any]:
        """
        Tạo tất cả ảnh (characters + scenes).
        
        Args:
            excel_path: Đường dẫn file Excel
            characters: Tạo ảnh characters
            scenes: Tạo ảnh scenes
            start_scene: Scene bắt đầu
            end_scene: Scene kết thúc
            overwrite: Ghi đè
            
        Returns:
            Dict với kết quả
        """
        results = {
            "characters": {"success": 0, "failed": 0, "errors": []},
            "scenes": {"success": 0, "failed": 0, "errors": []},
        }
        
        if characters:
            s, f, e = self.generate_character_images(excel_path, overwrite)
            results["characters"] = {"success": s, "failed": f, "errors": e}
        
        if scenes:
            s, f, e = self.generate_scene_images(excel_path, start_scene, end_scene, overwrite)
            results["scenes"] = {"success": s, "failed": f, "errors": e}
        
        # Print summary
        self._log("\n" + "=" * 60)
        self._log("SUMMARY")
        self._log("=" * 60)
        self._log(f"Characters: {results['characters']['success']} success, {results['characters']['failed']} failed")
        self._log(f"Scenes: {results['scenes']['success']} success, {results['scenes']['failed']} failed")
        
        return results
    
    def generate_and_save(
        self,
        prompt: str,
        output_path: str,
        token: str = None,
        project_id: str = None,
        reference_images: List[str] = None
    ) -> bool:
        """
        Tạo một ảnh đơn lẻ và lưu vào file.

        Args:
            prompt: Prompt mô tả ảnh
            output_path: Đường dẫn lưu ảnh
            token: Bearer token (nếu khác với token đã cấu hình)
            project_id: Project ID (nếu khác với ID đã cấu hình)
            reference_images: List of base64 encoded reference images

        Returns:
            True nếu thành công, False nếu thất bại
        """
        self._log(f"🎨 Generating single image...")
        self._log(f"   Prompt: {prompt[:80]}...")

        # Create new Flow client if token provided
        if token:
            from .google_flow_api import GoogleFlowAPI
            flow_client = GoogleFlowAPI(
                bearer_token=token,
                project_id=project_id or self.project_id,
                verbose=self.verbose
            )
        else:
            flow_client = self.flow_client

        try:
            # Generate image
            success, images, error = flow_client.generate_images(
                prompt=prompt,
                count=1,
                aspect_ratio=self.aspect_ratio,
                image_inputs=reference_images
            )

            if not success or not images:
                self._log(f"   ❌ Generation failed: {error}")
                return False

            # Download/save to output path
            output_dir = Path(output_path).parent
            filename = Path(output_path).stem

            downloaded = flow_client.download_image(
                images[0],
                output_dir,
                filename
            )

            if downloaded:
                self._log(f"   ✅ Saved to: {output_path}")
                return True
            else:
                self._log(f"   ❌ Download failed")
                return False

        except Exception as e:
            self._log(f"   ❌ Error: {e}")
            return False

    def get_stats(self) -> Dict[str, int]:
        """Lấy thống kê."""
        return self.stats.copy()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def load_config(config_path: str = "config/settings.yaml") -> Dict[str, Any]:
    """Load config từ file YAML."""
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def create_generator_from_config(
    project_path: str,
    config_path: str = "config/settings.yaml",
    verbose: bool = True
) -> FlowImageGenerator:
    """
    Tạo FlowImageGenerator từ config file.
    
    Args:
        project_path: Đường dẫn thư mục project
        config_path: Đường dẫn file config
        verbose: In log
        
    Returns:
        FlowImageGenerator instance
    """
    config = load_config(config_path)
    
    return FlowImageGenerator(
        project_path=Path(project_path),
        bearer_token=config.get("flow_bearer_token", ""),
        project_id=config.get("flow_project_id"),
        aspect_ratio=config.get("flow_aspect_ratio", "landscape"),
        delay_between_requests=config.get("flow_delay", 3.0),
        verbose=verbose
    )


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    import sys
    
    print("""
╔══════════════════════════════════════════════════════════════╗
║           FLOW IMAGE GENERATOR - VE3 TOOL                   ║
╠══════════════════════════════════════════════════════════════╣
║  Usage:                                                      ║
║    python flow_image_generator.py <project_path> [options]   ║
║                                                              ║
║  Options:                                                    ║
║    --characters    Generate character images only            ║
║    --scenes        Generate scene images only                ║
║    --all           Generate all (default)                    ║
║    --start N       Start from scene N                        ║
║    --end N         End at scene N                            ║
║    --overwrite     Overwrite existing images                 ║
╚══════════════════════════════════════════════════════════════╝
""")
    
    if len(sys.argv) < 2:
        print("Error: Please provide project path")
        print("Example: python flow_image_generator.py ./PROJECTS/KA1-0001")
        sys.exit(1)
    
    project_path = sys.argv[1]
    
    # Parse options
    do_characters = "--all" in sys.argv or "--characters" in sys.argv or (
        "--scenes" not in sys.argv and "--characters" not in sys.argv
    )
    do_scenes = "--all" in sys.argv or "--scenes" in sys.argv or (
        "--scenes" not in sys.argv and "--characters" not in sys.argv
    )
    overwrite = "--overwrite" in sys.argv
    
    start_scene = 1
    end_scene = None
    
    for i, arg in enumerate(sys.argv):
        if arg == "--start" and i + 1 < len(sys.argv):
            start_scene = int(sys.argv[i + 1])
        if arg == "--end" and i + 1 < len(sys.argv):
            end_scene = int(sys.argv[i + 1])
    
    # Create generator
    try:
        generator = create_generator_from_config(project_path)
        
        # Run
        results = generator.generate_all(
            characters=do_characters,
            scenes=do_scenes,
            start_scene=start_scene,
            end_scene=end_scene,
            overwrite=overwrite
        )
        
        # Exit code
        total_failed = results["characters"]["failed"] + results["scenes"]["failed"]
        sys.exit(0 if total_failed == 0 else 1)
        
    except FileNotFoundError as e:
        print(f"❌ Config file not found: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)
