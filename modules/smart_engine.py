"""
VE3 Tool - Smart Auto Engine (BROWSER JS MODE)
===============================================
Version: 2.0.0-browser
Date: 2024-12-17

Flow:
1. Voice -> SRT (Whisper)
2. SRT -> Prompts (AI)
3. Prompts -> Images (BROWSER JS - khong can API token)
4. Images -> Video (FFmpeg)
"""

__version__ = "2.0.0-browser"

import os
import json
import time
import shutil
import threading
from pathlib import Path
from typing import List, Dict, Optional, Callable, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime


# ============================================================================
# GLOBAL TOKEN EXTRACTION LOCK
# ============================================================================
# Đảm bảo chỉ 1 thread extract token tại 1 thời điểm
# Ngăn xung đột khi nhiều SmartEngine chạy song song
# ============================================================================
_global_token_lock = threading.Lock()
_token_extraction_queue = []  # Track waiting threads
_extraction_in_progress = False


def _acquire_token_extraction_slot(profile_name: str, logger=None):
    """
    Xin quyền extract token (serialize across all SmartEngine instances).

    Returns:
        True khi được phép extract
    """
    global _extraction_in_progress

    start_time = time.time()

    with _global_token_lock:
        if _extraction_in_progress:
            _token_extraction_queue.append(profile_name)
            if logger:
                queue_pos = len(_token_extraction_queue)
                logger(f"[TokenQueue] {profile_name}: Đang đợi ({queue_pos} trong hàng)...")

    # Đợi đến lượt
    while True:
        with _global_token_lock:
            if not _extraction_in_progress:
                _extraction_in_progress = True
                if profile_name in _token_extraction_queue:
                    _token_extraction_queue.remove(profile_name)
                break
        time.sleep(0.5)

    wait_time = time.time() - start_time
    if wait_time > 1 and logger:
        logger(f"[TokenQueue] {profile_name}: Đã đợi {wait_time:.1f}s")

    return True


def _release_token_extraction_slot(profile_name: str, logger=None):
    """Trả quyền extract token cho thread khác."""
    global _extraction_in_progress

    with _global_token_lock:
        _extraction_in_progress = False
        remaining = len(_token_extraction_queue)
        if remaining > 0 and logger:
            logger(f"[TokenQueue] {profile_name}: Xong! ({remaining} thread đang đợi)")


@dataclass
class Resource:
    """Tai nguyen (profile hoac API key)."""
    type: str  # 'profile', 'groq', 'gemini'
    value: str  # path hoac key
    token: str = ""
    project_id: str = ""
    status: str = "ready"  # ready, busy, exhausted, error
    fail_count: int = 0
    last_used: float = 0
    token_time: float = 0  # Thoi gian lay token (de check het han)


class SmartEngine:
    """
    Engine thong minh - dam bao output 100%.
    """

    # Token KHONG het han theo thoi gian
    # Chi refresh khi API tra loi 401 (authentication error)
    # Dieu nay toi uu hon vi token thuong valid lau hon 50 phut

    def __init__(self, config_path: str = None, assigned_profile: str = None):
        """
        Initialize SmartEngine.

        Args:
            config_path: Path to accounts.json config file
            assigned_profile: Specific Chrome profile name to use (for parallel processing)
        """
        # Support VE3_CONFIG_DIR environment variable
        if config_path:
            self.config_path = Path(config_path)
        else:
            config_dir = os.environ.get('VE3_CONFIG_DIR', 'config')
            self.config_path = Path(config_dir) / "accounts.json"

        # Config directory
        self.config_dir = self.config_path.parent

        # Token cache file
        self.tokens_path = self.config_path.parent / "tokens.json"

        self.chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

        # Assigned profile for parallel processing
        self.assigned_profile = assigned_profile

        # Resources
        self.profiles: List[Resource] = []
        self.headless_accounts: List[Resource] = []  # Headless accounts (Playwright)
        self.deepseek_keys: List[Resource] = []
        self.groq_keys: List[Resource] = []
        self.gemini_keys: List[Resource] = []

        # Ollama model (fallback when all APIs fail)
        # Default: qwen2.5:7b (fast, 32k context)
        self.ollama_model: str = "qwen2.5:7b"
        self.ollama_endpoint: str = "http://localhost:11434"

        # Settings - TOI UU TOC DO (PARALLEL OPTIMIZED)
        self.parallel = 20  # Tang len 20 - dung TAT CA tokens co san
        self.delay = 0.3    # Giam xuong 0.3s - may khoe chay nhanh
        self.max_retries = 3
        self.use_threadpool = True  # Dung ThreadPoolExecutor (hieu qua hon threading.Thread)
        self.images_per_worker = 5  # So anh moi worker xu ly truoc khi chuyen
        self.use_headless = True  # Uu tien headless mode (chay an)

        # State
        self.stop_flag = False
        self.callback = None
        self._lock = threading.Lock()

        # Parallel processing state
        self._character_gen_thread = None
        self._character_gen_result = None
        self._character_gen_error = None

        # Browser generator - reuse cho ca characters va scenes
        self._browser_generator = None

        # Cache media_name per profile: {profile_name: {image_id: media_name}}
        # QUAN TRONG: media_name chi valid cho token da tao ra no
        self.media_name_cache = {}

        # Video generation queue (parallel with image gen)
        self._video_queue = []
        self._video_queue_lock = threading.Lock()
        self._video_worker_thread = None
        self._video_worker_running = False
        self._video_results = {"success": 0, "failed": 0, "pending": 0}
        self._video_settings = {}

        self.load_config()
        self.load_cached_tokens()  # Load tokens da luu
        self.load_media_name_cache()  # Load media_name cache

    def log(self, msg: str, level: str = "INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        full_msg = f"[{ts}] [{level}] {msg}"
        if self.callback:
            self.callback(full_msg)
        else:
            # Fallback to print only when no GUI callback
            print(full_msg)

    def load_config(self):
        """Load config."""
        # === LOAD TỪ chrome_profiles/ DIRECTORY (ƯU TIÊN) ===
        # GUI tạo profiles ở đây, không phải accounts.json
        root_dir = Path(__file__).parent.parent.resolve()
        profiles_dir = root_dir / "chrome_profiles"

        if profiles_dir.exists():
            for profile_path in sorted(profiles_dir.iterdir()):
                if profile_path.is_dir() and not profile_path.name.startswith('.'):
                    self.profiles.append(Resource(
                        type='profile',
                        value=str(profile_path)
                    ))
            if self.profiles:
                self.log(f"[Config] Loaded {len(self.profiles)} profiles from chrome_profiles/")

        # === LOAD TỪ accounts.json (FALLBACK + API KEYS) ===
        if not self.config_path.exists():
            return

        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            self.chrome_path = data.get('chrome_path', self.chrome_path)

            # Load headless accounts (KHUYÊN DÙNG - chạy ẩn)
            for acc_id in data.get('headless_accounts', []):
                if acc_id and not acc_id.startswith('THAY_BANG'):
                    self.headless_accounts.append(Resource(
                        type='headless',
                        value=acc_id
                    ))

            # Load Chrome profiles từ accounts.json (bổ sung nếu chưa có)
            existing_paths = {p.value for p in self.profiles}
            for i, p in enumerate(data.get('chrome_profiles', [])):
                path = p if isinstance(p, str) else p.get('path', '')
                # Skip placeholders
                if not path or path.startswith('THAY_BANG'):
                    continue
                if Path(path).exists() and path not in existing_paths:
                    self.profiles.append(Resource(
                        type='profile',
                        value=path
                    ))

            # Load API keys
            api = data.get('api_keys', {})

            # DeepSeek (uu tien cao nhat - re va on dinh)
            for k in api.get('deepseek', []):
                if k and not k.startswith('THAY_BANG') and not k.startswith('sk-YOUR'):
                    self.deepseek_keys.append(Resource(type='deepseek', value=k))

            for k in api.get('groq', []):
                if k and not k.startswith('THAY_BANG') and not k.startswith('gsk_YOUR'):
                    self.groq_keys.append(Resource(type='groq', value=k))

            for k in api.get('gemini', []):
                if k and not k.startswith('THAY_BANG') and not k.startswith('AIzaSy_YOUR'):
                    self.gemini_keys.append(Resource(type='gemini', value=k))

            # Ollama local (fallback)
            ollama_cfg = api.get('ollama', {})
            if ollama_cfg:
                self.ollama_model = ollama_cfg.get('model', 'qwen2.5:7b')
                self.ollama_endpoint = ollama_cfg.get('endpoint', 'http://localhost:11434')

            # Settings
            settings = data.get('settings', {})
            self.parallel = settings.get('parallel', 2)
            self.delay = settings.get('delay_between_images', 2)
            self.use_headless = settings.get('use_headless', True)  # Mac dinh dung headless

        except Exception as e:
            self.log(f"Load config error: {e}", "ERROR")

    # ========== TOKEN CACHING ==========

    def load_cached_tokens(self):
        """Load tokens da luu tu file."""
        if not self.tokens_path.exists():
            return

        try:
            with open(self.tokens_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            loaded = 0

            for profile in self.profiles:
                profile_name = Path(profile.value).name
                if profile_name in data:
                    token_data = data[profile_name]

                    # KHONG check thoi gian - load tat ca tokens
                    # Chi refresh khi API tra 401
                    token = token_data.get('token', '')
                    if token:
                        profile.token = token
                        profile.project_id = token_data.get('project_id', '')
                        profile.token_time = token_data.get('time', 0)
                        profile.token_invalid = False  # Reset invalid flag
                        loaded += 1

            if loaded > 0:
                self.log(f"Loaded {loaded} cached tokens (khong check expire, chi refresh khi 401)")

        except Exception as e:
            self.log(f"Load cached tokens error: {e}", "WARN")

    # ========== MEDIA NAME CACHING ==========
    # QUAN TRONG: media_name chi valid cho token da tao ra no
    # Khi generate anh nv/loc -> luu media_name
    # Khi generate scene -> dung media_name tu cung token

    def load_media_name_cache(self):
        """Load media_name cache tu file."""
        cache_path = self.config_path.parent / "media_names.json"
        if not cache_path.exists():
            return

        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                self.media_name_cache = json.load(f)
            total = sum(len(v) for v in self.media_name_cache.values())
            if total > 0:
                self.log(f"Loaded {total} cached media_names")
        except Exception as e:
            self.log(f"Load media_name cache error: {e}", "WARN")

    def save_media_name_cache(self):
        """Luu media_name cache vao file."""
        try:
            cache_path = self.config_path.parent / "media_names.json"
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(self.media_name_cache, f, indent=2)
        except Exception as e:
            self.log(f"Save media_name cache error: {e}", "WARN")

    def get_cached_media_name(self, profile: 'Resource', image_id: str) -> str:
        """Lay media_name tu cache cho profile + image_id."""
        profile_name = Path(profile.value).name
        return self.media_name_cache.get(profile_name, {}).get(image_id, "")

    def set_cached_media_name(self, profile: 'Resource', image_id: str, media_name: str):
        """Luu media_name vao cache."""
        profile_name = Path(profile.value).name
        if profile_name not in self.media_name_cache:
            self.media_name_cache[profile_name] = {}
        self.media_name_cache[profile_name][image_id] = media_name
        self.save_media_name_cache()

    def save_cached_tokens(self):
        """Luu tokens vao file de dung lai."""
        try:
            data = {}
            for profile in self.profiles:
                if profile.token:
                    profile_name = Path(profile.value).name
                    data[profile_name] = {
                        'token': profile.token,
                        'project_id': profile.project_id,
                        'time': profile.token_time or time.time()
                    }

            self.tokens_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.tokens_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)

            self.log(f"Saved {len(data)} tokens to cache")

        except Exception as e:
            self.log(f"Save tokens error: {e}", "WARN")

    def is_token_valid(self, profile: Resource) -> bool:
        """Check token con valid khong.

        KHONG check theo thoi gian - chi refresh khi API loi 401.
        Token se duoc danh dau invalid khi API tra 401 -> trigger refresh.
        """
        if not profile.token:
            return False

        # Khong check thoi gian nua - chi refresh khi API that su loi
        # Token se duoc danh dau invalid khi API tra 401
        return not getattr(profile, 'token_invalid', False)

    def mark_token_invalid(self, profile: Resource, reason: str = "API 401"):
        """Danh dau token invalid khi API tra loi 401.

        Args:
            profile: Profile co token loi
            reason: Ly do (de log)
        """
        profile.token_invalid = True
        self.log(f"[Token] {Path(profile.value).name} bi danh dau INVALID: {reason}", "WARN")

    def refresh_token_on_error(self, profile: Resource) -> bool:
        """Refresh token khi API loi (401).

        Quan trong: Mo dung project_id cu de giu media_id da tao.

        Args:
            profile: Profile can refresh

        Returns:
            True neu refresh thanh cong
        """
        self.log(f"[Token] Refresh token cho {Path(profile.value).name} (giu project_id: {profile.project_id[:8] if profile.project_id else 'N/A'}...)")

        # Reset flag
        profile.token_invalid = False
        profile.token = ""

        # Lay token moi (se reuse project_id)
        return self.get_token_for_profile(profile)

    def get_valid_token_count(self) -> int:
        """Dem so token con valid."""
        return sum(1 for p in self.profiles if self.is_token_valid(p))

    def check_requirements(self, has_voice: bool = True) -> Tuple[bool, List[str]]:
        """
        Kiem tra thieu gi.
        BROWSER MODE: Khong can profiles trong accounts.json
        Return: (ok, list of missing items)
        """
        missing = []

        # BROWSER MODE: Khong can profiles - chi can Chrome duoc cai dat
        # if not self.profiles:
        #     missing.append("Chrome profiles (sua config/accounts.json)")

        if has_voice and not self.deepseek_keys and not self.groq_keys and not self.gemini_keys:
            missing.append("AI API keys cho voice (DeepSeek RE: platform.deepseek.com/api_keys)")

        if not Path(self.chrome_path).exists():
            missing.append(f"Chrome: {self.chrome_path}")

        return len(missing) == 0, missing

    # ========== RESOURCE MANAGEMENT ==========

    def get_available_profile(self) -> Optional[Resource]:
        """Lay profile san sang."""
        with self._lock:
            for p in self.profiles:
                if p.status == 'ready' and p.token:
                    return p
            # Neu khong co token, tra ve profile dau tien de lay token
            for p in self.profiles:
                if p.status in ['ready', 'error'] and p.fail_count < self.max_retries:
                    return p
        return None

    def get_available_ai_key(self) -> Optional[Resource]:
        """Lay AI key san sang (uu tien DeepSeek vi re va on dinh)."""
        with self._lock:
            # Uu tien DeepSeek (re nhat, on dinh)
            for k in self.deepseek_keys:
                if k.status == 'ready' and k.fail_count < self.max_retries:
                    return k
            # Fallback Groq (free nhung hay rate limit)
            for k in self.groq_keys:
                if k.status == 'ready' and k.fail_count < self.max_retries:
                    return k
            # Fallback Gemini
            for k in self.gemini_keys:
                if k.status == 'ready' and k.fail_count < self.max_retries:
                    return k
        return None

    def mark_resource_used(self, res: Resource, success: bool):
        """Danh dau resource da dung."""
        with self._lock:
            res.last_used = time.time()
            if success:
                res.fail_count = 0
                res.status = 'ready'
            else:
                res.fail_count += 1
                if res.fail_count >= self.max_retries:
                    res.status = 'exhausted'
                    self.log(f"Resource exhausted: {res.type} - {res.value[:20]}...", "WARN")

    def reset_resources(self):
        """Reset tat ca resources."""
        with self._lock:
            for r in self.profiles + self.deepseek_keys + self.groq_keys + self.gemini_keys:
                r.status = 'ready'
                r.fail_count = 0

    # ========== TOKEN MANAGEMENT ==========

    def get_token_headless(self, account: Resource) -> bool:
        """
        Lay token bang HEADLESS mode (chay an).
        Khong mo browser window.

        QUAN TRONG:
        - Reuse project_id da co de share media_ids giua nv va img.
        - Sử dụng global token queue để tránh xung đột.
        """
        try:
            from modules.headless_token import HeadlessTokenExtractor
        except ImportError:
            self.log("Chua cai Playwright! Chay: pip install playwright && playwright install chromium", "ERROR")
            return False

        account_name = account.value
        self.log(f"[Headless] Lay token: {account_name}...")

        # Log project_id status
        if account.project_id:
            self.log(f"  -> Reuse project_id: {account.project_id[:8]}...")
        else:
            self.log(f"  -> Chua co project_id, se tao moi")

        # === ACQUIRE TOKEN EXTRACTION SLOT (Queue) ===
        _acquire_token_extraction_slot(f"headless_{account_name}", self.log)

        try:
            extractor = HeadlessTokenExtractor(account.value)

            # Check co cookies chua
            if not extractor.has_valid_cookies():
                self.log(f"[Headless] {account_name} chua dang nhap!", "WARN")
                self.log(f"  -> Chay: python -m modules.headless_token login {account_name}", "WARN")
                return False

            # Lay token (headless) - truyen project_id de reuse
            token, proj_id, error = extractor.extract_token(
                headless=True,
                project_id=account.project_id  # Reuse existing project
            )

            if token:
                account.token = token
                # Chi update project_id neu co gia tri moi
                if proj_id:
                    account.project_id = proj_id
                account.token_time = time.time()
                account.status = 'ready'
                self.log(f"[Headless] OK: {account_name} - Token OK!", "OK")
                self.log(f"  -> Project ID: {account.project_id[:8]}..." if account.project_id else "  -> No project ID")
                self.save_cached_tokens()
                return True
            elif error == "need_login":
                self.log(f"[Headless] {account_name} can dang nhap lai!", "WARN")
                self.log(f"  -> Chay: python -m modules.headless_token login {account_name}", "WARN")
                return False
            else:
                account.fail_count += 1
                self.log(f"[Headless] FAIL: {account_name} - {error}", "ERROR")
                return False

        except Exception as e:
            account.fail_count += 1
            self.log(f"[Headless] ERROR: {e}", "ERROR")
            return False

        finally:
            # === RELEASE TOKEN EXTRACTION SLOT ===
            _release_token_extraction_slot(f"headless_{account_name}", self.log)

    def get_token_for_profile(self, profile: Resource) -> bool:
        """Lay token cho 1 profile (Chrome visible).

        QUAN TRONG:
        - Reuse project_id da co de share media_ids giua nv va img.
        - Sử dụng global token queue để tránh xung đột khi nhiều thread extract cùng lúc.
        """
        from modules.auto_token import ChromeAutoToken

        profile_name = Path(profile.value).name
        self.log(f"[Chrome] Lay token: {profile_name}...")

        # Log project_id status
        if profile.project_id:
            self.log(f"  -> Reuse project_id: {profile.project_id[:8]}...")
        else:
            self.log(f"  -> Chua co project_id, se tao moi")

        # === ACQUIRE TOKEN EXTRACTION SLOT (Queue) ===
        # Đảm bảo chỉ 1 Chrome extract token tại 1 thời điểm
        _acquire_token_extraction_slot(profile_name, self.log)

        try:
            # QUAN TRONG: Lay token LUON phai chay Chrome HIEN THI
            # Vi Google Flow detect headless mode va block!
            # Headless chi dung cho TAO ANH, khong dung cho LAY TOKEN
            self.log(f"[Chrome] ⚠️ Mo Chrome HIEN THI de lay token (Google block headless)")

            extractor = ChromeAutoToken(
                chrome_path=self.chrome_path,
                profile_path=profile.value,
                auto_close=True,  # Tu dong dong Chrome sau khi lay token
                headless=False  # LUON False - Google block headless khi lay token
            )

            # QUAN TRONG: Truyen project_id de reuse project da co
            # Neu co project_id -> mo project do de lay token moi (khong tao project moi)
            # Neu chua co -> tao project moi
            token, proj_id, error = extractor.extract_token(
                project_id=profile.project_id,  # Reuse existing project
                callback=self.callback
            )

            if token:
                profile.token = token
                # Chi update project_id neu co gia tri moi
                if proj_id:
                    profile.project_id = proj_id
                profile.token_time = time.time()  # Luu thoi gian lay token
                profile.status = 'ready'
                self.log(f"[Chrome] OK: {profile_name} - Token OK!", "OK")
                self.log(f"  -> Project ID: {profile.project_id[:8]}..." if profile.project_id else "  -> No project ID")
                self.save_cached_tokens()  # Luu ngay vao file
                return True
            else:
                profile.fail_count += 1
                self.log(f"[Chrome] FAIL: {profile_name} - {error}", "ERROR")
                return False

        except Exception as e:
            profile.fail_count += 1
            self.log(f"[Chrome] ERROR: {e}", "ERROR")
            return False

        finally:
            # === RELEASE TOKEN EXTRACTION SLOT ===
            _release_token_extraction_slot(profile_name, self.log)

    def _prefetch_token_for_worker(self, profile_path: str, worker_id: int) -> Optional[Dict]:
        """
        Pre-fetch token cho 1 worker (parallel batch processing).

        QUAN TRỌNG:
        - Mỗi worker cần 1 project RIÊNG (không reuse) vì media_ids unique per project
        - Token extraction phải serialized (1 Chrome tại 1 thời điểm)
        - Return dict với token + project_id để worker dùng

        Args:
            profile_path: Đường dẫn Chrome profile
            worker_id: ID của worker (0, 1, 2, ...)

        Returns:
            Dict {'token': str, 'project_id': str} hoặc None nếu fail
        """
        from modules.auto_token import ChromeAutoToken

        profile_name = Path(profile_path).name
        self.log(f"[Worker{worker_id}] Pre-fetch token từ profile: {profile_name}")

        # === ACQUIRE TOKEN EXTRACTION SLOT (Queue) ===
        # Đảm bảo chỉ 1 Chrome extract token tại 1 thời điểm
        _acquire_token_extraction_slot(f"worker_{worker_id}", self.log)

        try:
            # QUAN TRỌNG: Chrome HIỂN THỊ vì Google block headless
            self.log(f"[Worker{worker_id}] Mở Chrome HIỂN THỊ để lấy token...")

            extractor = ChromeAutoToken(
                chrome_path=self.chrome_path,
                profile_path=profile_path,
                auto_close=True,  # Tự động đóng sau khi lấy token
                headless=False  # LUÔN False - Google block headless
            )

            # QUAN TRỌNG: project_id=None để TẠO MỚI project
            # Mỗi worker cần project riêng vì media_ids unique per project
            token, project_id, error = extractor.extract_token(
                project_id=None,  # TẠO MỚI - không reuse
                callback=self.callback
            )

            if token:
                self.log(f"[Worker{worker_id}] ✅ Token OK!", "OK")
                if project_id:
                    self.log(f"[Worker{worker_id}]   Project: {project_id[:8]}...")

                return {
                    'token': token,
                    'project_id': project_id,
                    'profile_path': profile_path,
                    'worker_id': worker_id,
                    'timestamp': time.time()
                }
            else:
                self.log(f"[Worker{worker_id}] ❌ Token FAIL: {error}", "ERROR")
                return None

        except Exception as e:
            self.log(f"[Worker{worker_id}] ❌ Exception: {e}", "ERROR")
            return None

        finally:
            # === RELEASE TOKEN EXTRACTION SLOT ===
            _release_token_extraction_slot(f"worker_{worker_id}", self.log)

    def get_all_tokens(self) -> int:
        """
        Lay token cho tat ca accounts.
        Uu tien: Headless (an) > Chrome profiles (visible)
        """
        success = 0
        total = len(self.headless_accounts) + len(self.profiles)

        if total == 0:
            self.log("Khong co account nao! Cau hinh config/accounts.json", "WARN")
            return 0

        # 1. HEADLESS ACCOUNTS (uu tien - chay AN)
        if self.use_headless and self.headless_accounts:
            self.log(f"=== LAY TOKEN HEADLESS ({len(self.headless_accounts)} accounts) ===")

            for i, account in enumerate(self.headless_accounts):
                if self.stop_flag:
                    break

                self.log(f"[{i+1}/{len(self.headless_accounts)}] {account.value}")

                if self.is_token_valid(account):
                    self.log(f"  -> Da co token valid, skip")
                    success += 1
                    continue

                if self.get_token_headless(account):
                    success += 1

                # Delay nho giua cac accounts
                if i < len(self.headless_accounts) - 1:
                    time.sleep(0.5)

        # 2. CHROME PROFILES (backup - khi headless khong du)
        if self.profiles and (not self.use_headless or success < len(self.headless_accounts)):
            self.log(f"=== LAY TOKEN CHROME ({len(self.profiles)} profiles) ===")

            for i, profile in enumerate(self.profiles):
                if self.stop_flag:
                    break

                self.log(f"[{i+1}/{len(self.profiles)}] {Path(profile.value).name}")

                if self.is_token_valid(profile):
                    self.log(f"  -> Da co token valid, skip")
                    success += 1
                    continue

                if profile.token:
                    self.log(f"  -> Token het han, lay moi...")

                if self.get_token_for_profile(profile):
                    success += 1

                if i < len(self.profiles) - 1:
                    time.sleep(1)

        self.log(f"=== XONG: {success}/{total} tokens ===")
        return success

    def refresh_token_if_needed(self, profile: Resource) -> bool:
        """Refresh token neu can - check expiry."""
        if not self.is_token_valid(profile):
            return self.get_token_for_profile(profile)
        return True

    def _get_other_valid_profile(self, exclude_profile: Resource) -> Optional[Resource]:
        """Tim profile khac co token con valid."""
        with self._lock:
            for p in self.profiles:
                if p != exclude_profile and p.token and self.is_token_valid(p):
                    return p
        return None

    def refresh_expired_tokens(self) -> int:
        """
        Kiem tra va refresh cac token bi danh dau INVALID.
        Chi chay khi co token bi mark invalid (do API 401).

        Returns:
            So token da refresh thanh cong
        """
        refreshed = 0
        invalid_profiles = []

        # Tim cac profile co token bi mark invalid
        for profile in self.profiles:
            if profile.token and not self.is_token_valid(profile):
                invalid_profiles.append(profile)
                profile.token = ""  # Clear invalid token

        if not invalid_profiles:
            return 0

        self.log(f"Tim thay {len(invalid_profiles)} token INVALID, dang refresh...")

        for profile in invalid_profiles:
            if self.stop_flag:
                break
            if self.get_token_for_profile(profile):
                refreshed += 1

        self.log(f"Da refresh {refreshed}/{len(invalid_profiles)} tokens")
        return refreshed

    # ========== SRT PROCESSING ==========

    def make_srt(self, voice_path: Path, srt_path: Path) -> bool:
        """Tao SRT tu voice."""
        if srt_path.exists():
            self.log(f"SRT da ton tai: {srt_path.name}")
            return True

        self.log("Transcribe voice -> SRT...")

        try:
            from modules.voice_to_srt import VoiceToSrt
            conv = VoiceToSrt(model_name="base", language="vi")
            conv.transcribe(voice_path, srt_path)
            self.log(f"OK: {srt_path.name}", "OK")
            return True
        except Exception as e:
            self.log(f"SRT error: {e}", "ERROR")
            return False

    # ========== PROMPT GENERATION ==========

    def make_prompts(self, proj_dir: Path, name: str, excel_path: Path) -> bool:
        """Tao prompts tu SRT. Retry voi cac AI keys khac neu fail."""
        if excel_path.exists():
            self.log(f"Prompts da ton tai: {excel_path.name}")
            return True

        self.log("Generate prompts...")

        # Load config
        import yaml
        cfg = {}
        cfg_file = Path("config/settings.yaml")
        if cfg_file.exists():
            with open(cfg_file, "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f) or {}

        # Add API keys (thu tu uu tien: Gemini > Groq > DeepSeek > Ollama)
        cfg['gemini_api_keys'] = [k.value for k in self.gemini_keys if k.status != 'exhausted']
        cfg['groq_api_keys'] = [k.value for k in self.groq_keys if k.status != 'exhausted']
        cfg['deepseek_api_keys'] = [k.value for k in self.deepseek_keys if k.status != 'exhausted']
        cfg['preferred_provider'] = 'gemini' if self.gemini_keys else ('groq' if self.groq_keys else 'deepseek')

        # Ollama local model (fallback khi tat ca API fail)
        cfg['ollama_model'] = self.ollama_model
        cfg['ollama_endpoint'] = self.ollama_endpoint

        # Retry with different keys
        for attempt in range(self.max_retries):
            if self.stop_flag:
                return False

            ai_key = self.get_available_ai_key()
            if not ai_key:
                self.log("Het AI keys!", "ERROR")
                return False

            self.log(f"Thu AI key: {ai_key.type} (attempt {attempt+1})")

            try:
                from modules.prompts_generator import PromptGenerator
                gen = PromptGenerator(cfg)

                # Pass callback de bat dau character generation song song
                if gen.generate_for_project(
                    proj_dir, name,
                    on_characters_ready=lambda ep, pd: self._on_characters_ready(ep, pd)
                ):
                    self.mark_resource_used(ai_key, True)
                    self.log(f"OK: {excel_path.name}", "OK")
                    return True
                else:
                    self.mark_resource_used(ai_key, False)

            except Exception as e:
                self.log(f"Prompt error: {e}", "ERROR")
                self.mark_resource_used(ai_key, False)

        return False

    # ========== PARALLEL CHARACTER GENERATION ==========

    def _load_character_prompts(self, excel_path: Path, proj_dir: Path) -> List[Dict]:
        """Load CHI character prompts (nv*, loc*) tu Excel - cho parallel generation."""
        import openpyxl

        prompts = []
        wb = openpyxl.load_workbook(excel_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            if not headers or all(h is None for h in headers):
                continue

            # Tim cot ID va Prompt
            id_col = None
            prompt_col = None

            for i, h in enumerate(headers):
                if h is None:
                    continue
                h_lower = str(h).lower().strip()

                if id_col is None and ('id' in h_lower):
                    id_col = i
                if 'english' in h_lower and 'prompt' in h_lower:
                    prompt_col = i
                elif prompt_col is None and h_lower == 'img_prompt':
                    prompt_col = i
                elif prompt_col is None and 'prompt' in h_lower and 'video' not in h_lower:
                    prompt_col = i

            if id_col is None or prompt_col is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or id_col >= len(row) or prompt_col >= len(row):
                    continue

                pid = row[id_col]
                prompt = row[prompt_col]

                if not pid or not prompt:
                    continue

                pid_str = str(pid).strip()
                prompt_str = str(prompt).strip()

                # Skip children (DO_NOT_GENERATE)
                if prompt_str == "DO_NOT_GENERATE":
                    continue

                # CHI lay character/location prompts (nv*, loc*)
                if pid_str.startswith('nv') or pid_str.startswith('loc'):
                    out_path = proj_dir / "nv" / f"{pid_str}.png"
                    if not out_path.exists():  # Chi lay chua co anh
                        prompts.append({
                            'id': pid_str,
                            'prompt': prompt_str,
                            'output_path': str(out_path),
                            'reference_files': "",
                            'nv_path': str(proj_dir / "nv")
                        })

        return prompts

    def _generate_characters_async(self, excel_path: Path, proj_dir: Path):
        """
        Generate character images trong background thread.
        Duoc goi tu callback khi characters ready.
        Respect generation_mode setting (api hoac chrome).
        """
        def _worker():
            try:
                self.log("[PARALLEL] Bat dau tao anh nhan vat (background)...")

                # Load character prompts
                char_prompts = self._load_character_prompts(excel_path, proj_dir)

                if not char_prompts:
                    self.log("[PARALLEL] Khong co character prompts moi can tao")
                    self._character_gen_result = {"success": 0, "failed": 0}
                    return

                self.log(f"[PARALLEL] Tao {len(char_prompts)} anh nhan vat...")

                # Check generation_mode setting
                generation_mode = 'api'  # Default
                try:
                    import yaml
                    settings_path = self.config_dir / "settings.yaml"
                    if settings_path.exists():
                        with open(settings_path, 'r', encoding='utf-8') as f:
                            settings = yaml.safe_load(f) or {}
                        generation_mode = settings.get('generation_mode', 'api')
                except:
                    pass

                # Generate using correct mode
                if generation_mode == 'api':
                    self.log("[PARALLEL] Dung API MODE cho characters...")
                    results = self.generate_images_api(char_prompts, proj_dir)
                else:
                    self.log("[PARALLEL] Dung BROWSER MODE cho characters...")
                    results = self.generate_images_browser(char_prompts, proj_dir)

                self._character_gen_result = results
                self.log(f"[PARALLEL] Xong! Success={results.get('success', 0)}, Failed={results.get('failed', 0)}")

            except Exception as e:
                self._character_gen_error = str(e)
                self.log(f"[PARALLEL] Loi: {e}", "ERROR")

        # Start thread
        self._character_gen_thread = threading.Thread(target=_worker, daemon=True)
        self._character_gen_thread.start()
        self.log("[PARALLEL] Character generation thread started!")

    def _wait_for_character_generation(self, timeout: int = 600) -> bool:
        """
        Doi character generation thread hoan thanh.
        Returns True neu thanh cong, False neu loi hoac timeout.
        """
        if self._character_gen_thread is None:
            return True  # Khong co thread nao dang chay

        self.log("[PARALLEL] Doi character generation hoan thanh...")
        self._character_gen_thread.join(timeout=timeout)

        if self._character_gen_thread.is_alive():
            self.log("[PARALLEL] Character generation timeout!", "WARN")
            return False

        if self._character_gen_error:
            self.log(f"[PARALLEL] Character generation error: {self._character_gen_error}", "ERROR")
            return False

        return True

    def _on_characters_ready(self, excel_path: Path, proj_dir: Path):
        """
        Callback duoc goi khi characters prompts da duoc save.
        Bat dau generate character images song song.
        """
        self._generate_characters_async(excel_path, proj_dir)

    # ========== IMAGE GENERATION ==========

    def _sanitize_prompt(self, prompt: str) -> str:
        """
        Điều chỉnh prompt để tránh vi phạm policy.
        Loại bỏ các từ/cụm từ nhạy cảm và debug tags.
        """
        import re

        # === LOẠI BỎ DEBUG TAGS ===
        # Các tag như [FALLBACK], [DEBUG], [TEST] không nên gửi đến API
        prompt = re.sub(r'\[FALLBACK\]\s*', '', prompt, flags=re.IGNORECASE)
        prompt = re.sub(r'\[DEBUG\]\s*', '', prompt, flags=re.IGNORECASE)
        prompt = re.sub(r'\[TEST\]\s*', '', prompt, flags=re.IGNORECASE)
        prompt = re.sub(r'\[TIER\s*\d+\]\s*', '', prompt, flags=re.IGNORECASE)

        # Các từ/cụm từ cần loại bỏ hoặc thay thế
        sensitive_words = [
            # Violence
            (r'\b(kill|murder|blood|gore|violent|weapon|gun|knife|death|dead|corpse)\b', ''),
            (r'\b(fight|attack|hurt|injure|wound)\b', 'interact'),
            # Adult content
            (r'\b(naked|nude|sexy|seductive|erotic|sensual)\b', ''),
            (r'\b(kiss|embrace|romantic)\b', 'close'),
            # Sensitive topics
            (r'\b(drugs|alcohol|smoking|cigarette)\b', ''),
            (r'\b(scared|terrified|horror|scary)\b', 'surprised'),
            # Simplify complex scenes
            (r'\b(crying|tears|sad|depressed)\b', 'emotional'),
        ]

        sanitized = prompt
        for pattern, replacement in sensitive_words:
            sanitized = re.sub(pattern, replacement, sanitized, flags=re.IGNORECASE)

        # Loại bỏ khoảng trắng thừa
        sanitized = re.sub(r'\s+', ' ', sanitized).strip()

        return sanitized

    def _simplify_prompt(self, prompt: str) -> str:
        """
        Đơn giản hóa prompt khi các retry khác thất bại.
        Giữ lại phần cơ bản nhất: shot type + subject + action.
        """
        import re

        # Loại bỏ các phần phức tạp
        simplified = prompt

        # Bỏ phần style dài ở cuối (Cinematic, 4K, ...)
        simplified = re.sub(r'Cinematic.*$', '', simplified, flags=re.IGNORECASE)
        simplified = re.sub(r'4K.*$', '', simplified, flags=re.IGNORECASE)
        simplified = re.sub(r'photorealistic.*$', '', simplified, flags=re.IGNORECASE)

        # Bỏ các chi tiết về lens, camera
        simplified = re.sub(r'\d+mm lens[,.]?\s*', '', simplified, flags=re.IGNORECASE)
        simplified = re.sub(r'shot on [^,]+,?\s*', '', simplified, flags=re.IGNORECASE)

        # Bỏ các chi tiết về lighting
        simplified = re.sub(r'(soft|warm|cold|dramatic|natural) (light|lighting)[^,]*,?\s*', '', simplified, flags=re.IGNORECASE)

        # Bỏ các emotion words có thể gây policy violation
        simplified = re.sub(r'\b(devastated|terrified|horrified|anguished|tormented)\b', 'thoughtful', simplified, flags=re.IGNORECASE)

        # Giữ lại tối đa 150 ký tự (phần đầu quan trọng nhất)
        if len(simplified) > 150:
            simplified = simplified[:150].rsplit(' ', 1)[0]

        # Thêm style đơn giản
        simplified = simplified.strip().rstrip(',.')
        simplified += ". High quality, professional photograph."

        return simplified

    def generate_single_image(self, prompt_data: Dict, profile: Resource, retry_count: int = 0) -> tuple:
        """
        Tao 1 anh voi 1 profile, ho tro reference images.

        FLOW QUAN TRONG:
        1. Neu la nv*/loc* -> generate va LUU media_name vao cache
        2. Neu la scene -> dung media_name tu cache (cung profile/token)
           Neu khong co trong cache -> upload de lay media_name moi

        Returns:
            tuple: (success: bool, token_expired: bool)
        """
        from modules.google_flow_api import GoogleFlowAPI, AspectRatio, ImageInput

        pid = prompt_data.get('id', '')
        prompt = prompt_data.get('prompt', '')
        output = prompt_data.get('output_path', '')
        reference_files = prompt_data.get('reference_files', '')
        nv_path = prompt_data.get('nv_path', '')

        if not prompt or not output:
            return False, False

        # Skip if exists
        if Path(output).exists():
            return True, False

        is_reference_image = pid.startswith('nv') or pid.startswith('loc')

        try:
            # Bat verbose cho nv/loc de debug media_name
            api = GoogleFlowAPI(
                bearer_token=profile.token,
                project_id=profile.project_id,
                verbose=is_reference_image  # Verbose for reference images to debug
            )

            Path(output).parent.mkdir(parents=True, exist_ok=True)

            # === SCENE IMAGES: Su dung reference images ===
            image_inputs = []
            if nv_path and not is_reference_image:
                # Parse reference_files (JSON array or comma-separated)
                file_list = []
                if reference_files:
                    try:
                        parsed = json.loads(reference_files)
                        if isinstance(parsed, list):
                            file_list = parsed
                        elif isinstance(parsed, str):
                            file_list = [parsed]
                    except (json.JSONDecodeError, TypeError):
                        file_list = [f.strip() for f in str(reference_files).split(",") if f.strip()]

                # FALLBACK: Dam bao LUON co nhan vat trong reference
                has_character = any(f.lower().startswith('nv') for f in file_list)

                if not file_list:
                    # Khong co reference nao -> dung nvc
                    file_list = ["nvc.png"]
                    self.log(f"  -> No reference, using default nvc.png")
                elif not has_character:
                    # Chi co loc, khong co nhan vat -> them nvc vao dau
                    file_list.insert(0, "nvc.png")
                    self.log(f"  -> No character in refs, adding nvc.png → {file_list}")

                # Tim media_name cho moi reference image
                # LUU Y: API CHI CHAP NHAN media_name, KHONG chap nhan base64!
                skipped_refs = []
                for filename in file_list:
                    # Extract image_id tu filename (vd: "nv1.png" -> "nv1")
                    image_id = Path(filename).stem

                    # CHI dung cached media_name - base64 KHONG hoat dong
                    cached_media_name = self.get_cached_media_name(profile, image_id)

                    if cached_media_name:
                        # API chi chap nhan REFERENCE type
                        # SUBJECT/STYLE khong duoc ho tro
                        image_inputs.append(ImageInput(name=cached_media_name))
                        self.log(f"  -> Ref OK: {image_id}")
                    else:
                        # Khong co media_name -> SKIP (khong the dung base64)
                        skipped_refs.append(image_id)

                if skipped_refs:
                    self.log(f"  -> SKIP refs (no media_name): {skipped_refs}", "WARN")
                    self.log(f"  -> Tao anh KHONG CO reference (chua co media_name)", "WARN")
                    # Clear image_inputs neu co bat ky ref nao thieu
                    # Vi API yeu cau TAT CA refs phai co media_name
                    image_inputs = []

                if image_inputs:
                    self.log(f"  -> Using {len(image_inputs)} reference images for {pid}")

            # === GENERATE IMAGE ===
            success, images, error = api.generate_images(
                prompt=prompt,
                count=1,
                aspect_ratio=AspectRatio.LANDSCAPE,
                image_inputs=image_inputs if image_inputs else None
            )

            if success and images:
                # === DEBUG: Xem API tra ve gi ===
                img = images[0]
                self.log(f"  -> DEBUG: media_name={img.media_name}, media_id={img.media_id}, workflow_id={img.workflow_id}")

                # === LUU MEDIA_NAME neu la nv/loc ===
                if is_reference_image:
                    # Thu lay media_name, fallback to workflow_id or media_id
                    ref_id = img.media_name or img.workflow_id or img.media_id
                    if ref_id:
                        self.set_cached_media_name(profile, pid, ref_id)
                        self.log(f"  -> Saved ref_id for {pid}: {ref_id[:40]}...")
                    else:
                        self.log(f"  -> WARNING: No identifier returned for {pid}!", "WARN")
                        self.log(f"  -> Available: media_name={img.media_name}, workflow_id={img.workflow_id}, media_id={img.media_id}", "DEBUG")

                # Download image
                downloaded = api.download_image(images[0], Path(output).parent, pid)
                if downloaded:
                    # Rename to correct filename if needed
                    final_path = Path(output)
                    if downloaded.exists() and str(downloaded) != output:
                        if final_path.exists():
                            final_path.unlink()
                        downloaded.rename(output)

                    # Queue video generation if enabled (parallel)
                    video_prompt = prompt_data.get('video_prompt', '')
                    self._queue_video_generation(final_path, pid, video_prompt)

                    return True, False
                else:
                    self.log(f"Download failed {pid}", "ERROR")
                    return False, False
            else:
                # Check if token expired (API 401)
                error_str = str(error).lower()
                token_expired = 'expired' in error_str or 'unauthorized' in error_str or '401' in error_str or 'authentication' in error_str
                if token_expired:
                    self.log(f"Token het han cho {pid} (API 401), thu refresh...", "WARN")
                    self.mark_token_invalid(profile, f"API 401 - {pid}")

                    # Thu refresh ngay va retry
                    if retry_count < 2 and self.refresh_token_on_error(profile):
                        self.log(f"  -> Refresh OK, retry {pid}...", "OK")
                        return self.generate_single_image(prompt_data, profile, retry_count + 1)

                    return False, token_expired

                # Check for 403 Forbidden - could be rate limit or account issue
                is_forbidden = '403' in error_str or 'forbidden' in error_str
                if is_forbidden:
                    self.log(f"403 Forbidden cho {pid} - co the do rate limit hoac account bi han che", "WARN")
                    self.log(f"  -> Thu chuyen sang account khac hoac doi 30s...", "WARN")
                    # Mark this profile as needing refresh (might need re-login)
                    profile.token = ""  # Force re-login
                    return False, True  # Return token_expired=True to trigger account switch

                # Check for policy violation or invalid argument - retry with fixes
                is_policy_error = '400' in error_str or 'policy' in error_str or 'blocked' in error_str or 'safety' in error_str or 'invalid' in error_str
                if is_policy_error and retry_count < 3:
                    prompt_data_copy = prompt_data.copy()

                    if retry_count == 0:
                        # Retry 1: Sanitize prompt
                        self.log(f"  -> Error 400, thu lai voi prompt da dieu chinh...", "WARN")
                        sanitized = self._sanitize_prompt(prompt)
                        if sanitized != prompt:
                            prompt_data_copy['prompt'] = sanitized
                            return self.generate_single_image(prompt_data_copy, profile, retry_count + 1)

                    if retry_count <= 1:
                        # Retry 2: Remove reference images (might be expired/invalid)
                        self.log(f"  -> Thu lai KHONG co reference images...", "WARN")
                        prompt_data_copy['reference_images'] = []
                        prompt_data_copy['prompt'] = self._sanitize_prompt(prompt)
                        return self.generate_single_image(prompt_data_copy, profile, retry_count + 1)

                    if retry_count == 2:
                        # Retry 3: Simplify prompt completely
                        self.log(f"  -> Thu lai voi prompt don gian...", "WARN")
                        simple_prompt = self._simplify_prompt(prompt)
                        prompt_data_copy['prompt'] = simple_prompt
                        prompt_data_copy['reference_images'] = []
                        return self.generate_single_image(prompt_data_copy, profile, retry_count + 1)

                self.log(f"Generate failed {pid}: {error}", "ERROR")
                return False, False

        except Exception as e:
            error_str = str(e).lower()
            token_expired = 'expired' in error_str or 'unauthorized' in error_str or '401' in error_str or 'authentication' in error_str
            if token_expired:
                self.log(f"Token het han cho {pid} (Exception), thu refresh...", "WARN")
                self.mark_token_invalid(profile, f"Exception 401 - {pid}")

                # Thu refresh ngay va retry
                if retry_count < 2 and self.refresh_token_on_error(profile):
                    self.log(f"  -> Refresh OK, retry {pid}...", "OK")
                    return self.generate_single_image(prompt_data, profile, retry_count + 1)
            else:
                self.log(f"Image error {pid}: {e}", "ERROR")
            return False, token_expired

    def generate_images_api(self, prompts: List[Dict], proj_dir: Path) -> Dict:
        """
        Tao anh bang API MODE (nhanh hon, khong can mo browser).
        Su dung Proxy API de bypass captcha.

        Flow:
        1. Lay bearer token tu Chrome profile (1 lan)
        2. Goi Proxy API de tao anh
        3. Download anh ve local

        Args:
            prompts: List prompts [{'id': '1', 'prompt': '...', 'output_path': '...'}]
            proj_dir: Thu muc project

        Returns:
            Dict {"success": int, "failed": int}
        """
        self.log("=== TAO ANH BANG API MODE ===")

        # Tim Excel file
        excel_files = list((proj_dir / "prompts").glob("*_prompts.xlsx"))
        if not excel_files:
            self.log("Khong tim thay file Excel!", "ERROR")
            return {"success": 0, "failed": len(prompts)}

        try:
            from modules.browser_flow_generator import BrowserFlowGenerator
        except ImportError as e:
            self.log(f"Khong import duoc BrowserFlowGenerator: {e}", "ERROR")
            return {"success": 0, "failed": len(prompts)}

        # Load settings
        import yaml
        headless = True
        try:
            config_path = self.config_dir / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    settings = yaml.safe_load(f) or {}
                headless = settings.get('browser_headless', True)
        except:
            pass

        # Tim profile co san
        profile_name = "main"
        if self.assigned_profile:
            profile_name = self.assigned_profile
        elif self.profiles:
            profile_name = Path(self.profiles[0].value).name

        self.log(f"API mode voi profile: {profile_name}")

        # Config path - dung settings.yaml, khong phai accounts.json
        settings_path = self.config_dir / "settings.yaml"

        # Check bearer token - ưu tiên pre-fetched token (parallel batch mode)
        bearer_token = ""
        proxy_token = ""
        prefetched_project_id = None

        # === CHECK PRE-FETCHED TOKEN (parallel batch mode) ===
        if hasattr(self, '_prefetched_token') and self._prefetched_token:
            prefetch = self._prefetched_token
            bearer_token = prefetch.get('token', '')
            prefetched_project_id = prefetch.get('project_id', '')
            worker_id = prefetch.get('worker_id', -1)
            if bearer_token:
                self.log(f"[Worker{worker_id}] Dùng pre-fetched token (project: {prefetched_project_id[:8] if prefetched_project_id else 'N/A'}...)")

        # Load settings.yaml for proxy_token và fallback
        try:
            with open(settings_path, 'r', encoding='utf-8') as f:
                cfg = yaml.safe_load(f) or {}
            proxy_token = cfg.get('proxy_api_token', '')
            # Only use settings.yaml bearer_token if no pre-fetched token
            if not bearer_token:
                bearer_token = cfg.get('flow_bearer_token', '')
        except:
            pass

        if not proxy_token:
            self.log("THIEU proxy_api_token trong settings.yaml!", "ERROR")
            self.log("API mode can proxy token tu nanoai.pics", "ERROR")
            return {"success": 0, "failed": len(prompts)}

        if not bearer_token:
            self.log("Chua co flow_bearer_token, se tu dong lay...", "WARN")
            self.log("Neu loi Chrome, hay mo https://labs.google/fx/vi/tools/flow", "WARN")
            self.log("va copy bearer token vao settings.yaml (flow_bearer_token)", "WARN")

        try:
            # Tao BrowserFlowGenerator
            generator = BrowserFlowGenerator(
                project_path=str(proj_dir),
                profile_name=profile_name,
                headless=headless,
                verbose=True,
                config_path=str(settings_path)
            )

            # === INJECT PRE-FETCHED PROJECT_ID nếu có ===
            # Quan trọng: mỗi worker cần project riêng để tránh conflict media_ids
            if prefetched_project_id:
                generator.config['flow_project_id'] = prefetched_project_id
                self.log(f"  -> Inject project_id: {prefetched_project_id[:8]}...")

            # Goi generate_from_prompts_auto - tu dong chon API hoac Chrome
            result = generator.generate_from_prompts_auto(
                prompts=prompts,
                excel_path=excel_files[0],
                bearer_token=bearer_token if bearer_token else None
            )

            if result.get("success") == False:
                error_msg = result.get('error', '')
                self.log(f"API mode error: {error_msg}", "ERROR")
                if 'Chrome' in error_msg or 'session' in error_msg:
                    self.log("=== HUONG DAN LAY TOKEN THU CONG ===", "WARN")
                    self.log("1. Mo https://labs.google/fx/vi/tools/flow trong Chrome", "WARN")
                    self.log("2. Tao 1 anh bat ky de trigger API", "WARN")
                    self.log("3. Mo DevTools (F12) > Network > tim request 'batchGenerateImages'", "WARN")
                    self.log("4. Copy 'Authorization: Bearer ya29.xxx...' vao settings.yaml", "WARN")
                return {"success": 0, "failed": len(prompts)}

            # Parse results
            success_count = result.get("success", 0)
            failed_count = result.get("failed", 0)

            self.log(f"API mode: {success_count} thanh cong, {failed_count} that bai")
            return {"success": success_count, "failed": failed_count}

        except Exception as e:
            self.log(f"API mode exception: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return {"success": 0, "failed": len(prompts)}

    def generate_images_browser(self, prompts: List[Dict], proj_dir: Path) -> Dict:
        """
        Tao anh bang BROWSER AUTOMATION (khong can API token).
        Day la phuong phap CHINH - chay headless, song song.

        Neu parallel_browsers > 1: Dung ParallelFlowGenerator (nhanh hon)
        Neu parallel_browsers = 1: Dung BrowserFlowGenerator (tuan tu)
        """
        self.log("=== TAO ANH BANG BROWSER ===")

        # QUAN TRONG: Sort prompts - nv/loc truoc, scene sau
        # De reference images (nhan vat, dia diem) duoc tao truoc scenes
        def sort_key(p):
            pid = p.get('id', '')
            if pid.startswith('nv'):
                return (0, pid)  # nvc, nv1, nv2... first
            elif pid.startswith('loc'):
                return (1, pid)  # loc1, loc2... second
            else:
                return (2, pid)  # scenes last

        prompts = sorted(prompts, key=sort_key)
        self.log(f"Da sap xep: {[p.get('id') for p in prompts[:5]]}... (nv/loc truoc, scene sau)")

        # Load settings
        headless = True
        parallel_browsers = 1
        try:
            import yaml
            config_path = self.config_dir / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    settings = yaml.safe_load(f) or {}
                headless = settings.get('browser_headless', True)
                parallel_browsers = max(1, min(5, settings.get('parallel_browsers', 1)))
        except:
            pass

        # Tim Excel file
        excel_files = list((proj_dir / "prompts").glob("*_prompts.xlsx"))
        if not excel_files:
            self.log("Khong tim thay file Excel!", "ERROR")
            return {"success": 0, "failed": len(prompts)}

        # =====================================================================
        # PARALLEL MODE: Nhieu browser song song
        # =====================================================================
        if parallel_browsers > 1:
            self.log(f"[PARALLEL] Chay {parallel_browsers} browsers song song")
            try:
                from modules.parallel_flow_generator import ParallelFlowGenerator

                generator = ParallelFlowGenerator(
                    project_path=str(proj_dir),
                    num_browsers=parallel_browsers,
                    headless=headless,
                    verbose=True
                )

                result = generator.generate_parallel(
                    excel_path=excel_files[0],
                    overwrite=False
                )

                if result.get("success"):
                    stats = result.get("stats", {})
                    return {
                        "success": stats.get("success", 0),
                        "failed": stats.get("failed", 0),
                        "parallel": True,
                        "speedup": stats.get("speedup", 1)
                    }
                else:
                    self.log(f"Parallel error: {result.get('error')}", "ERROR")
                    return {"success": 0, "failed": len(prompts)}

            except ImportError as e:
                self.log(f"Khong import duoc ParallelFlowGenerator: {e}", "WARN")
                self.log("Fallback ve mode 1 browser...", "WARN")
                # Fallback to single browser
            except Exception as e:
                self.log(f"Parallel error: {e}", "ERROR")
                import traceback
                traceback.print_exc()
                return {"success": 0, "failed": len(prompts)}

        # =====================================================================
        # SEQUENTIAL MODE: 1 browser
        # =====================================================================
        try:
            from modules.browser_flow_generator import BrowserFlowGenerator
        except ImportError:
            self.log("Khong import duoc BrowserFlowGenerator!", "ERROR")
            return {"success": 0, "failed": len(prompts)}

        # Find profile to use
        profile_name = "main"  # Default

        # Use assigned profile if set (for parallel processing)
        if self.assigned_profile:
            profile_name = self.assigned_profile
            self.log(f"Dung profile (assigned): {profile_name}")
        else:
            # Find first available profile from chrome_profiles directory
            # ROOT = thu muc chua ve3_pro.py (D:\AUTO\ve3-tool)
            root_dir = Path(__file__).parent.parent.resolve()  # modules/../ = root
            profiles_dir = root_dir / "chrome_profiles"

            self.log(f"Tim profile tai: {profiles_dir}")

            if profiles_dir.exists():
                # Liet ke tat ca profiles
                all_items = list(profiles_dir.iterdir())
                available_profiles = [p.name for p in all_items if p.is_dir() and not p.name.startswith('.')]
                self.log(f"Cac profiles: {available_profiles}")

                if available_profiles:
                    # Uu tien profile KHONG phai "main" (user da tao)
                    non_main = [p for p in available_profiles if p != "main"]
                    profile_name = non_main[0] if non_main else available_profiles[0]
                    self.log(f">>> Dung profile: {profile_name}")
            else:
                self.log(f"Chua co thu muc chrome_profiles, tao moi...")
                profiles_dir.mkdir(exist_ok=True)
                (profiles_dir / "main").mkdir(exist_ok=True)
                available_profiles = ["main"]

        try:
            # REUSE browser generator neu da co (giu nguyen session)
            need_new_generator = False
            saved_project_id = None  # Luu project_id tu generator cu

            if self._browser_generator is None or self._browser_generator.driver is None:
                need_new_generator = True
                self.log("Tao browser generator moi...")
                # Lay project_id tu generator cu neu co
                if self._browser_generator:
                    saved_project_id = self._browser_generator.config.get('flow_project_id', '')
                    if saved_project_id:
                        self.log(f"  -> Luu project_id tu generator cu: {saved_project_id[:8]}...")
            else:
                # Check if existing session is still valid
                try:
                    # Try a simple operation to verify session is alive
                    _ = self._browser_generator.driver.current_url
                    self.log("Reuse browser generator (session con valid)...")
                except Exception as e:
                    self.log(f"Session cu da het han ({type(e).__name__}), tao moi...")
                    need_new_generator = True
                    # Luu project_id truoc khi cleanup
                    saved_project_id = self._browser_generator.config.get('flow_project_id', '')
                    if saved_project_id:
                        self.log(f"  -> Luu project_id tu session cu: {saved_project_id[:8]}...")
                    # Clean up old generator
                    try:
                        self._browser_generator.stop_browser()
                    except:
                        pass
                    self._browser_generator = None

            if need_new_generator:
                generator = BrowserFlowGenerator(
                    project_path=str(proj_dir),
                    profile_name=profile_name,
                    headless=headless,
                    verbose=True
                )
                self._browser_generator = generator
                # Restore project_id tu generator cu
                if saved_project_id:
                    generator.config['flow_project_id'] = saved_project_id
                    self.log(f"  -> Restore project_id: {saved_project_id[:8]}...")
            else:
                generator = self._browser_generator
                # Update project path neu khac
                if str(generator.project_path) != str(proj_dir):
                    generator.project_path = Path(proj_dir)
                    generator.project_code = proj_dir.name

            # === QUAN TRONG: Share project_id tu profile sang generator ===
            # Tim profile matching voi profile_name de lay project_id
            matching_profile = None
            for p in self.profiles:
                if Path(p.value).name == profile_name:
                    matching_profile = p
                    break

            if matching_profile and matching_profile.project_id:
                generator.config['flow_project_id'] = matching_profile.project_id
                self.log(f"  -> Share project_id: {matching_profile.project_id[:8]}...")
            else:
                self.log(f"  -> Chua co project_id trong profile, se tao moi khi can")

            # Override callback
            def custom_log(msg, level="info"):
                self.log(msg, level.upper() if level else "INFO")
            generator._log = custom_log

            # Truyen prompts da load san (tu _load_prompts) thay vi doc lai Excel
            # Su dung generate_from_prompts_auto de tu dong chon mode (chrome/api)
            result = generator.generate_from_prompts_auto(
                prompts=prompts,
                excel_path=excel_files[0]
            )

            # === SYNC project_id nguoc lai tu generator ve profile ===
            # Neu generator da lay token moi va co project_id, cap nhat ve profile
            new_project_id = generator.config.get('flow_project_id', '')
            if new_project_id and matching_profile:
                if matching_profile.project_id != new_project_id:
                    self.log(f"  -> Sync project_id ve profile: {new_project_id[:8]}...")
                    matching_profile.project_id = new_project_id
                    self.save_cached_tokens()  # Luu lai de lan sau dung

            if result.get("success"):
                stats = result.get("stats", {})
                return {
                    "success": stats.get("success", 0),
                    "failed": stats.get("failed", 0)
                }
            else:
                return {"success": 0, "failed": len(prompts), "error": result.get("error")}

        except Exception as e:
            self.log(f"Browser error: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return {"success": 0, "failed": len(prompts)}

    def generate_images_parallel(self, prompts: List[Dict]) -> Dict:
        """
        Tao anh - DUNG 1 PROFILE DUY NHAT.
        Dam bao media_name consistent (cung token tao ref va scene).

        DON GIAN & AN TOAN:
        - 1 voice = 1 profile
        - Tat ca anh (nv/loc + scenes) dung chung 1 profile
        - media_name luon valid vi cung token
        """
        self.log(f"=== TAO {len(prompts)} ANH (1 PROFILE) ===")

        # Sort: nv/loc truoc, scene sau (de co media_name khi tao scene)
        def sort_key(p):
            pid = p.get('id', '')
            if pid.startswith('nv'):
                return (0, pid)  # nv first
            elif pid.startswith('loc'):
                return (1, pid)  # loc second
            else:
                return (2, pid)  # scenes last

        sorted_prompts = sorted(prompts, key=sort_key)

        return self._generate_images_single_profile(sorted_prompts)

    def _generate_images_single_profile(self, prompts: List[Dict]) -> Dict:
        """
        Tao TAT CA images bang 1 PROFILE DUY NHAT.
        Don gian, an toan, media_name luon consistent.
        """
        results = {"success": 0, "failed": 0, "pending": list(prompts)}

        # Loop until all done or no resources left
        attempt = 0
        while results["pending"] and attempt < self.max_retries * 2:
            if self.stop_flag:
                break

            attempt += 1
            self.log(f"=== ROUND {attempt} - {len(results['pending'])} pending ===")

            # Get 1 profile with valid token
            all_accounts = self.headless_accounts + self.profiles
            active_profile = None
            for p in all_accounts:
                if p.token and p.status != 'exhausted':
                    active_profile = p
                    break

            if not active_profile:
                self.log("Khong co account nao co token!", "WARN")
                n = self.get_all_tokens()
                if n == 0:
                    break
                for p in all_accounts:
                    if p.token:
                        active_profile = p
                        break

            if not active_profile:
                break

            profile_name = Path(active_profile.value).name
            self.log(f"Dung profile: {profile_name}")

            # Process images sequentially with 1 profile
            pending = results["pending"]
            results["pending"] = []
            done_count = 0

            for prompt_data in pending:
                if self.stop_flag:
                    results["pending"].append(prompt_data)
                    continue

                pid = prompt_data.get('id', '')
                self.log(f"[{pid}] Dang tao...")

                # Check token still valid
                if not active_profile.token:
                    self.log(f"[{pid}] Token het han, dung lai!", "WARN")
                    results["pending"].append(prompt_data)
                    # Add remaining to pending
                    idx = pending.index(prompt_data)
                    results["pending"].extend(pending[idx+1:])
                    break

                success, token_expired = self.generate_single_image(prompt_data, active_profile)

                if token_expired:
                    active_profile.token = ""
                    self.log(f"[{pid}] Token het han!", "WARN")
                    results["pending"].append(prompt_data)
                    # Add remaining to pending
                    idx = pending.index(prompt_data)
                    results["pending"].extend(pending[idx+1:])
                    break

                if success:
                    self.log(f"[{pid}] OK!", "OK")
                    done_count += 1
                    results["success"] += 1
                else:
                    self.log(f"[{pid}] FAIL", "WARN")
                    results["pending"].append(prompt_data)

                # Small delay
                time.sleep(self.delay)

                # Progress log
                if done_count % 5 == 0:
                    self.log(f"[Progress] {done_count}/{len(pending)}")

            self.log(f"Round {attempt}: +{done_count} OK, {len(results['pending'])} pending")

            # If still have pending, wait a bit before retry
            if results["pending"]:
                time.sleep(1)

        results["failed"] = len(results["pending"])
        self.log(f"=== XONG: {results['success']} OK, {results['failed']} FAIL ===")

        return results

    # ========== MAIN PIPELINE ==========

    def run(
        self,
        input_path: str,
        output_dir: str = None,
        callback: Callable = None
    ) -> Dict:
        """
        BROWSER MODE PIPELINE - Tao anh bang JS automation.

        Flow:
        1. Check requirements (Chrome, AI keys)
        2. Tao SRT tu voice (Whisper)
        3. Tao prompts tu SRT (AI)
        4. Tao anh bang BROWSER JS (khong can API token)
        5. Ghep video (FFmpeg)

        Args:
            input_path: Voice file (.mp3, .wav) hoac Excel (.xlsx)
            output_dir: Thu muc output (optional)
            callback: Ham log callback

        Returns:
            Dict with success/failed counts
        """
        self.callback = callback
        self.stop_flag = False

        inp = Path(input_path)
        ext = inp.suffix.lower()
        name = inp.stem

        # Setup output dir
        if output_dir:
            proj_dir = Path(output_dir)
        else:
            proj_dir = Path("PROJECTS") / name

        proj_dir.mkdir(parents=True, exist_ok=True)
        for d in ["srt", "prompts", "nv", "img"]:
            (proj_dir / d).mkdir(exist_ok=True)

        excel_path = proj_dir / "prompts" / f"{name}_prompts.xlsx"
        srt_path = proj_dir / "srt" / f"{name}.srt"

        # Read generation mode from settings
        generation_mode = 'api'  # Default
        try:
            import yaml
            settings_path = self.config_dir / "settings.yaml"
            if settings_path.exists():
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = yaml.safe_load(f) or {}
                generation_mode = settings.get('generation_mode', 'api')
        except:
            pass

        mode_display = "API MODE" if generation_mode == 'api' else "BROWSER JS MODE"

        self.log("="*50)
        self.log(f"VE3 TOOL v{__version__} - {mode_display}")
        self.log(f"INPUT: {inp}")
        self.log(f"OUTPUT: {proj_dir}")
        self.log("="*50)

        # === RESUME CHECK: Kiểm tra đã làm đến bước nào ===
        final_video = proj_dir / f"{name}_final.mp4"

        # Nếu video cuối đã tồn tại → hoàn thành rồi
        if final_video.exists():
            self.log("✅ RESUME: Video đã hoàn thành, skip!", "OK")
            return {"success": True, "skipped": "video_exists", "video": str(final_video)}

        # Log trạng thái resume
        resume_status = []
        if srt_path.exists():
            resume_status.append("SRT ✓")
        if excel_path.exists():
            resume_status.append("Excel ✓")

        img_dir = proj_dir / "img"
        existing_images = len(list(img_dir.glob("*.png"))) + len(list(img_dir.glob("*.mp4"))) if img_dir.exists() else 0
        if existing_images > 0:
            resume_status.append(f"Images: {existing_images} ✓")

        if resume_status:
            self.log(f"📌 RESUME: {' | '.join(resume_status)}")

        # === 1. CHECK REQUIREMENTS ===
        self.log("[STEP 1] Kiem tra yeu cau...")

        ok, missing = self.check_requirements(has_voice=(ext in ['.mp3', '.wav']))
        if not ok:
            self.log("THIEU:", "ERROR")
            for m in missing:
                self.log(f"  - {m}", "ERROR")
            return {"error": "missing_requirements", "missing": missing}

        # BROWSER MODE - khong can tokens
        self.log("  MODE: Browser JS automation (khong can API token)")
        self.log(f"  AI keys: DeepSeek={len(self.deepseek_keys)}, Groq={len(self.groq_keys)}, Gemini={len(self.gemini_keys)}")

        # === 2. TAO SRT + PROMPTS ===
        self.log("[STEP 2] Tao SRT + Prompts...")

        voice_path = None
        if ext in ['.mp3', '.wav']:
            voice_path = proj_dir / f"{name}{ext}"
            if inp != voice_path:
                shutil.copy2(inp, voice_path)

            # Tao SRT (skip nếu đã có)
            if srt_path.exists():
                self.log("  ⏭️ SRT đã tồn tại, skip!")
            else:
                if not self.make_srt(voice_path, srt_path):
                    return {"error": "srt_failed"}

        # Tao Prompts (skip nếu đã có)
        if ext == '.xlsx':
            if inp != excel_path:
                shutil.copy2(inp, excel_path)
        elif excel_path.exists():
            self.log("  ⏭️ Excel đã tồn tại, skip!")
        else:
            if not self.make_prompts(proj_dir, name, excel_path):
                return {"error": "prompts_failed"}

        self.log("BROWSER MODE: Khong can token, su dung JS automation")

        # === 3. DOI CHARACTER GENERATION (PARALLEL) ===
        # Neu character generation dang chay song song, doi no xong
        if self._character_gen_thread is not None:
            self.log("[STEP 3] Doi character generation hoan thanh...")
            self._wait_for_character_generation(timeout=600)

            # Lay ket qua character generation
            char_results = self._character_gen_result or {"success": 0, "failed": 0}
            self.log(f"  Character images: success={char_results.get('success', 0)}, failed={char_results.get('failed', 0)}")
        else:
            self.log("[STEP 3] Character generation khong chay song song")
            char_results = {"success": 0, "failed": 0}

        # === 4. LOAD SCENE PROMPTS (chi scenes, bo qua characters da tao) ===
        self.log("[STEP 4] Load scene prompts...")

        all_prompts = self._load_prompts(excel_path, proj_dir)

        if not all_prompts:
            return {"error": "no_prompts"}

        # Filter: chi lay prompts CHUA co anh
        prompts = [p for p in all_prompts if not Path(p['output_path']).exists()]
        existing_count = len(all_prompts) - len(prompts)

        if existing_count > 0:
            self.log(f"  ⏭️ Đã có {existing_count}/{len(all_prompts)} ảnh (resume)")
            self.log(f"  📌 Còn {len(prompts)} ảnh cần tạo")
        else:
            self.log(f"  Tổng: {len(all_prompts)} prompts")

        if not prompts:
            self.log("  ✅ Tất cả ảnh đã tồn tại, skip tạo ảnh!", "OK")
            # Merge results with character generation
            results = {
                "success": char_results.get("success", 0),
                "failed": char_results.get("failed", 0),
                "skipped": "all_exist"
            }
        else:
            # === 5. TAO IMAGES - CHON MODE DUA TREN SETTINGS ===
            # Start video worker (parallel with image generation)
            self._start_video_worker(proj_dir)

            if generation_mode == 'api':
                self.log("[STEP 5] Tao images bang API MODE...")
                scene_results = self.generate_images_api(prompts, proj_dir)
            else:
                self.log("[STEP 5] Tao images bang BROWSER MODE...")
                scene_results = self.generate_images_browser(prompts, proj_dir)

            # Merge results
            results = {
                "success": char_results.get("success", 0) + scene_results.get("success", 0),
                "failed": char_results.get("failed", 0) + scene_results.get("failed", 0)
            }

        # === 6. FINAL CHECK ===
        self.log("[STEP 6] Kiem tra ket qua...")

        if results.get("failed", 0) > 0:
            self.log(f"CON {results['failed']} ANH CHUA XONG!", "WARN")
        else:
            self.log("TAT CA ANH DA HOAN THANH!", "OK")

        # === 7. EXPORT TXT & SRT ===
        self.log("[STEP 7] Xuat TXT & SRT...")
        self._export_scenes(excel_path, proj_dir, name)

        # === 8. COMPOSE VIDEO (LUON chay - du co vai anh fail) ===
        self.log("[STEP 8] Ghep video...")
        if results.get("failed", 0) > 0:
            self.log(f"  CANH BAO: {results['failed']} anh fail, nhung van ghep video voi anh co san!", "WARN")

        video_path = self._compose_video(proj_dir, excel_path, name)
        if video_path:
            self.log(f"  -> Video: {video_path.name}", "OK")
            results["video"] = str(video_path)
        else:
            self.log("  Video composer khong kha dung hoac thieu file", "WARN")

        # === 9. DONG BROWSER ===
        self._close_browser()

        # === 10. WAIT FOR VIDEO GENERATION (if running) ===
        if self._video_worker_running:
            self.log("[STEP 10] Doi video generation hoan thanh...")
            # Wait for queue to empty (with timeout)
            wait_start = time.time()
            max_wait = 600  # 10 minutes max
            while self._video_worker_running and time.time() - wait_start < max_wait:
                with self._video_queue_lock:
                    if not self._video_queue:
                        break
                time.sleep(2)

            # Stop worker and get results
            self._stop_video_worker()
            video_results = self.get_video_results()
            self.log(f"[VIDEO] Ket qua: {video_results['success']} OK, {video_results['failed']} failed")
            results["video_gen"] = video_results

        return results

    def _close_browser(self):
        """Dong browser generator (giu nguyen working profile)."""
        if self._browser_generator is not None:
            try:
                self.log("[CLEANUP] Dong browser...")
                self._browser_generator.stop_browser()
            except:
                pass
            self._browser_generator = None

    def _export_scenes(self, excel_path: Path, proj_dir: Path, name: str) -> None:
        """Export scenes ra TXT va SRT de ho tro video editing."""
        try:
            from modules.excel_manager import PromptWorkbook

            wb = PromptWorkbook(excel_path)
            wb.load_or_create()

            # Export TXT - danh sach phan canh
            txt_path = proj_dir / f"{name}_scenes.txt"
            if wb.export_scenes_txt(txt_path):
                self.log(f"  -> TXT: {txt_path.name}", "OK")

            # Export SRT - thoi gian phan canh
            srt_output_path = proj_dir / f"{name}_scenes.srt"
            if wb.export_scenes_srt(srt_output_path):
                self.log(f"  -> SRT: {srt_output_path.name}", "OK")

        except Exception as e:
            self.log(f"  Export error: {e}", "WARN")

    def _process_srt_for_video(self, srt_path: Path, output_path: Path, max_chars: int = 50) -> Path:
        """
        Xử lý SRT: tách dòng dài thành nhiều dòng ngắn (max 50 ký tự).
        Chia đều timestamp theo số từ.
        """
        import re

        def parse_time(time_str: str) -> float:
            """Parse SRT timestamp to seconds."""
            h, m, s = time_str.replace(',', '.').split(':')
            return int(h) * 3600 + int(m) * 60 + float(s)

        def format_time(seconds: float) -> str:
            """Format seconds to SRT timestamp."""
            h = int(seconds // 3600)
            m = int((seconds % 3600) // 60)
            s = seconds % 60
            return f"{h:02d}:{m:02d}:{s:06.3f}".replace('.', ',')

        def split_text(text: str, max_len: int) -> list:
            """Tách text thành các đoạn <= max_len ký tự, tách theo từ."""
            words = text.split()
            chunks = []
            current = []
            current_len = 0

            for word in words:
                word_len = len(word) + (1 if current else 0)  # +1 for space
                if current_len + word_len <= max_len:
                    current.append(word)
                    current_len += word_len
                else:
                    if current:
                        chunks.append(' '.join(current))
                    current = [word]
                    current_len = len(word)

            if current:
                chunks.append(' '.join(current))

            return chunks if chunks else [text[:max_len]]

        try:
            with open(srt_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Parse SRT entries
            pattern = r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)(?=\n\n|\Z)'
            entries = re.findall(pattern, content, re.DOTALL)

            new_entries = []
            new_index = 1

            for idx, start, end, text in entries:
                text = text.strip().replace('\n', ' ').upper()  # Viết hoa
                start_sec = parse_time(start)
                end_sec = parse_time(end)
                duration = end_sec - start_sec

                # Tách text nếu quá dài
                if len(text) <= max_chars:
                    new_entries.append((new_index, start, end, text))
                    new_index += 1
                else:
                    chunks = split_text(text, max_chars)
                    chunk_duration = duration / len(chunks)

                    for i, chunk in enumerate(chunks):
                        chunk_start = start_sec + i * chunk_duration
                        chunk_end = start_sec + (i + 1) * chunk_duration
                        new_entries.append((
                            new_index,
                            format_time(chunk_start),
                            format_time(chunk_end),
                            chunk
                        ))
                        new_index += 1

            # Write new SRT
            with open(output_path, 'w', encoding='utf-8') as f:
                for idx, start, end, text in new_entries:
                    f.write(f"{idx}\n{start} --> {end}\n{text}\n\n")

            self.log(f"  SRT processed: {len(entries)} -> {len(new_entries)} entries (max {max_chars} chars)")
            return output_path

        except Exception as e:
            self.log(f"  SRT process error: {e}", "WARN")
            return srt_path  # Return original if error

    def _compose_video(self, proj_dir: Path, excel_path: Path, name: str) -> Optional[Path]:
        """
        Tự động ghép video từ ảnh + voice + SRT.
        Đọc trực tiếp từ Excel format của prompts generator.
        """
        import subprocess
        import openpyxl
        import tempfile

        # Check FFmpeg
        try:
            result = subprocess.run(["ffmpeg", "-version"], capture_output=True, text=True)
            if result.returncode != 0:
                self.log("  FFmpeg khong hoat dong!", "ERROR")
                return None
        except FileNotFoundError:
            self.log("  FFmpeg chua cai! https://ffmpeg.org/download.html", "ERROR")
            return None

        # Tìm voice file
        voice_files = list(proj_dir.glob("*.mp3")) + list(proj_dir.glob("*.wav"))
        if not voice_files:
            self.log("  Khong tim thay voice file (.mp3/.wav)", "WARN")
            return None
        voice_path = voice_files[0]

        # Tìm SRT file
        srt_files = list(proj_dir.glob("srt/*.srt")) + list(proj_dir.glob("*.srt"))
        srt_path = srt_files[0] if srt_files else None

        # Xử lý SRT: tách dòng dài (max 50 ký tự)
        if srt_path:
            processed_srt = proj_dir / f"{name}_video.srt"
            srt_path = self._process_srt_for_video(srt_path, processed_srt, max_chars=50)

        output_path = proj_dir / f"{name}_final.mp4"
        img_dir = proj_dir / "img"

        self.log(f"  Voice: {voice_path.name}")
        self.log(f"  SRT: {srt_path.name if srt_path else 'None'}")
        self.log(f"  Excel: {excel_path.name}")

        try:
            # 1. Load scenes từ Excel (Scenes sheet)
            wb = openpyxl.load_workbook(excel_path)

            # Tìm sheet Scenes
            scenes_sheet = None
            for sheet_name in wb.sheetnames:
                if 'scene' in sheet_name.lower():
                    scenes_sheet = wb[sheet_name]
                    break

            if not scenes_sheet:
                self.log("  Khong tim thay sheet 'Scenes' trong Excel!", "ERROR")
                return None

            # Đọc headers
            headers = [cell.value for cell in scenes_sheet[1]]
            self.log(f"  Headers: {headers[:5]}...")

            # Tìm cột cần thiết (chỉ cần ID và Start Time)
            id_col = start_col = None
            for i, h in enumerate(headers):
                if h is None:
                    continue
                h_lower = str(h).lower()
                if 'id' in h_lower and id_col is None:
                    id_col = i
                if 'start' in h_lower and 'time' in h_lower:
                    start_col = i

            if id_col is None:
                self.log("  Khong tim thay cot ID!", "ERROR")
                return None

            self.log(f"  Columns: ID={id_col}, Start={start_col}")

            # 2. Load media (video clips hoặc images) với timestamps
            # Ưu tiên: video clip (.mp4) > image (.png)
            media_items = []
            video_count = 0
            image_count = 0

            for row in scenes_sheet.iter_rows(min_row=2, values_only=True):
                if row[id_col] is None:
                    continue

                scene_id = str(row[id_col]).strip()

                # Chỉ lấy scenes có số (1, 2, 3...), bỏ qua nv1, loc1
                if not scene_id.isdigit():
                    continue

                # Ưu tiên video clip (.mp4), fallback to image (.png)
                video_path = img_dir / f"{scene_id}.mp4"
                img_path = img_dir / f"{scene_id}.png"

                if video_path.exists():
                    media_path = video_path
                    is_video = True
                    video_count += 1
                elif img_path.exists():
                    media_path = img_path
                    is_video = False
                    image_count += 1
                else:
                    continue

                # Parse start_time
                start_time = 0.0
                if start_col is not None and row[start_col]:
                    start_time = self._parse_timestamp(str(row[start_col]))

                media_items.append({
                    'id': scene_id,
                    'path': str(media_path),
                    'start': start_time,
                    'is_video': is_video
                })

            if not media_items:
                self.log("  Khong tim thay media nao trong img/ folder!", "ERROR")
                return None

            # Sắp xếp theo start_time
            media_items.sort(key=lambda x: x['start'])
            self.log(f"  Tim thay {len(media_items)} media: {video_count} video clips, {image_count} images")

            # 3. Tính duration cho mỗi media (CHỈ dựa vào start_time)
            # Lấy tổng thời lượng từ voice
            probe_cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                        "-of", "default=noprint_wrappers=1:nokey=1", str(voice_path)]
            result = subprocess.run(probe_cmd, capture_output=True, text=True)
            total_duration = float(result.stdout.strip()) if result.stdout.strip() else 60.0
            self.log(f"  Voice duration: {total_duration:.1f}s")

            # Tính duration mỗi media = start_time[i+1] - start_time[i]
            for i, item in enumerate(media_items):
                if i < len(media_items) - 1:
                    item['duration'] = media_items[i + 1]['start'] - item['start']
                else:
                    # Media cuối: kéo dài đến hết voice
                    item['duration'] = total_duration - item['start']

                # Đảm bảo duration hợp lệ (tối thiểu 0.5s)
                if item['duration'] <= 0:
                    item['duration'] = max(0.5, (total_duration - item['start']) / max(1, len(media_items) - i))

            # 4. Tạo video với FFmpeg + Fade Transitions
            # Windows fix: Don't use context manager - manual cleanup with retry
            temp_dir = tempfile.mkdtemp()
            try:
                # Debug: show first few media items
                self.log(f"  First media: {Path(media_items[0]['path']).resolve()}")
                for i in range(min(3, len(media_items))):
                    item = media_items[i]
                    media_type = "🎬" if item['is_video'] else "🖼️"
                    self.log(f"    {media_type} #{item['id']}: start={item['start']:.1f}s, dur={item['duration']:.1f}s")

                # Video không có audio
                temp_video = Path(temp_dir) / "temp_video.mp4"

                # Fade in/out cho mỗi clip
                import random
                FADE_DURATION = 0.4  # 0.4 giây fade
                self.log(f"  Dang tao {len(media_items)} clips ({video_count} video, {image_count} image)...")

                # Tạo từng clip
                clip_paths = []
                for i, item in enumerate(media_items):
                    clip_path = Path(temp_dir) / f"clip_{i:03d}.mp4"
                    abs_path = str(Path(item['path']).resolve()).replace('\\', '/')
                    target_duration = item['duration']

                    # === TRANSITION EFFECTS ===
                    # Chọn ngẫu nhiên hiệu ứng: mix (crossfade) hoặc tối dần (fade to black)
                    # Đơn giản và phù hợp với thời lượng ngắn của mỗi clip
                    transition_type = random.choice(['fade_black', 'fade_black', 'mix'])  # 2/3 fade to black
                    fade_out_start = max(0, target_duration - FADE_DURATION)

                    if transition_type == 'fade_black':
                        # Tối dần: fade in/out to black
                        fade_filter = f"fade=t=in:st=0:d={FADE_DURATION},fade=t=out:st={fade_out_start}:d={FADE_DURATION}"
                    else:
                        # Mix: fade với alpha (crossfade effect khi concat)
                        fade_filter = f"fade=t=in:st=0:d={FADE_DURATION}:alpha=1,fade=t=out:st={fade_out_start}:d={FADE_DURATION}:alpha=1"

                    if item['is_video']:
                        # === VIDEO CLIP: Cắt lấy phần giữa + thêm transitions ===
                        # Lấy duration của video gốc
                        probe_cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                                    "-of", "default=noprint_wrappers=1:nokey=1", abs_path]
                        probe_result = subprocess.run(probe_cmd, capture_output=True, text=True)
                        video_duration = float(probe_result.stdout.strip()) if probe_result.stdout.strip() else 8.0

                        # Base filter: scale + pad + transitions
                        base_vf = f"scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2,{fade_filter}"

                        if video_duration > target_duration:
                            # Cắt lấy phần giữa: bỏ đầu và cuối bằng nhau
                            trim_total = video_duration - target_duration
                            trim_start = trim_total / 2
                            # Sử dụng -ss (seek) và -t (duration)
                            cmd_clip = [
                                "ffmpeg", "-y",
                                "-ss", str(trim_start),
                                "-i", abs_path,
                                "-t", str(target_duration),
                                "-vf", base_vf,
                                "-c:v", "libx264", "-pix_fmt", "yuv420p",
                                "-an",  # Bỏ audio từ video clip
                                "-r", "30", str(clip_path)
                            ]
                        else:
                            # Video ngắn hơn target → dùng nguyên video
                            cmd_clip = [
                                "ffmpeg", "-y",
                                "-i", abs_path,
                                "-t", str(target_duration),
                                "-vf", base_vf,
                                "-c:v", "libx264", "-pix_fmt", "yuv420p",
                                "-an",
                                "-r", "30", str(clip_path)
                            ]
                    else:
                        # === IMAGE: Tạo static clip với transitions ===
                        vf = f"scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2,{fade_filter}"

                        cmd_clip = [
                            "ffmpeg", "-y",
                            "-loop", "1", "-t", str(target_duration),
                            "-i", abs_path,
                            "-vf", vf,
                            "-c:v", "libx264", "-pix_fmt", "yuv420p",
                            "-r", "30", str(clip_path)
                        ]

                    result = subprocess.run(cmd_clip, capture_output=True, text=True)
                    if result.returncode != 0:
                        self.log(f"  Clip {i} failed: {result.stderr[-200:]}", "ERROR")
                        continue

                    clip_paths.append(clip_path)

                    # Progress log mỗi 10 clips
                    if (i + 1) % 10 == 0:
                        self.log(f"  ... {i + 1}/{len(media_items)} clips")

                if not clip_paths:
                    self.log("  Khong tao duoc clip nao!", "ERROR")
                    return None

                self.log(f"  Da tao {len(clip_paths)} clips, dang ghep...")

                # Concat tất cả clips (timing chính xác, không rút ngắn)
                list_file = Path(temp_dir) / "clips.txt"
                with open(list_file, 'w', encoding='utf-8') as f:
                    for cp in clip_paths:
                        f.write(f"file '{str(cp).replace(chr(92), '/')}'\n")

                cmd_concat = [
                    "ffmpeg", "-y", "-f", "concat", "-safe", "0",
                    "-i", str(list_file),
                    "-c", "copy", str(temp_video)
                ]
                result = subprocess.run(cmd_concat, capture_output=True, text=True)
                if result.returncode != 0:
                    error_lines = result.stderr.strip().split('\n')
                    self.log(f"  Concat error: {error_lines[-1]}", "ERROR")
                    return None

                # Thêm audio
                temp_with_audio = Path(temp_dir) / "with_audio.mp4"
                self.log("  Dang them voice...")
                cmd2 = [
                    "ffmpeg", "-y",
                    "-i", str(temp_video),
                    "-i", str(voice_path),
                    "-c:v", "copy", "-c:a", "aac", "-b:a", "192k",
                    "-shortest", str(temp_with_audio)
                ]
                result = subprocess.run(cmd2, capture_output=True, text=True)
                if result.returncode != 0:
                    self.log(f"  FFmpeg error: {result.stderr[:200]}", "ERROR")
                    return None

                # Burn subtitles nếu có
                if srt_path and srt_path.exists():
                    self.log("  Dang burn phu de...")
                    # Escape SRT path cho FFmpeg filter
                    srt_escaped = str(srt_path).replace('\\', '/').replace(':', '\\:')

                    # Font path - Anton Regular
                    font_dir = "C\\:/Users/admin/AppData/Local/Microsoft/Windows/Fonts"

                    # Style: Chữ trắng, viền đen, font Anton
                    # PrimaryColour format: &HAABBGGRR (Alpha, Blue, Green, Red)
                    # &H00FFFFFF = white, &H00000000 = black
                    subtitle_style = (
                        "FontName=Anton,"
                        "FontSize=32,"  # Vua phai, 1 dong ~ 50 ky tu
                        "PrimaryColour=&H00FFFFFF,"  # Trắng
                        "OutlineColour=&H00000000,"  # Đen
                        "BorderStyle=1,"
                        "Outline=2,"  # Vien mong
                        "Shadow=0,"
                        "MarginV=30,"
                        "Alignment=2"  # Bottom center
                    )

                    # FFmpeg command với custom font
                    vf_filter = f"subtitles='{srt_escaped}':fontsdir='{font_dir}':force_style='{subtitle_style}'"

                    cmd3 = [
                        "ffmpeg", "-y",
                        "-i", str(temp_with_audio),
                        "-vf", vf_filter,
                        "-c:a", "copy", str(output_path)
                    ]
                    result = subprocess.run(cmd3, capture_output=True, text=True)
                    if result.returncode != 0:
                        self.log(f"  Subtitle burn failed: {result.stderr[-200:]}", "WARN")
                        # Fallback: thử không có custom font
                        self.log("  Thu lai voi font mac dinh...", "WARN")
                        vf_simple = f"subtitles='{srt_escaped}':force_style='FontSize=32,PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,BorderStyle=1,Outline=2'"
                        cmd3_simple = [
                            "ffmpeg", "-y",
                            "-i", str(temp_with_audio),
                            "-vf", vf_simple,
                            "-c:a", "copy", str(output_path)
                        ]
                        result = subprocess.run(cmd3_simple, capture_output=True, text=True)
                        if result.returncode != 0:
                            import shutil
                            shutil.copy(temp_with_audio, output_path)
                else:
                    import shutil
                    shutil.copy(temp_with_audio, output_path)

                self.log(f"  Video hoan thanh: {output_path.name}", "OK")
                return output_path

            finally:
                # Windows fix: Wait for FFmpeg to release file handles, then cleanup with retry
                import gc
                import shutil
                gc.collect()
                time.sleep(1)  # Wait longer for file handles

                # Retry cleanup up to 5 times
                for attempt in range(5):
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=False)
                        break
                    except PermissionError:
                        gc.collect()
                        time.sleep(1)
                        if attempt == 4:
                            # Last attempt - ignore errors
                            shutil.rmtree(temp_dir, ignore_errors=True)

        except Exception as e:
            self.log(f"  Video compose error: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return None

    def _parse_timestamp(self, timestamp: str) -> float:
        """Parse timestamp SRT format (00:01:23,456) sang giây."""
        if not timestamp:
            return 0.0
        timestamp = timestamp.replace(",", ".")
        parts = timestamp.split(":")
        if len(parts) == 3:
            h, m, s = parts
            return int(h) * 3600 + int(m) * 60 + float(s)
        elif len(parts) == 2:
            m, s = parts
            return int(m) * 60 + float(s)
        return float(timestamp) if timestamp else 0.0

    def _compose_video_simple(self, proj_dir: Path, excel_path: Path, name: str,
                               images: list, voice_path: Path, srt_path: Path,
                               temp_dir: str) -> Optional[Path]:
        """Fallback: Ghép video đơn giản không có transition (nếu xfade fail)."""
        import subprocess
        import shutil

        output_path = proj_dir / f"{name}_final.mp4"
        temp_video = Path(temp_dir) / "temp_video.mp4"
        temp_with_audio = Path(temp_dir) / "with_audio.mp4"

        # Tạo file list cho FFmpeg concat
        list_file = Path(temp_dir) / "images.txt"
        with open(list_file, 'w', encoding='utf-8') as f:
            for img in images:
                abs_path = str(Path(img['path']).resolve()).replace('\\', '/')
                f.write(f"file '{abs_path}'\n")
                f.write(f"duration {img['duration']:.3f}\n")
            f.write(f"file '{abs_path}'\n")

        # Concat đơn giản
        cmd1 = [
            "ffmpeg", "-y", "-f", "concat", "-safe", "0",
            "-i", str(list_file),
            "-vf", "scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2",
            "-c:v", "libx264", "-pix_fmt", "yuv420p",
            "-r", "30", str(temp_video)
        ]
        result = subprocess.run(cmd1, capture_output=True, text=True)
        if result.returncode != 0:
            self.log(f"  Simple concat cung that bai!", "ERROR")
            return None

        # Thêm audio
        cmd2 = [
            "ffmpeg", "-y",
            "-i", str(temp_video),
            "-i", str(voice_path),
            "-c:v", "copy", "-c:a", "aac", "-b:a", "192k",
            "-shortest", str(temp_with_audio)
        ]
        result = subprocess.run(cmd2, capture_output=True, text=True)
        if result.returncode != 0:
            return None

        # Burn subtitles nếu có
        if srt_path and srt_path.exists():
            srt_escaped = str(srt_path).replace('\\', '/').replace(':', '\\:')
            # Style: Chữ trắng viền đen, font 32px
            vf_filter = f"subtitles='{srt_escaped}':force_style='FontSize=32,PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,BorderStyle=1,Outline=2'"
            cmd3 = [
                "ffmpeg", "-y",
                "-i", str(temp_with_audio),
                "-vf", vf_filter,
                "-c:a", "copy", str(output_path)
            ]
            result = subprocess.run(cmd3, capture_output=True, text=True)
            if result.returncode != 0:
                shutil.copy(temp_with_audio, output_path)
        else:
            shutil.copy(temp_with_audio, output_path)

        self.log(f"  Video (simple): {output_path.name}", "OK")
        return output_path

    def _load_prompts(self, excel_path: Path, proj_dir: Path) -> List[Dict]:
        """Load prompts tu Excel - doc TAT CA sheets."""
        import openpyxl

        prompts = []
        wb = openpyxl.load_workbook(excel_path)

        self.log(f"Excel co {len(wb.sheetnames)} sheets: {wb.sheetnames}")

        # Doc TAT CA sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            self.log(f"  Sheet '{sheet_name}' headers: {headers}")

            if not headers or all(h is None for h in headers):
                self.log(f"  -> Skip: no headers")
                continue

            # Tim cot ID, Prompt, video_prompt va reference_files
            id_col = None
            prompt_col = None
            video_prompt_col = None  # video_prompt column for I2V
            ref_col = None  # reference_files column
            chars_col = None  # characters_used column
            loc_col = None  # location_used column

            for i, h in enumerate(headers):
                if h is None:
                    continue
                h_lower = str(h).lower().strip()

                # Tim cot ID
                if id_col is None:
                    if h_lower == 'id' or h_lower == 'scene_id' or h_lower == 'sceneid':
                        id_col = i
                    elif 'id' in h_lower and ('scene' in h_lower or 'nv' in h_lower or 'char' in h_lower):
                        id_col = i

                # Tim cot video_prompt (cho Image-to-Video)
                if 'video' in h_lower and 'prompt' in h_lower:
                    video_prompt_col = i

                # Tim cot Prompt - uu tien english_prompt, sau do img_prompt
                if 'english' in h_lower and 'prompt' in h_lower:
                    prompt_col = i
                elif h_lower == 'img_prompt' and prompt_col is None:
                    prompt_col = i
                elif h_lower == 'image_prompt' and prompt_col is None:
                    prompt_col = i
                elif prompt_col is None and h_lower == 'prompt':
                    prompt_col = i
                elif prompt_col is None and 'prompt' in h_lower and 'video' not in h_lower and 'viet' not in h_lower:
                    prompt_col = i

                # Tim cot reference_files (cho scene images)
                if 'reference' in h_lower and 'file' in h_lower:
                    ref_col = i

                # Tim cot characters_used va location_used (de build reference_files neu can)
                if h_lower == 'characters_used':
                    chars_col = i
                if h_lower == 'location_used':
                    loc_col = i

            # Neu khong tim thay, thu cot dau = ID, tim cot co "prompt"
            if id_col is None and len(headers) > 0 and headers[0]:
                # Cot dau tien co the la ID
                first_col = str(headers[0]).lower()
                if 'id' in first_col or first_col in ['scene_id', 'nv_id', 'character_id']:
                    id_col = 0

            if id_col is None or prompt_col is None:
                self.log(f"  -> Skip: id_col={id_col}, prompt_col={prompt_col}")
                continue

            self.log(f"  -> Found: id_col={id_col} ({headers[id_col]}), prompt_col={prompt_col} ({headers[prompt_col]}), ref_col={ref_col}")

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                if id_col >= len(row) or prompt_col >= len(row):
                    continue

                pid = row[id_col]
                prompt = row[prompt_col]

                if not pid or not prompt:
                    continue

                pid_str = str(pid).strip()
                prompt_str = str(prompt).strip()

                # Skip DO_NOT_GENERATE markers (child characters, placeholders)
                if prompt_str == "DO_NOT_GENERATE" or prompt_str.upper() == "DO_NOT_GENERATE":
                    continue

                # Get video_prompt if available (for Image-to-Video)
                video_prompt = ""
                if video_prompt_col is not None and video_prompt_col < len(row):
                    video_prompt = row[video_prompt_col] or ""

                # Get reference_files if available
                reference_files = ""
                if ref_col is not None and ref_col < len(row):
                    reference_files = row[ref_col] or ""

                # === FALLBACK: Build reference_files tu characters_used + location_used ===
                # Neu reference_files rong, tao tu cac cot khac (dao dien da set)
                if not reference_files:
                    ref_list = []

                    # Lay characters_used
                    if chars_col is not None and chars_col < len(row):
                        chars_val = row[chars_col]
                        if chars_val:
                            try:
                                chars = json.loads(str(chars_val)) if str(chars_val).startswith('[') else [c.strip() for c in str(chars_val).split(',') if c.strip()]
                                for c in chars:
                                    c_id = c.replace('.png', '').strip()
                                    if c_id and c_id not in ref_list:
                                        ref_list.append(f"{c_id}.png")
                            except:
                                pass

                    # Lay location_used
                    if loc_col is not None and loc_col < len(row):
                        loc_val = row[loc_col]
                        if loc_val:
                            loc_id = str(loc_val).replace('.png', '').strip()
                            if loc_id and f"{loc_id}.png" not in ref_list:
                                ref_list.append(f"{loc_id}.png")

                    if ref_list:
                        reference_files = json.dumps(ref_list)

                # Xac dinh output folder
                # Characters (nv*) and Locations (loc*) -> nv/ folder
                if pid_str.startswith('nv') or pid_str.startswith('loc'):
                    out_path = proj_dir / "nv" / f"{pid_str}.png"
                else:
                    out_path = proj_dir / "img" / f"{pid_str}.png"

                prompts.append({
                    'id': pid_str,
                    'prompt': str(prompt).strip(),
                    'output_path': str(out_path),
                    'sheet': sheet_name,
                    'reference_files': str(reference_files).strip() if reference_files else "",
                    'nv_path': str(proj_dir / "nv"),  # Path to reference images folder
                    'video_prompt': str(video_prompt).strip() if video_prompt else ""  # For I2V
                })
                count += 1

            self.log(f"  -> Loaded {count} prompts from '{sheet_name}'")

        self.log(f"TONG CONG: {len(prompts)} prompts")
        return prompts

    def stop(self):
        """Dung."""
        self.stop_flag = True
        self._stop_video_worker()

    # =========================================================================
    # VIDEO GENERATION (Parallel with Image Gen)
    # =========================================================================

    def _load_video_settings(self, proj_dir: Path = None):
        """Load video generation settings from config + project cache."""
        try:
            import yaml
            import json
            settings_path = self.config_dir / "settings.yaml"
            if settings_path.exists():
                with open(settings_path, 'r', encoding='utf-8') as f:
                    settings = yaml.safe_load(f) or {}

                bearer_token = ''
                project_id = ''

                # 1. ƯU TIÊN: Token từ project cache (.media_cache.json)
                # Đây là token đã được cache khi tạo ảnh, dùng chung cho video
                if proj_dir:
                    cache_path = proj_dir / "prompts" / ".media_cache.json"
                    if cache_path.exists():
                        try:
                            with open(cache_path, 'r', encoding='utf-8') as f:
                                cache_data = json.load(f)
                            bearer_token = cache_data.get('_bearer_token', '')
                            project_id = cache_data.get('_project_id', '')
                            if bearer_token:
                                self.log(f"[VIDEO] Dùng token từ project cache")
                        except:
                            pass

                # 2. FALLBACK: Token từ profiles
                if not bearer_token:
                    for profile in self.profiles:
                        if profile.token and self.is_token_valid(profile):
                            bearer_token = profile.token
                            project_id = profile.project_id or ''
                            self.log(f"[VIDEO] Dùng token từ profile: {Path(profile.value).name}")
                            break

                # 3. FALLBACK: Token từ settings.yaml
                if not bearer_token:
                    bearer_token = settings.get('flow_bearer_token', '')
                    project_id = settings.get('flow_project_id', '')
                    if bearer_token:
                        self.log(f"[VIDEO] Dùng token từ settings.yaml")

                self._video_settings = {
                    'count': settings.get('video_count', '10'),  # Default 10 images → video
                    'model': settings.get('video_model', 'fast'),
                    'replace_image': settings.get('video_replace_image', True),
                    'bearer_token': bearer_token,
                    'project_id': project_id,
                    'proxy_token': settings.get('proxy_api_token', '')
                }

                # Parse count
                count_str = str(self._video_settings['count']).strip().lower()
                if count_str == 'full':
                    self._video_settings['count_num'] = -1  # -1 = full
                else:
                    try:
                        self._video_settings['count_num'] = int(count_str)
                    except ValueError:
                        self._video_settings['count_num'] = 0

                return self._video_settings['count_num'] != 0
        except Exception as e:
            self.log(f"Load video settings error: {e}", "WARN")

        return False

    def _start_video_worker(self, proj_dir: Path):
        """Start video generation worker thread."""
        if self._video_worker_running:
            return

        # Load settings với proj_dir để đọc được project cache
        if not self._load_video_settings(proj_dir):
            self.log("[VIDEO] Video generation disabled (count = 0)", "INFO")
            return

        # Kiểm tra token: ưu tiên prefetched → đã load từ cache
        if not self._video_settings.get('bearer_token'):
            # Thử dùng prefetched token từ parallel batch mode
            if hasattr(self, '_prefetched_token') and self._prefetched_token:
                self._video_settings['bearer_token'] = self._prefetched_token.get('token', '')
                self._video_settings['project_id'] = self._prefetched_token.get('project_id', '')
                self.log("[VIDEO] Dùng pre-fetched token cho video generation")

        if not self._video_settings.get('bearer_token'):
            self.log("[VIDEO] Chưa có token - sẽ tạo video sau khi có ảnh", "INFO")
            return

        self._video_worker_running = True
        self._video_results = {"success": 0, "failed": 0, "pending": 0}

        self._video_worker_thread = threading.Thread(
            target=self._video_worker_loop,
            args=(proj_dir,),
            daemon=True
        )
        self._video_worker_thread.start()
        self.log("[VIDEO] Video worker started (parallel with image gen)")

    def _stop_video_worker(self):
        """Stop video generation worker."""
        self._video_worker_running = False
        if self._video_worker_thread:
            self._video_worker_thread.join(timeout=5)
            self._video_worker_thread = None

    def _queue_video_generation(self, image_path: Path, image_id: str, video_prompt: str = ""):
        """Add image to video generation queue."""
        if not self._video_worker_running:
            return

        # Only queue scene images (not nv/loc)
        if image_id.startswith('nv') or image_id.startswith('loc'):
            return

        with self._video_queue_lock:
            # Check count limit
            count_num = self._video_settings.get('count_num', 0)
            current_queued = len(self._video_queue) + self._video_results['success'] + self._video_results['failed']

            if count_num != -1 and current_queued >= count_num:
                return  # Limit reached

            self._video_queue.append({
                'image_path': image_path,
                'image_id': image_id,
                'video_prompt': video_prompt
            })
            self._video_results['pending'] = len(self._video_queue)
            self.log(f"[VIDEO] Queued: {image_id} (pending: {len(self._video_queue)})")

    def _video_worker_loop(self, proj_dir: Path):
        """Video generation worker loop."""
        from modules.image_to_video import ImageToVideoConverter

        self.log("[VIDEO] Worker loop started")

        # Create converter
        try:
            converter = ImageToVideoConverter(
                project_path=str(proj_dir),
                bearer_token=self._video_settings['bearer_token'],
                project_id=self._video_settings['project_id'] or 'default',
                proxy_token=self._video_settings.get('proxy_token'),
                use_proxy=bool(self._video_settings.get('proxy_token')),
                video_model=self._video_settings.get('model', 'fast'),
                log_callback=lambda msg, lvl: self.log(f"[VIDEO] {msg}", lvl.upper())
            )
        except Exception as e:
            self.log(f"[VIDEO] Failed to create converter: {e}", "ERROR")
            self._video_worker_running = False
            return

        while self._video_worker_running and not self.stop_flag:
            # Get next item from queue
            item = None
            with self._video_queue_lock:
                if self._video_queue:
                    item = self._video_queue.pop(0)
                    self._video_results['pending'] = len(self._video_queue)

            if not item:
                time.sleep(1)  # Wait for new items
                continue

            image_path = item['image_path']
            image_id = item['image_id']
            video_prompt = item.get('video_prompt', '')

            self.log(f"[VIDEO] Processing: {image_id}")

            try:
                result = converter.convert_image_to_video(
                    image_path=image_path,
                    prompt=video_prompt,
                    replace_image=self._video_settings.get('replace_image', True)
                )

                if result.is_completed:
                    self._video_results['success'] += 1
                    self.log(f"[VIDEO] OK: {image_id} -> {result.video_path.name}")
                else:
                    self._video_results['failed'] += 1
                    self.log(f"[VIDEO] FAILED: {image_id} - {result.error}", "ERROR")

            except Exception as e:
                self._video_results['failed'] += 1
                self.log(f"[VIDEO] Error {image_id}: {e}", "ERROR")

            # Delay between videos
            time.sleep(2)

        self.log(f"[VIDEO] Worker stopped. Results: {self._video_results['success']} OK, {self._video_results['failed']} failed")

    def get_video_results(self) -> Dict:
        """Get video generation results."""
        return self._video_results.copy()


# ============================================================================
# SIMPLE API
# ============================================================================

def run_auto(input_path: str, callback: Callable = None) -> Dict:
    """
    1 HAM DUY NHAT - Chay tat ca tu dong.

    Args:
        input_path: File voice (.mp3, .wav) hoac Excel (.xlsx)
        callback: Ham nhan log messages

    Returns:
        Dict: {"success": N, "failed": M}
    """
    engine = SmartEngine()
    return engine.run(input_path, callback=callback)


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        result = run_auto(sys.argv[1], callback=print)
        print(f"\nResult: {result}")
    else:
        print("Usage: python smart_engine.py <voice.mp3>")
