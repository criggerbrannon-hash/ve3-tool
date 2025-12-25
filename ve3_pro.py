#!/usr/bin/env python3
"""
Uni-x Voice to Video v2.3
=========================
Beautiful, Smart, Powerful
1 Click: Voice → Video Images

v2.3 Updates:
- Rebranded to Uni-x Voice to Video
- New color scheme and logo
- Hide console window on Windows

v2.2 Updates:
- Unified Preview & Edit tab (combined preview + prompts)
- Settings dialog with all tools (token, prompts template, config)
- Auto-refresh preview during processing
- Progress with time estimation
- Detail panel with reference + result images

v2.1 Updates:
- Improved UI layout (scrollable controls)
- Preview tab: edit prompts, regenerate images
- Prompts tab: inline editing
"""

import os
import sys
import json
import shutil
import threading
import webbrowser
import time
import io
import urllib.request
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

# Setup
ROOT_DIR = Path(__file__).parent
sys.path.insert(0, str(ROOT_DIR))

# Support external config/projects directories (for auto-update setup)
CONFIG_DIR = Path(os.environ.get('VE3_CONFIG_DIR', ROOT_DIR / "config"))
PROJECTS_DIR = Path(os.environ.get('VE3_PROJECTS_DIR', ROOT_DIR / "PROJECTS"))


def get_git_info():
    """Get git commit info: hash, date, message."""
    import subprocess
    git_dir = ROOT_DIR / ".git"
    if not git_dir.exists():
        return None

    try:
        # Get commit info: hash, unix timestamp, message
        # %ct = unix timestamp (seconds since epoch)
        result = subprocess.run(
            ["git", "log", "-1", "--format=%h|%ct|%s"],
            cwd=ROOT_DIR, capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            parts = result.stdout.strip().split("|")
            if len(parts) >= 3:
                # Convert unix timestamp to local time
                timestamp = int(parts[1])
                local_time = datetime.fromtimestamp(timestamp)
                date_str = local_time.strftime("%Y-%m-%d %H:%M")

                return {
                    "hash": parts[0],
                    "date": date_str,  # Gio theo may tinh
                    "message": parts[2][:50]
                }
    except:
        pass
    return None


def auto_update_from_git():
    """Auto pull latest code from git if available."""
    import subprocess
    git_dir = ROOT_DIR / ".git"
    if not git_dir.exists():
        return False, "Not a git repo"

    # Doc branch tu file config (de de dang chuyen session)
    branch_file = ROOT_DIR / "config" / "current_branch.txt"
    if branch_file.exists():
        TARGET_BRANCH = branch_file.read_text(encoding='utf-8').strip()
    else:
        TARGET_BRANCH = "main"  # Fallback to main

    try:
        # Fetch and reset to target branch
        subprocess.run(
            ["git", "fetch", "origin", TARGET_BRANCH],
            cwd=ROOT_DIR, capture_output=True, text=True, timeout=30
        )

        result = subprocess.run(
            ["git", "reset", "--hard", f"origin/{TARGET_BRANCH}"],
            cwd=ROOT_DIR, capture_output=True, text=True, timeout=30
        )

        if result.returncode == 0:
            return True, f"Updated to {TARGET_BRANCH}"
        else:
            return False, result.stderr.strip()[:100]
    except Exception as e:
        return False, str(e)[:100]


# Auto-update on startup
_update_ok, _update_msg = auto_update_from_git()
GIT_INFO = get_git_info()  # Store for GUI display

if _update_ok:
    print(f"[Auto-Update] {_update_msg}")
else:
    print(f"[Auto-Update] Skip: {_update_msg}")

if GIT_INFO:
    print(f"[Version] {GIT_INFO['hash']} - {GIT_INFO['date']} - {GIT_INFO['message']}")

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except:
    HAS_PIL = False


class UnixVoiceToVideo:
    """Uni-x Voice to Video - Beautiful GUI."""

    VERSION = "2.3"
    APP_NAME = "Uni-x Voice to Video"
    LOGO_URL = "https://cdn-new.topcv.vn/unsafe/https://static.topcv.vn/company_logos/682bdc2e715781747704878.jpg"

    # Color scheme - Modern Purple/Blue gradient theme
    COLORS = {
        'primary': '#6366f1',      # Indigo
        'primary_dark': '#4f46e5', # Darker indigo
        'secondary': '#8b5cf6',    # Purple
        'accent': '#06b6d4',       # Cyan
        'success': '#10b981',      # Emerald
        'warning': '#f59e0b',      # Amber
        'error': '#ef4444',        # Red
        'bg': '#0f172a',           # Slate 900
        'bg_light': '#1e293b',     # Slate 800
        'bg_card': '#334155',      # Slate 700
        'text': '#f8fafc',         # Slate 50
        'text_muted': '#94a3b8',   # Slate 400
    }

    def __init__(self):
        self.root = tk.Tk()
        # Title with version and last update time
        title = self.APP_NAME
        if GIT_INFO:
            title += f"  |  Updated: {GIT_INFO['date']}  ({GIT_INFO['hash']})"
        self.root.title(title)
        self.root.geometry("1280x800")
        self.root.minsize(1100, 700)

        # Set icon from URL
        self._set_icon()

        # Config background - dark theme
        self.root.configure(bg=self.COLORS['bg'])

        # Variables
        self.input_mode = tk.StringVar(value="file")
        self.input_path = tk.StringVar()

        # Batch mode paths (relative to tool's parent folder)
        # Tool: D:\AUTO\ve3-tool → Parent: D:\AUTO
        # Voice: D:\AUTO\voice, Done: D:\AUTO\done
        self.batch_voice_folder = ROOT_DIR.parent / "voice"
        self.batch_done_folder = ROOT_DIR.parent / "done"

        # Data
        self.profiles: List[str] = []
        self.groq_keys: List[str] = []
        self.gemini_keys: List[str] = []
        self.deepseek_keys: List[str] = []
        self.chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        
        # State
        self._running = False
        self._stop = False
        self._engine = None

        # Progress tracking
        self._start_time = None
        self._items_done = 0
        self._items_total = 0
        self._auto_refresh_id = None

        # Current project data
        self.current_project_dir: Optional[Path] = None
        self.characters: List[Dict] = []
        self.scenes: List[Dict] = []
        
        # Image cache
        self.image_cache = {}
        
        # Load config
        self.load_config()
        
        # Create UI
        self.setup_styles()
        self.create_ui()
        
        # Initial state
        self.update_resource_display()

    def _set_icon(self):
        """Set application icon from URL."""
        if not HAS_PIL:
            return
        try:
            # Download icon
            with urllib.request.urlopen(self.LOGO_URL, timeout=5) as response:
                img_data = response.read()
            img = Image.open(io.BytesIO(img_data))
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            self._icon_photo = ImageTk.PhotoImage(img)
            self.root.iconphoto(True, self._icon_photo)
        except Exception as e:
            print(f"Could not load icon: {e}")

    def setup_styles(self):
        """Setup ttk styles with modern dark theme."""
        style = ttk.Style()

        # Use clam theme as base
        style.theme_use('clam')

        C = self.COLORS

        # Configure main theme colors
        style.configure('.',
            font=('Segoe UI', 10),
            background=C['bg_light'],
            foreground=C['text'])

        # Title style
        style.configure('Title.TLabel',
            font=('Segoe UI', 22, 'bold'),
            foreground=C['primary'],
            background=C['bg_light'])
        style.configure('Subtitle.TLabel',
            font=('Segoe UI', 11),
            foreground=C['text_muted'],
            background=C['bg_light'])

        style.configure('Section.TLabelframe.Label',
            font=('Segoe UI', 10, 'bold'),
            foreground=C['accent'],
            background=C['bg_light'])

        style.configure('TLabelframe',
            background=C['bg_light'],
            bordercolor=C['bg_card'])

        style.configure('TFrame', background=C['bg_light'])
        style.configure('TLabel', background=C['bg_light'], foreground=C['text'])

        # Button styles
        style.configure('TButton',
            font=('Segoe UI', 10),
            background=C['bg_card'],
            foreground=C['text'],
            borderwidth=0,
            padding=(10, 6))
        style.map('TButton',
            background=[('active', C['primary']), ('pressed', C['primary_dark'])],
            foreground=[('active', 'white')])

        style.configure('Accent.TButton',
            font=('Segoe UI', 10, 'bold'),
            background=C['primary'],
            foreground='white')
        style.map('Accent.TButton',
            background=[('active', C['primary_dark'])])

        style.configure('Big.TButton',
            font=('Segoe UI', 14, 'bold'),
            padding=(20, 15))

        style.configure('Status.TLabel',
            font=('Segoe UI', 9),
            background=C['bg_light'])

        # Entry style
        style.configure('TEntry',
            fieldbackground=C['bg_card'],
            foreground=C['text'],
            insertcolor=C['text'])

        # Combobox
        style.configure('TCombobox',
            fieldbackground=C['bg_card'],
            background=C['bg_card'],
            foreground=C['text'])

        # Notebook (tabs)
        style.configure('TNotebook',
            background=C['bg'],
            borderwidth=0)
        style.configure('TNotebook.Tab',
            background=C['bg_card'],
            foreground=C['text_muted'],
            padding=(15, 8),
            font=('Segoe UI', 10))
        style.map('TNotebook.Tab',
            background=[('selected', C['primary'])],
            foreground=[('selected', 'white')])

        # Treeview
        style.configure('Treeview',
            background=C['bg_card'],
            foreground=C['text'],
            fieldbackground=C['bg_card'],
            borderwidth=0)
        style.configure('Treeview.Heading',
            background=C['bg_light'],
            foreground=C['accent'],
            font=('Segoe UI', 9, 'bold'))
        style.map('Treeview',
            background=[('selected', C['primary'])],
            foreground=[('selected', 'white')])

        # Progressbar
        style.configure('TProgressbar',
            background=C['primary'],
            troughcolor=C['bg_card'],
            borderwidth=0,
            thickness=8)

        # Scrollbar
        style.configure('TScrollbar',
            background=C['bg_card'],
            troughcolor=C['bg_light'],
            borderwidth=0)

        # Radiobutton
        style.configure('TRadiobutton',
            background=C['bg_light'],
            foreground=C['text'])
    
    def create_ui(self):
        """Create main UI layout."""
        # Main container
        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # === LEFT PANEL (Controls) ===
        left_panel = ttk.Frame(main, width=350)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 15))
        left_panel.pack_propagate(False)
        
        self.create_controls(left_panel)
        
        # === RIGHT PANEL (Preview & Log) ===
        right_panel = ttk.Frame(main)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.create_preview(right_panel)
    
    def create_controls(self, parent):
        """Create left control panel with scrollable area."""

        C = self.COLORS

        # Create canvas with scrollbar for scrollable controls
        canvas = tk.Canvas(parent, highlightthickness=0, bg=C['bg_light'])
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor=tk.NW, width=335)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # === HEADER ===
        header = ttk.Frame(scrollable_frame)
        header.pack(fill=tk.X, pady=(0, 15), padx=5)

        ttk.Label(header, text="🎬 Uni-x", style='Title.TLabel').pack(anchor=tk.W)
        ttk.Label(header, text="Voice → Video (1 Click)", style='Subtitle.TLabel').pack(anchor=tk.W)

        # Version info - visible in UI
        if GIT_INFO:
            version_text = f"v{self.VERSION} • {GIT_INFO['date']} ({GIT_INFO['hash']})"
            ttk.Label(header, text=version_text, foreground=self.COLORS['text_muted'],
                     font=('Segoe UI', 8)).pack(anchor=tk.W, pady=(2, 0))

        # === 1. INPUT ===
        input_frame = ttk.LabelFrame(scrollable_frame, text=" 📁 Đầu vào ", padding=10)
        input_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        # Mode selection - Row 1
        mode_row = ttk.Frame(input_frame)
        mode_row.pack(fill=tk.X, pady=(0, 4))

        ttk.Radiobutton(mode_row, text="📄 File đơn", variable=self.input_mode,
                        value="file", command=self.on_mode_change).pack(side=tk.LEFT)
        ttk.Radiobutton(mode_row, text="📂 Thư mục", variable=self.input_mode,
                        value="folder", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)

        # Mode selection - Row 2 (Auto Batch)
        mode_row2 = ttk.Frame(input_frame)
        mode_row2.pack(fill=tk.X, pady=(0, 8))

        ttk.Radiobutton(mode_row2, text="🔄 Auto Batch (../voice → ../done)", variable=self.input_mode,
                        value="batch", command=self.on_mode_change).pack(side=tk.LEFT)

        # Path row
        path_row = ttk.Frame(input_frame)
        path_row.pack(fill=tk.X)

        self.path_entry = ttk.Entry(path_row, textvariable=self.input_path, font=('Consolas', 9))
        self.path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        ttk.Button(path_row, text="Chọn...", width=8, command=self.browse_input).pack(side=tk.LEFT)

        # Input info
        self.input_info_label = ttk.Label(input_frame, text="Hỗ trợ: .mp3, .wav, .xlsx",
                                          foreground='gray', font=('Segoe UI', 9))
        self.input_info_label.pack(anchor=tk.W, pady=(5, 0))

        # === 2. START BUTTON ===
        self.start_btn = tk.Button(
            scrollable_frame, text="▶  BẮT ĐẦU",
            font=('Segoe UI', 14, 'bold'),
            bg=C['primary'], fg='white', activebackground=C['primary_dark'],
            relief=tk.FLAT, cursor='hand2',
            command=self.start_processing
        )
        self.start_btn.pack(fill=tk.X, pady=10, ipady=12, padx=5)

        # Stop button
        self.stop_btn = tk.Button(
            scrollable_frame, text="⏹  Dừng",
            font=('Segoe UI', 10),
            bg=C['error'], fg='white', activebackground='#dc2626',
            relief=tk.FLAT, state=tk.DISABLED,
            command=self.stop_processing
        )
        self.stop_btn.pack(fill=tk.X, pady=(0, 10), padx=5)

        # === 3. PROGRESS ===
        progress_frame = ttk.LabelFrame(scrollable_frame, text=" 📊 Tiến độ ", padding=10)
        progress_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))

        self.progress_label = ttk.Label(progress_frame, text="Sẵn sàng", font=('Segoe UI', 10, 'bold'))
        self.progress_label.pack(anchor=tk.W)

        self.progress_detail = ttk.Label(progress_frame, text="", foreground='gray')
        self.progress_detail.pack(anchor=tk.W)

        # === 4. RESOURCES ===
        res_frame = ttk.LabelFrame(scrollable_frame, text=" 🔧 Tài nguyên ", padding=10)
        res_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        self.res_profiles = ttk.Label(res_frame, text="👤 Profiles: 0")
        self.res_profiles.pack(anchor=tk.W)

        # API keys (theo thu tu uu tien)
        self.res_gemini = ttk.Label(res_frame, text="🔑 Gemini: 0")
        self.res_gemini.pack(anchor=tk.W)

        self.res_groq = ttk.Label(res_frame, text="🔑 Groq: 0")
        self.res_groq.pack(anchor=tk.W)

        self.res_deepseek = ttk.Label(res_frame, text="🔑 DeepSeek: 0")
        self.res_deepseek.pack(anchor=tk.W)

        # Config buttons row 1
        btn_row = ttk.Frame(res_frame)
        btn_row.pack(fill=tk.X, pady=(8, 0))

        ttk.Button(btn_row, text="⚙️ Cài đặt", command=self.open_settings).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="🔄 Reload", command=self.reload_config).pack(side=tk.LEFT)

        # === 5. QUICK ACTIONS (simplified) ===
        actions_frame = ttk.LabelFrame(scrollable_frame, text=" ⚡ Thao tác nhanh ", padding=10)
        actions_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        ttk.Button(actions_frame, text="📂 Mở Output", command=self.open_output_folder).pack(fill=tk.X)
    
    def create_preview(self, parent):
        """Create right preview panel - unified view."""

        # Notebook for tabs
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # === TAB 1: UNIFIED PREVIEW (combined Preview + Prompts) ===
        preview_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(preview_tab, text="  🖼️ Preview & Edit  ")

        # Top bar: Progress + Actions
        top_bar = ttk.Frame(preview_tab)
        top_bar.pack(fill=tk.X, pady=(0, 10))

        self.thumb_progress = ttk.Label(top_bar, text="Chọn file để bắt đầu", font=('Segoe UI', 10, 'bold'))
        self.thumb_progress.pack(side=tk.LEFT)

        ttk.Button(top_bar, text="🔄 Tạo lại chưa xong", command=self.regenerate_all_pending).pack(side=tk.RIGHT, padx=(5, 0))

        # Main content: Left (table) + Right (detail)
        main_paned = ttk.PanedWindow(preview_tab, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)

        # LEFT: Table list
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)

        # Unified treeview for all items (chars + scenes)
        cols = ('id', 'type', 'prompt', 'status')
        self.main_tree = ttk.Treeview(left_frame, columns=cols, show='headings', selectmode='browse')
        self.main_tree.heading('id', text='ID')
        self.main_tree.heading('type', text='Loại')
        self.main_tree.heading('prompt', text='Prompt (double-click để sửa)')
        self.main_tree.heading('status', text='')
        self.main_tree.column('id', width=60, anchor=tk.CENTER)
        self.main_tree.column('type', width=50, anchor=tk.CENTER)
        self.main_tree.column('prompt', width=300)
        self.main_tree.column('status', width=30, anchor=tk.CENTER)
        self.main_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tree_scroll = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.main_tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.main_tree.configure(yscrollcommand=tree_scroll.set)

        # Bind events
        self.main_tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        self.main_tree.bind('<Double-1>', self.on_tree_double_click)

        # RIGHT: Detail panel
        right_frame = ttk.LabelFrame(main_paned, text=" Chi tiết ", padding=10)
        main_paned.add(right_frame, weight=1)

        # Detail: Image preview
        img_frame = ttk.Frame(right_frame)
        img_frame.pack(fill=tk.BOTH, expand=True)

        # Reference image (for scenes)
        self.ref_frame = ttk.LabelFrame(img_frame, text="Tham chiếu", padding=5)
        self.ref_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        self.ref_image_label = ttk.Label(self.ref_frame, text="N/A", anchor=tk.CENTER, background='#ecf0f1')
        self.ref_image_label.pack(fill=tk.BOTH, expand=True)

        # Result image
        self.result_frame = ttk.LabelFrame(img_frame, text="Kết quả", padding=5)
        self.result_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.result_image_label = ttk.Label(self.result_frame, text="Chưa có", anchor=tk.CENTER, background='#ecf0f1')
        self.result_image_label.pack(fill=tk.BOTH, expand=True)

        # Detail: Prompt editor
        ttk.Label(right_frame, text="Prompt:", font=('Segoe UI', 9, 'bold')).pack(anchor=tk.W, pady=(10, 0))
        self.detail_prompt_text = tk.Text(right_frame, height=4, wrap=tk.WORD, font=('Segoe UI', 9), bg='#f9f9f9')
        self.detail_prompt_text.pack(fill=tk.X, pady=(5, 10))

        # Detail: Action buttons
        action_row = ttk.Frame(right_frame)
        action_row.pack(fill=tk.X)
        ttk.Button(action_row, text="💾 Lưu prompt", command=self.save_current_prompt).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_row, text="🔄 Tạo lại ảnh", command=self.regenerate_current_image).pack(side=tk.LEFT)
        self.detail_status = ttk.Label(action_row, text="", foreground='gray')
        self.detail_status.pack(side=tk.RIGHT)

        # Store current selection
        self._current_item_id = None
        self._current_item_type = None  # 'char' or 'scene'

        # Keep old variables for compatibility
        self.char_combo = None
        self.scene_combo = None
        self.char_tree = self.main_tree  # Alias for compatibility
        self.scene_tree = self.main_tree  # Alias for compatibility

        # === TAB 2: LOG ===
        log_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(log_tab, text="  📝 Log  ")

        self.log_text = tk.Text(log_tab, wrap=tk.WORD, font=('Consolas', 9), bg='#1e1e1e', fg='#d4d4d4')
        self.log_text.pack(fill=tk.BOTH, expand=True)

        log_scroll = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=log_scroll.set)

        log_btn_row = ttk.Frame(log_tab)
        log_btn_row.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(log_btn_row, text="🗑️ Xóa", command=self.clear_log).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(log_btn_row, text="💾 Lưu", command=self.save_log).pack(side=tk.LEFT)
    
    # ========== ACTIONS ==========
    
    def on_mode_change(self):
        """Handle mode change."""
        mode = self.input_mode.get()

        if mode == "batch":
            # Auto Batch mode - show voice folder path
            self.input_path.set(str(self.batch_voice_folder))
            pending = self._count_pending_voices()
            self.input_info_label.config(
                text=f"📂 voice → done | {pending} file chờ xử lý"
            )
            # Disable path entry for batch mode
            self.path_entry.config(state='disabled')
        else:
            self.input_path.set("")
            self.path_entry.config(state='normal')
            if mode == "folder":
                self.input_info_label.config(text="Mỗi file voice trong thư mục = 1 dự án")
            else:
                self.input_info_label.config(text="Hỗ trợ: .mp3, .wav, .xlsx")

    def _count_pending_voices(self) -> int:
        """Count pending voice files (in voice folder but not in done folder)."""
        if not self.batch_voice_folder.exists():
            return 0

        pending = 0
        for subfolder in self.batch_voice_folder.iterdir():
            if subfolder.is_dir():
                for voice_file in subfolder.glob("*.mp3"):
                    # Check if video already exists (done/voice_stem/voice_stem_final.mp4)
                    done_video = self.batch_done_folder / voice_file.stem / f"{voice_file.stem}_final.mp4"
                    if not done_video.exists():
                        pending += 1
                for voice_file in subfolder.glob("*.wav"):
                    done_video = self.batch_done_folder / voice_file.stem / f"{voice_file.stem}_final.mp4"
                    if not done_video.exists():
                        pending += 1
        return pending
    
    def browse_input(self):
        """Browse for input."""
        if self.input_mode.get() == "folder":
            path = filedialog.askdirectory(title="Chọn thư mục chứa voice")
        else:
            path = filedialog.askopenfilename(
                title="Chọn file",
                filetypes=[
                    ("Supported", "*.mp3 *.wav *.xlsx"),
                    ("Audio", "*.mp3 *.wav"),
                    ("Excel", "*.xlsx"),
                    ("All", "*.*")
                ]
            )
        
        if path:
            self.input_path.set(path)
            self.update_input_info()
    
    def update_input_info(self):
        """Update input info display."""
        path = self.input_path.get()
        if not path:
            return
        
        p = Path(path)
        if self.input_mode.get() == "folder":
            voices = list(p.glob("*.mp3")) + list(p.glob("*.wav"))
            self.input_info_label.config(text=f"📁 {len(voices)} file voice")
        else:
            size = p.stat().st_size / 1024 if p.exists() else 0
            self.input_info_label.config(text=f"📄 {p.name} ({size:.1f} KB)")
    
    def log(self, msg: str, level: str = "INFO"):
        """Add log message."""
        ts = datetime.now().strftime("%H:%M:%S")
        
        # Color tags
        colors = {
            "OK": "#2ecc71",
            "ERROR": "#e74c3c", 
            "WARN": "#f39c12",
            "INFO": "#3498db"
        }
        
        prefix = {"OK": "✅", "ERROR": "❌", "WARN": "⚠️", "INFO": "ℹ️"}.get(level, "•")
        
        self.log_text.insert(tk.END, f"[{ts}] {prefix} {msg}\n")
        self.log_text.see(tk.END)
    
    def clear_log(self):
        self.log_text.delete(1.0, tk.END)
    
    def save_log(self):
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text", "*.txt")])
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(self.log_text.get(1.0, tk.END))
            self.log(f"Saved log to {path}", "OK")
    
    def update_progress(self, percent: float, text: str = "", detail: str = ""):
        """Update progress display."""
        self.progress_var.set(percent)
        if text:
            self.progress_label.config(text=text)
        if detail:
            self.progress_detail.config(text=detail)

    def update_progress_with_time(self, done: int, total: int, current_item: str = ""):
        """Update progress with time estimation."""
        if total <= 0:
            return

        self._items_done = done
        self._items_total = total

        # Calculate progress
        percent = (done / total) * 100
        self.progress_var.set(percent)

        # Calculate time estimation
        time_str = ""
        if self._start_time and done > 0:
            elapsed = time.time() - self._start_time
            avg_per_item = elapsed / done
            remaining_items = total - done
            eta_seconds = remaining_items * avg_per_item

            # Format ETA
            if eta_seconds < 60:
                time_str = f"~{int(eta_seconds)}s còn lại"
            elif eta_seconds < 3600:
                mins = int(eta_seconds / 60)
                secs = int(eta_seconds % 60)
                time_str = f"~{mins}m {secs}s còn lại"
            else:
                hours = int(eta_seconds / 3600)
                mins = int((eta_seconds % 3600) / 60)
                time_str = f"~{hours}h {mins}m còn lại"

        # Update labels
        self.progress_label.config(text=f"Tiến độ: {done}/{total} ({percent:.1f}%)")
        detail_text = current_item
        if time_str:
            detail_text = f"{current_item} | {time_str}" if current_item else time_str
        self.progress_detail.config(text=detail_text)

    def _start_auto_refresh(self):
        """Start auto-refresh timer for preview during processing."""
        self._start_time = time.time()

        def tick():
            if self._running:
                self.refresh_preview()
                self._auto_refresh_id = self.root.after(5000, tick)  # Refresh every 5s

        self._auto_refresh_id = self.root.after(5000, tick)

    def _stop_auto_refresh(self):
        """Stop auto-refresh timer."""
        if self._auto_refresh_id:
            self.root.after_cancel(self._auto_refresh_id)
            self._auto_refresh_id = None
        self._start_time = None
    
    def update_resource_display(self):
        """Update resource display."""
        self.res_profiles.config(text=f"👤 Profiles: {len(self.profiles)}" +
                                 (" ✅" if self.profiles else " ⚠️"))
        # API keys theo thu tu uu tien
        self.res_gemini.config(text=f"🔑 Gemini: {len(self.gemini_keys)}" +
                              (" ✅" if self.gemini_keys else ""))
        self.res_groq.config(text=f"🔑 Groq: {len(self.groq_keys)}" +
                            (" ✅" if self.groq_keys else ""))
        self.res_deepseek.config(text=f"🔑 DeepSeek: {len(self.deepseek_keys)}" +
                                (" ✅" if self.deepseek_keys else ""))
    
    # ========== CONFIG ==========
    
    def load_config(self):
        """Load config from chrome_profiles/ and accounts.json."""
        # Luôn scan thư mục chrome_profiles/ trước (tạo từ GUI)
        self.profiles = []
        profiles_dir = ROOT_DIR / "chrome_profiles"
        if profiles_dir.exists():
            for item in profiles_dir.iterdir():
                if item.is_dir() and not item.name.startswith('.'):
                    self.profiles.append(str(item))

        accounts_file = CONFIG_DIR / "accounts.json"

        if not accounts_file.exists():
            self.create_default_config()
            return

        try:
            with open(accounts_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.chrome_path = data.get('chrome_path', self.chrome_path)

            # Thêm profiles từ accounts.json (nếu chưa có trong chrome_profiles/)
            existing_paths = set(self.profiles)
            for p in data.get('chrome_profiles', []):
                path = p if isinstance(p, str) else p.get('path', '')
                if path and not path.startswith('THAY_BANG') and Path(path).exists():
                    if path not in existing_paths:
                        self.profiles.append(path)
            
            # API keys (thu tu uu tien: Ollama > Gemini > Groq > DeepSeek)
            api = data.get('api_keys', {})
            self.gemini_keys = [k for k in api.get('gemini', [])
                              if k and not k.startswith('THAY_BANG')]
            self.groq_keys = [k for k in api.get('groq', [])
                            if k and not k.startswith('THAY_BANG')]
            self.deepseek_keys = [k for k in api.get('deepseek', [])
                                if k and not k.startswith('THAY_BANG')]

            # Ollama local config
            ollama_config = api.get('ollama', {})
            self.ollama_model = ollama_config.get('model', 'qwen2.5:7b')
            self.ollama_endpoint = ollama_config.get('endpoint', 'http://localhost:11434')
            self.ollama_priority = ollama_config.get('priority', False)
            
        except Exception as e:
            print(f"Load config error: {e}")
    
    def create_default_config(self):
        """Create default config file."""
        CONFIG_DIR.mkdir(exist_ok=True)
        
        default = {
            "_README": [
                "=== VE3 TOOL CONFIG ===",
                "Dien thong tin Chrome profiles va API keys vao day",
                "Thu tu uu tien API: Gemini > Groq > DeepSeek",
                "Gemini: https://aistudio.google.com/app/apikey",
                "Groq (FREE): https://console.groq.com/keys",
                "DeepSeek: https://platform.deepseek.com/api_keys"
            ],
            "chrome_path": r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            "chrome_profiles": [
                "THAY_BANG_DUONG_DAN_PROFILE_1",
                "THAY_BANG_DUONG_DAN_PROFILE_2"
            ],
            "api_keys": {
                "gemini": [],
                "groq": [],
                "deepseek": []
            },
            "settings": {
                "parallel": 2,
                "delay_between_images": 2
            }
        }
        
        with open(CONFIG_DIR / "accounts.json", 'w', encoding='utf-8') as f:
            json.dump(default, f, indent=4, ensure_ascii=False)
    
    def reload_config(self):
        """Reload config."""
        self.load_config()
        self.update_resource_display()
        self.log("Đã reload config", "OK")
    
    def open_settings(self):
        """Open settings dialog with all tools."""
        win = tk.Toplevel(self.root)
        win.title("⚙️ Cài đặt")
        win.geometry("700x550")
        win.transient(self.root)
        win.grab_set()

        notebook = ttk.Notebook(win)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Browser Profiles (Enhanced)
        prof_tab = ttk.Frame(notebook, padding=15)
        notebook.add(prof_tab, text="  🌐 Trình duyệt  ")

        ttk.Label(prof_tab, text="Quản lý Chrome Profiles:",
                  font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        ttk.Label(prof_tab, text="Mỗi profile = 1 tài khoản Google riêng, chạy song song",
                  foreground='gray', font=('Segoe UI', 9)).pack(anchor=tk.W, pady=(0, 10))

        # Profile list
        prof_list_frame = ttk.Frame(prof_tab)
        prof_list_frame.pack(fill=tk.BOTH, expand=True)

        prof_list = tk.Listbox(prof_list_frame, height=6, font=('Consolas', 9))
        prof_scroll = ttk.Scrollbar(prof_list_frame, orient="vertical", command=prof_list.yview)
        prof_list.configure(yscrollcommand=prof_scroll.set)
        prof_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        prof_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh_profile_list():
            prof_list.delete(0, tk.END)
            profiles_dir = ROOT_DIR / "chrome_profiles"
            if profiles_dir.exists():
                for p in profiles_dir.iterdir():
                    if p.is_dir():
                        prof_list.insert(tk.END, p.name)
            if prof_list.size() == 0:
                prof_list.insert(tk.END, "(Chưa có profile - Ấn 'Thêm mới')")

        refresh_profile_list()

        # Headless toggle
        headless_var = tk.BooleanVar(value=self._get_headless_setting())
        headless_frame = ttk.Frame(prof_tab)
        headless_frame.pack(fill=tk.X, pady=(10, 5))
        ttk.Checkbutton(headless_frame, text="Chạy ẩn (Headless) - Khuyến nghị khi đã đăng nhập",
                        variable=headless_var, command=lambda: self._save_headless_setting(headless_var.get())
                        ).pack(side=tk.LEFT)

        # Generation Mode toggle (Chrome vs API)
        ttk.Separator(prof_tab, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(15, 10))
        ttk.Label(prof_tab, text="Chế độ tạo ảnh:",
                  font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        ttk.Label(prof_tab, text="Chrome: Tự động hóa trình duyệt  |  API: Gọi trực tiếp API (cần token)",
                  foreground='gray', font=('Segoe UI', 9)).pack(anchor=tk.W, pady=(0, 5))

        gen_mode_frame = ttk.Frame(prof_tab)
        gen_mode_frame.pack(fill=tk.X, pady=(5, 10))

        current_mode = self._get_generation_mode()
        gen_mode_var = tk.StringVar(value=current_mode)

        def on_mode_change():
            mode = gen_mode_var.get()
            self._save_generation_mode(mode)

        ttk.Radiobutton(gen_mode_frame, text="🌐 Chrome (Browser Automation)",
                        variable=gen_mode_var, value="chrome",
                        command=on_mode_change).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Radiobutton(gen_mode_frame, text="⚡ API (Direct - cần Bearer Token)",
                        variable=gen_mode_var, value="api",
                        command=on_mode_change).pack(side=tk.LEFT)

        # Parallel workers setting (for batch processing)
        ttk.Separator(prof_tab, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(15, 10))
        ttk.Label(prof_tab, text="Số luồng xử lý song song (Auto Batch):",
                  font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        ttk.Label(prof_tab, text="Nhiều luồng = nhanh hơn | Mỗi luồng cần 1 project riêng (token)",
                  foreground='gray', font=('Segoe UI', 9)).pack(anchor=tk.W, pady=(0, 5))

        parallel_frame = ttk.Frame(prof_tab)
        parallel_frame.pack(fill=tk.X, pady=(5, 10))

        current_parallel = self._get_parallel_workers()
        parallel_var = tk.IntVar(value=current_parallel)
        parallel_label = ttk.Label(parallel_frame, text=f"{current_parallel} luồng",
                                   font=('Segoe UI', 10, 'bold'), width=12)

        def update_parallel(val):
            num = int(float(val))
            parallel_var.set(num)
            parallel_label.config(text=f"{num} luồng")
            self._save_parallel_workers(num)

        ttk.Label(parallel_frame, text="1").pack(side=tk.LEFT)
        parallel_scale = ttk.Scale(parallel_frame, from_=1, to=10, orient=tk.HORIZONTAL,
                                   variable=parallel_var, command=update_parallel, length=180)
        parallel_scale.pack(side=tk.LEFT, padx=5)
        ttk.Label(parallel_frame, text="10").pack(side=tk.LEFT)
        parallel_label.pack(side=tk.LEFT, padx=10)

        # Buttons row 1
        prof_btn_row1 = ttk.Frame(prof_tab)
        prof_btn_row1.pack(fill=tk.X, pady=(5, 5))

        def add_new_profile():
            """Add new browser profile."""
            name = simpledialog.askstring("Thêm Profile",
                "Nhập tên profile (VD: account1, work, test...):",
                parent=win)
            if name:
                name = name.strip().replace(" ", "_")
                profiles_dir = ROOT_DIR / "chrome_profiles"
                profiles_dir.mkdir(exist_ok=True)
                profile_path = profiles_dir / name
                if profile_path.exists():
                    messagebox.showwarning("Cảnh báo", f"Profile '{name}' đã tồn tại!")
                    return
                profile_path.mkdir(exist_ok=True)
                refresh_profile_list()
                messagebox.showinfo("OK", f"Đã tạo profile '{name}'.\nẤn 'Mở đăng nhập' để login Google.")

        def open_profile_login():
            """Open browser for login."""
            sel = prof_list.curselection()
            if not sel:
                messagebox.showwarning("Chọn profile", "Vui lòng chọn 1 profile từ danh sách!")
                return
            profile_name = prof_list.get(sel[0])
            if profile_name.startswith("("):
                return

            profiles_dir = ROOT_DIR / "chrome_profiles"
            profile_path = profiles_dir / profile_name

            win.config(cursor="wait")
            win.update()

            try:
                self._open_browser_for_login(str(profile_path), profile_name)
            finally:
                win.config(cursor="")

        def delete_profile():
            """Delete selected profile."""
            sel = prof_list.curselection()
            if not sel:
                messagebox.showwarning("Chọn profile", "Vui lòng chọn 1 profile!")
                return
            profile_name = prof_list.get(sel[0])
            if profile_name.startswith("("):
                return

            if messagebox.askyesno("Xác nhận", f"Xóa profile '{profile_name}'?\nDữ liệu đăng nhập sẽ bị mất!"):
                import shutil
                profiles_dir = ROOT_DIR / "chrome_profiles"
                profile_path = profiles_dir / profile_name
                try:
                    shutil.rmtree(profile_path)
                    refresh_profile_list()
                    messagebox.showinfo("OK", f"Đã xóa profile '{profile_name}'")
                except Exception as e:
                    messagebox.showerror("Lỗi", f"Không thể xóa: {e}")

        def test_token_visible():
            """Test lấy token với Chrome hiển thị (không ẩn) để debug."""
            sel = prof_list.curselection()
            if not sel:
                messagebox.showwarning("Chọn profile", "Vui lòng chọn 1 profile từ danh sách!")
                return
            profile_name = prof_list.get(sel[0])
            if profile_name.startswith("("):
                return

            profiles_dir = ROOT_DIR / "chrome_profiles"
            profile_path = str(profiles_dir / profile_name)

            self.log(f"🔍 Test lấy token (KHÔNG ẨN) cho: {profile_name}")

            def run_test():
                try:
                    from modules.chrome_token_extractor import ChromeTokenExtractor

                    extractor = ChromeTokenExtractor(
                        chrome_path=self.chrome_path,
                        profile_path=profile_path,
                        headless=False,  # KHÔNG ẨN để debug
                        timeout=120
                    )

                    def log_cb(msg, level="info"):
                        self.root.after(0, lambda: self.log(f"[Test] {msg}", level.upper()))

                    token, proj_id, error = extractor.extract_token(callback=log_cb)

                    if token:
                        self.root.after(0, lambda: self.log(f"✅ Token OK! Length: {len(token)}", "OK"))
                        self.root.after(0, lambda: messagebox.showinfo("Thành công", f"Lấy token thành công!\n\nProfile: {profile_name}\nToken length: {len(token)}\nProject ID: {proj_id[:20] if proj_id else 'N/A'}..."))
                    else:
                        self.root.after(0, lambda: self.log(f"❌ Lỗi: {error}", "ERROR"))
                        self.root.after(0, lambda: messagebox.showerror("Lỗi", f"Không lấy được token!\n\nLỗi: {error}\n\nHãy thử:\n1. Mở đăng nhập lại\n2. Làm theo hướng dẫn allow pasting"))

                except Exception as e:
                    self.root.after(0, lambda: self.log(f"❌ Exception: {e}", "ERROR"))
                    self.root.after(0, lambda: messagebox.showerror("Lỗi", str(e)))

            import threading
            threading.Thread(target=run_test, daemon=True).start()

        ttk.Button(prof_btn_row1, text="➕ Thêm mới", command=add_new_profile).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(prof_btn_row1, text="🔓 Mở đăng nhập", command=open_profile_login).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(prof_btn_row1, text="🗑️ Xóa", command=delete_profile).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(prof_btn_row1, text="🔄", command=refresh_profile_list, width=3).pack(side=tk.LEFT)
        ttk.Button(prof_btn_row1, text="🧪 Test", command=test_token_visible).pack(side=tk.LEFT, padx=(5, 0))

        # Info
        ttk.Label(prof_tab, text="💡 Mỗi voice sẽ dùng 1 profile khác nhau khi chạy song song",
                  foreground='#666', font=('Segoe UI', 9)).pack(anchor=tk.W, pady=(10, 0))

        # Tab 2: API Keys
        api_tab = ttk.Frame(notebook, padding=15)
        notebook.add(api_tab, text="  🔑 API Keys  ")

        # Scroll frame for API keys
        api_canvas = tk.Canvas(api_tab, highlightthickness=0)
        api_scrollbar = ttk.Scrollbar(api_tab, orient="vertical", command=api_canvas.yview)
        api_scroll_frame = ttk.Frame(api_canvas)

        api_scroll_frame.bind("<Configure>", lambda e: api_canvas.configure(scrollregion=api_canvas.bbox("all")))
        api_canvas.create_window((0, 0), window=api_scroll_frame, anchor="nw")
        api_canvas.configure(yscrollcommand=api_scrollbar.set)

        # Header with priority note
        ttk.Label(api_scroll_frame, text="Thu tu uu tien: Gemini > Groq > DeepSeek",
                  foreground='gray', font=('Segoe UI', 9, 'italic')).pack(anchor=tk.W, pady=(0, 10))

        # Store entry references for saving
        api_entries = {}

        # 1. Gemini (highest priority)
        ttk.Label(api_scroll_frame, text="1. Gemini Keys (Uu tien cao nhat):",
                  font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        gem_link = ttk.Label(api_scroll_frame, text="🔗 aistudio.google.com/app/apikey",
                             foreground='blue', cursor='hand2')
        gem_link.pack(anchor=tk.W)
        gem_link.bind('<Button-1>', lambda e: webbrowser.open("https://aistudio.google.com/app/apikey"))

        gem_entry = tk.Text(api_scroll_frame, height=2, font=('Consolas', 9), wrap=tk.WORD)
        gem_entry.pack(fill=tk.X, pady=(5, 10))
        gem_entry.insert(tk.END, '\n'.join(self.gemini_keys))
        api_entries['gemini'] = gem_entry

        # 2. Groq (second priority)
        ttk.Label(api_scroll_frame, text="2. Groq Keys (FREE, nhanh):",
                  font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        groq_link = ttk.Label(api_scroll_frame, text="🔗 console.groq.com/keys",
                             foreground='blue', cursor='hand2')
        groq_link.pack(anchor=tk.W)
        groq_link.bind('<Button-1>', lambda e: webbrowser.open("https://console.groq.com/keys"))

        groq_entry = tk.Text(api_scroll_frame, height=2, font=('Consolas', 9), wrap=tk.WORD)
        groq_entry.pack(fill=tk.X, pady=(5, 10))
        groq_entry.insert(tk.END, '\n'.join(self.groq_keys))
        api_entries['groq'] = groq_entry

        # 3. DeepSeek (lowest priority, cheapest)
        ttk.Label(api_scroll_frame, text="3. DeepSeek Keys (Re, cham):",
                  font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        ds_link = ttk.Label(api_scroll_frame, text="🔗 platform.deepseek.com/api_keys",
                           foreground='blue', cursor='hand2')
        ds_link.pack(anchor=tk.W)
        ds_link.bind('<Button-1>', lambda e: webbrowser.open("https://platform.deepseek.com/api_keys"))

        ds_entry = tk.Text(api_scroll_frame, height=2, font=('Consolas', 9), wrap=tk.WORD)
        ds_entry.pack(fill=tk.X, pady=(5, 10))
        ds_entry.insert(tk.END, '\n'.join(self.deepseek_keys))
        api_entries['deepseek'] = ds_entry

        ttk.Label(api_scroll_frame, text="(Moi key 1 dong, Enter de xuong dong)",
                  foreground='gray', font=('Segoe UI', 8)).pack(anchor=tk.W)

        # Buttons
        btn_frame = ttk.Frame(api_scroll_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        def save_api_keys():
            """Save API keys to accounts.json."""
            try:
                config_file = CONFIG_DIR / "accounts.json"
                with open(config_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Update API keys
                if 'api_keys' not in data:
                    data['api_keys'] = {}

                for key_type, entry in api_entries.items():
                    text = entry.get("1.0", tk.END).strip()
                    keys = [k.strip() for k in text.split('\n') if k.strip()]
                    data['api_keys'][key_type] = keys

                with open(config_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)

                self.reload_config()
                messagebox.showinfo("OK", "Da luu API keys!")
            except Exception as e:
                messagebox.showerror("Loi", f"Khong the luu: {e}")

        def test_api_keys():
            """Test all API keys."""
            win.config(cursor="wait")
            win.update()

            results = []

            # Test Gemini
            for i, key in enumerate(self.gemini_keys):
                try:
                    from modules.ai_providers import GeminiClient
                    client = GeminiClient(key)
                    r = client.generate("Say OK", max_tokens=10)
                    status = "OK" if r else "FAIL"
                except Exception as e:
                    status = f"FAIL: {str(e)[:30]}"
                results.append(f"Gemini #{i+1}: {status}")

            # Test Groq
            for i, key in enumerate(self.groq_keys):
                try:
                    from modules.ai_providers import GroqClient
                    client = GroqClient(key)
                    r = client.generate("Say OK", max_tokens=10)
                    status = "OK" if r else "FAIL"
                except Exception as e:
                    status = f"FAIL: {str(e)[:30]}"
                results.append(f"Groq #{i+1}: {status}")

            # Test DeepSeek
            for i, key in enumerate(self.deepseek_keys):
                try:
                    from modules.ai_providers import DeepSeekClient
                    client = DeepSeekClient(key)
                    r = client.generate("Say OK", max_tokens=10)
                    status = "OK" if r else "FAIL"
                except Exception as e:
                    status = f"FAIL: {str(e)[:30]}"
                results.append(f"DeepSeek #{i+1}: {status}")

            win.config(cursor="")

            if results:
                messagebox.showinfo("Ket qua Test", '\n'.join(results))
            else:
                messagebox.showwarning("Chua co key", "Chua co API key nao de test!")

        ttk.Button(btn_frame, text="💾 Luu API Keys", command=save_api_keys).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_frame, text="🧪 Test APIs", command=test_api_keys).pack(side=tk.LEFT)

        api_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        api_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Tab 3: Token
        token_tab = ttk.Frame(notebook, padding=15)
        notebook.add(token_tab, text="  🔑 Token  ")

        ttk.Label(token_tab, text="Lấy Token thủ công",
                  font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        ttk.Label(token_tab, text="Dùng khi cần lấy token mới hoặc test profile.",
                  foreground='gray').pack(anchor=tk.W, pady=(0, 10))

        ttk.Button(token_tab, text="🔑 Lấy Token ngay",
                   command=lambda: [win.destroy(), self.get_token_manual()]).pack(anchor=tk.W)

        ttk.Separator(token_tab, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)

        ttk.Label(token_tab, text="Tokens đã cache:", font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        self._show_cached_tokens(token_tab)

        # Tab 4: Video Generation
        video_tab = ttk.Frame(notebook, padding=15)
        notebook.add(video_tab, text="  🎬 Video  ")

        ttk.Label(video_tab, text="Chuyển đổi Ảnh sang Video (I2V)",
                  font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        ttk.Label(video_tab, text="Sử dụng Google Veo 3 để chuyển ảnh thành video",
                  foreground='gray').pack(anchor=tk.W, pady=(0, 10))

        # Video count setting
        video_count_frame = ttk.Frame(video_tab)
        video_count_frame.pack(fill=tk.X, pady=(5, 10))

        ttk.Label(video_count_frame, text="Số ảnh chuyển video:").pack(side=tk.LEFT)

        video_count_var = tk.StringVar(value=self._get_video_count_setting())
        video_count_entry = ttk.Entry(video_count_frame, textvariable=video_count_var, width=10)
        video_count_entry.pack(side=tk.LEFT, padx=5)

        ttk.Label(video_count_frame, text="(số hoặc 'full' = tất cả)", foreground='gray').pack(side=tk.LEFT)

        # Video model setting
        video_model_frame = ttk.Frame(video_tab)
        video_model_frame.pack(fill=tk.X, pady=(5, 10))

        ttk.Label(video_model_frame, text="Model video:").pack(side=tk.LEFT)

        video_model_var = tk.StringVar(value=self._get_video_model_setting())
        ttk.Radiobutton(video_model_frame, text="Fast (nhanh)", variable=video_model_var, value="fast").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(video_model_frame, text="Quality (chất lượng)", variable=video_model_var, value="quality").pack(side=tk.LEFT)

        # Replace image option
        replace_var = tk.BooleanVar(value=self._get_video_replace_setting())
        ttk.Checkbutton(video_tab, text="Thay thế ảnh gốc bằng video (backup vào img_backup/)",
                        variable=replace_var).pack(anchor=tk.W, pady=(5, 10))

        def save_video_settings():
            """Save video generation settings."""
            try:
                settings_file = CONFIG_DIR / "settings.yaml"
                settings = {}
                if settings_file.exists():
                    import yaml
                    with open(settings_file, 'r', encoding='utf-8') as f:
                        settings = yaml.safe_load(f) or {}

                settings['video_count'] = video_count_var.get().strip()
                settings['video_model'] = video_model_var.get()
                settings['video_replace_image'] = replace_var.get()

                import yaml
                with open(settings_file, 'w', encoding='utf-8') as f:
                    yaml.dump(settings, f, allow_unicode=True, default_flow_style=False)

                messagebox.showinfo("OK", "Đã lưu cài đặt video!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu: {e}")

        ttk.Button(video_tab, text="💾 Lưu cài đặt Video", command=save_video_settings).pack(anchor=tk.W, pady=(10, 5))

        ttk.Separator(video_tab, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)

        ttk.Label(video_tab, text="💡 Lưu ý:", font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)
        ttk.Label(video_tab, text="• Cần Bearer Token (lấy từ tab Token)\n• Mỗi video mất 1-3 phút để tạo\n• Video được lưu vào thư mục video/\n• Prompt lấy từ cột 'video_prompt' trong Excel",
                  foreground='gray', justify=tk.LEFT).pack(anchor=tk.W)

        # Tab 5: Prompts Template
        prompts_tab = ttk.Frame(notebook, padding=15)
        notebook.add(prompts_tab, text="  📝 Prompts  ")

        ttk.Label(prompts_tab, text="Prompts Template",
                  font=('Segoe UI', 11, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        ttk.Label(prompts_tab, text="Chỉnh sửa prompt AI để tạo ảnh theo phong cách mong muốn.",
                  foreground='gray').pack(anchor=tk.W, pady=(0, 10))

        ttk.Button(prompts_tab, text="📝 Mở Prompts Editor",
                   command=lambda: [win.destroy(), self.open_prompts_editor()]).pack(anchor=tk.W, pady=(0, 10))

        prompts_file = CONFIG_DIR / "prompts.yaml"
        if prompts_file.exists():
            ttk.Label(prompts_tab, text=f"File: {prompts_file}",
                      foreground='gray', font=('Consolas', 9)).pack(anchor=tk.W)

        # Tab 5: Help
        help_tab = ttk.Frame(notebook, padding=15)
        notebook.add(help_tab, text="  ❓ Help  ")

        help_text = """🎯 CÁCH SỬ DỤNG:

1️⃣ Thêm Chrome Profile:
   • Mở chrome://version → copy "Profile Path"
   • Dán vào accounts.json

2️⃣ Thêm Groq API Key (FREE):
   • Vào console.groq.com/keys
   • Tạo key → dán vào accounts.json

3️⃣ Chạy Tool:
   • Chọn file voice → BẮT ĐẦU
   • Tool tự động: Voice → SRT → Prompts → Images

⚠️ LƯU Ý:
   • Đóng Chrome trước khi chạy
   • Không di chuột khi đang lấy token"""

        ttk.Label(help_tab, text=help_text, justify=tk.LEFT, font=('Segoe UI', 10)).pack(anchor=tk.W)

        # Close button
        ttk.Button(win, text="Đóng", command=win.destroy).pack(pady=10)

    def _show_cached_tokens(self, parent):
        """Show cached tokens info."""
        tokens_file = CONFIG_DIR / "tokens.json"
        if tokens_file.exists():
            try:
                import time
                with open(tokens_file, 'r') as f:
                    tokens = json.load(f)
                for profile, data in tokens.items():
                    token_time = data.get('token_time', 0)
                    age_mins = int((time.time() - token_time) / 60) if token_time else 999
                    status = "✅" if age_mins < 50 else "⚠️ expired"
                    name = Path(profile).name
                    ttk.Label(parent, text=f"  • {name}: {age_mins}m ago {status}",
                              font=('Consolas', 9)).pack(anchor=tk.W)
            except:
                ttk.Label(parent, text="  (Không đọc được)", foreground='gray').pack(anchor=tk.W)
        else:
            ttk.Label(parent, text="  (Chưa có token nào)", foreground='gray').pack(anchor=tk.W)
    
    def open_config_file(self):
        """Open config file in editor."""
        config_file = CONFIG_DIR / "accounts.json"
        
        if not config_file.exists():
            self.create_default_config()
        
        if sys.platform == "win32":
            os.startfile(str(config_file))
        else:
            import subprocess
            subprocess.Popen(["xdg-open", str(config_file)])
    
    def open_output_folder(self):
        """Open output folder."""
        PROJECTS_DIR.mkdir(exist_ok=True)

        if sys.platform == "win32":
            os.startfile(str(PROJECTS_DIR))
        else:
            import subprocess
            subprocess.Popen(["xdg-open", str(PROJECTS_DIR)])

    # ========== BROWSER PROFILE MANAGEMENT ==========

    def _get_headless_setting(self) -> bool:
        """Get headless setting from config."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                return config.get('browser_headless', True)
        except:
            pass
        return True  # Default: headless

    def _save_headless_setting(self, headless: bool):
        """Save headless setting to config."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            config = {}
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
            config['browser_headless'] = headless
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config, f, default_flow_style=False, allow_unicode=True)
            self.log(f"Headless mode: {'ON' if headless else 'OFF'}", "OK")
        except Exception as e:
            print(f"Save headless error: {e}")

    def _get_generation_mode(self) -> str:
        """Get generation mode from config: 'chrome' or 'api'."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                return config.get('generation_mode', 'api')
        except:
            pass
        return 'api'  # Default: API mode (user preference)

    def _save_generation_mode(self, mode: str):
        """Save generation mode to config: 'chrome' or 'api'."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            config = {}
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
            config['generation_mode'] = mode
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config, f, default_flow_style=False, allow_unicode=True)
            mode_name = "Chrome (Browser)" if mode == 'chrome' else "API (Direct)"
            self.log(f"Generation mode: {mode_name}", "OK")
        except Exception as e:
            print(f"Save generation_mode error: {e}")

    def _get_parallel_workers(self) -> int:
        """Get number of parallel workers from config."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                # Backwards compatible with old key
                value = config.get('parallel_workers', config.get('parallel_browsers', 3))
                return max(1, min(10, value))
        except:
            pass
        return 3  # Default: 3 workers

    def _save_parallel_workers(self, num: int):
        """Save number of parallel workers to config."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            config = {}
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
            config['parallel_workers'] = max(1, min(10, num))
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config, f, default_flow_style=False, allow_unicode=True)
            self.log(f"Parallel workers: {num}", "OK")
        except Exception as e:
            print(f"Save parallel_workers error: {e}")

    # ======= VIDEO SETTINGS =======
    def _get_video_count_setting(self) -> str:
        """Get video count setting (number or 'full')."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                return str(config.get('video_count', '10'))
        except:
            pass
        return '10'  # Default: 10 images to video

    def _get_video_model_setting(self) -> str:
        """Get video model setting ('fast' or 'quality')."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                return config.get('video_model', 'fast')
        except:
            pass
        return 'fast'

    def _get_video_replace_setting(self) -> bool:
        """Get video replace image setting."""
        try:
            import yaml
            config_path = CONFIG_DIR / "settings.yaml"
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                return config.get('video_replace_image', True)
        except:
            pass
        return True

    def _open_browser_for_login(self, profile_path: str, profile_name: str):
        """Open Chrome browser with profile for Google login."""
        FLOW_URL = "https://labs.google/fx/vi/tools/flow"

        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.by import By

            self.log(f"Mở trình duyệt cho profile: {profile_name}...")

            options = Options()
            options.add_argument(f"--user-data-dir={profile_path}")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)
            options.add_argument("--window-size=1280,900")

            # Random port to avoid conflicts
            import random
            debug_port = random.randint(9300, 9500)
            options.add_argument(f"--remote-debugging-port={debug_port}")

            driver = webdriver.Chrome(options=options)
            driver.execute_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )

            driver.get(FLOW_URL)
            self.log("Trình duyệt đã mở - Hãy đăng nhập Google!", "OK")
            self.log("Đóng trình duyệt khi hoàn tất đăng nhập.")

            # Show message with detailed instructions
            messagebox.showinfo(
                "Đăng nhập Google + Kích hoạt Token",
                f"Trình duyệt đã mở cho profile '{profile_name}'.\n\n"
                "📋 LÀM THEO CÁC BƯỚC SAU:\n\n"
                "1️⃣ Đăng nhập tài khoản Google\n"
                "2️⃣ Đợi trang Google Flow hiện lên\n"
                "3️⃣ Nhấn F12 để mở DevTools\n"
                "4️⃣ Chọn tab 'Console'\n"
                "5️⃣ Gõ: allow pasting  rồi Enter\n"
                "6️⃣ Paste lệnh: console.log('OK')  rồi Enter\n"
                "7️⃣ Đóng trình duyệt khi xong\n\n"
                "⚠️ Bước 5-6 cần làm 1 LẦN để cho phép paste code!"
            )

        except Exception as e:
            self.log(f"Lỗi mở trình duyệt: {e}", "ERROR")
            messagebox.showerror("Lỗi", f"Không thể mở trình duyệt:\n{e}\n\nCần cài selenium:\npip install selenium")

    # ========== MAIN PROCESSING ==========
    
    def start_processing(self):
        """Start main processing."""
        mode = self.input_mode.get()

        # Batch mode validation
        if mode == "batch":
            if not self.batch_voice_folder.exists():
                messagebox.showerror(
                    "Lỗi",
                    f"Không tìm thấy thư mục voice:\n{self.batch_voice_folder}\n\n"
                    "Vui lòng tạo thư mục này và thêm các file voice cần xử lý."
                )
                return

            pending = self._count_pending_voices()
            if pending == 0:
                messagebox.showinfo(
                    "Thông báo",
                    "Không có file voice mới cần xử lý!\n\n"
                    f"Thư mục voice: {self.batch_voice_folder}\n"
                    f"Thư mục done: {self.batch_done_folder}"
                )
                return
        else:
            # Normal mode validation
            path = self.input_path.get()

            if not path:
                messagebox.showerror("Lỗi", "Vui lòng chọn file hoặc thư mục đầu vào!")
                return

            if not Path(path).exists():
                messagebox.showerror("Lỗi", f"Không tìm thấy:\n{path}")
                return

        # Reload config
        self.load_config()

        if not self.profiles:
            result = messagebox.askyesno(
                "Thiếu Chrome Profile",
                "Chưa có Chrome profile nào!\n\n"
                "Bạn cần thêm profile vào file config.\n\n"
                "Mở file config ngay?"
            )
            if result:
                self.open_config_file()
            return

        # Check AI keys for voice (for non-batch modes)
        if mode != "batch":
            path = self.input_path.get()
            ext = Path(path).suffix.lower() if Path(path).is_file() else ""
            has_ai_keys = self.gemini_keys or self.groq_keys or self.deepseek_keys
            if ext in ['.mp3', '.wav'] and not has_ai_keys:
                result = messagebox.askyesno(
                    "Thieu AI API Key",
                    "Can Gemini, Groq hoac DeepSeek API key de xu ly voice!\n\n"
                    "Thu tu uu tien: Gemini > Groq (FREE) > DeepSeek\n\n"
                    "Mo Cai dat de nhap API keys?"
                )
                if result:
                    self.open_settings()
                return
        else:
            # Batch mode also needs AI keys
            has_ai_keys = self.gemini_keys or self.groq_keys or self.deepseek_keys
            if not has_ai_keys:
                result = messagebox.askyesno(
                    "Thieu AI API Key",
                    "Can Gemini, Groq hoac DeepSeek API key de xu ly voice!\n\n"
                    "Mo Cai dat de nhap API keys?"
                )
                if result:
                    self.open_settings()
                return

        # Start
        self._running = True
        self._stop = False
        self.start_btn.config(state=tk.DISABLED, bg=self.COLORS['bg_card'])
        self.stop_btn.config(state=tk.NORMAL)

        # Start auto-refresh for preview
        self._start_auto_refresh()

        self.clear_log()
        self.log("=" * 50)
        self.log("🚀 BẮT ĐẦU XỬ LÝ")
        self.log("=" * 50)

        if mode == "batch":
            self.log(f"📂 Mode: Auto Batch")
            self.log(f"   Voice: {self.batch_voice_folder}")
            self.log(f"   Done: {self.batch_done_folder}")
            threading.Thread(target=self._process_batch, daemon=True).start()
        elif mode == "folder":
            threading.Thread(target=self._process_folder, daemon=True).start()
        else:
            threading.Thread(target=self._process_single, daemon=True).start()
    
    def stop_processing(self):
        """Stop processing."""
        self._stop = True
        if self._engine:
            self._engine.stop()
        self.log("⏹️ Đang dừng...", "WARN")
    
    def _process_single(self):
        """Process single file in background thread."""
        try:
            from modules.smart_engine import SmartEngine
            
            path = self.input_path.get()
            
            engine = SmartEngine()
            self._engine = engine
            
            def log_cb(msg):
                # Parse level from message
                level = "INFO"
                if "[OK]" in msg or "OK!" in msg:
                    level = "OK"
                elif "[ERROR]" in msg or "ERROR" in msg:
                    level = "ERROR"
                elif "[WARN]" in msg:
                    level = "WARN"
                
                self.root.after(0, lambda: self.log(msg, level))
            
            results = engine.run(path, callback=log_cb)
            
            # Update UI
            self.root.after(0, lambda: self._on_complete(results))
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda err=e: self.log(f"Lỗi: {err}", "ERROR"))
            self.root.after(0, lambda: messagebox.showerror("Lỗi", str(e)))
        finally:
            self._running = False
            self.root.after(0, self._reset_ui)

    def _process_batch(self):
        """
        Process all pending voice files from batch_voice_folder.
        PARALLEL PROCESSING with pre-fetched tokens.

        Structure:
        - voice/AR16-T1/AR16-0035.mp3  →  done/AR16-T1/AR16-0035/...
        - voice/AR16-T1/AR16-0036.mp3  →  done/AR16-T1/AR16-0036/...

        Flow:
        1. Scan pending files
        2. Pre-fetch N tokens (1 per worker, serialized)
        3. Run N workers in parallel (each worker processes voices sequentially)
        """
        try:
            from modules.smart_engine import SmartEngine
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import queue
            import threading

            # Create done folder if not exists
            self.batch_done_folder.mkdir(parents=True, exist_ok=True)

            # Get settings
            num_workers = self._get_parallel_workers()

            # Scan for pending voice files
            pending_files = []
            for subfolder in sorted(self.batch_voice_folder.iterdir()):
                if not subfolder.is_dir():
                    continue

                for voice_file in sorted(subfolder.glob("*.mp3")) + sorted(subfolder.glob("*.wav")):
                    # Output folder = done/voice_stem (không có subfolder)
                    # Ví dụ: voice/ar34-t1/ar34-0023.mp3 → done/ar34-0023/
                    output_folder = self.batch_done_folder / voice_file.stem

                    # Check if video already exists
                    final_video = output_folder / f"{voice_file.stem}_final.mp4"
                    if final_video.exists():
                        continue

                    pending_files.append({
                        'voice_path': voice_file,
                        'output_folder': output_folder,
                        'subfolder': subfolder.name
                    })

            if not pending_files:
                self.root.after(0, lambda: self.log("✅ Không có file mới cần xử lý!", "OK"))
                return

            total = len(pending_files)
            actual_workers = min(num_workers, total)

            self.log(f"📋 Tìm thấy {total} file cần xử lý")
            self.log(f"⚡ Chế độ song song: {actual_workers} luồng")
            for i, f in enumerate(pending_files[:5]):
                self.log(f"   {i+1}. {f['subfolder']}/{f['voice_path'].name}")
            if total > 5:
                self.log(f"   ... và {total - 5} file khác")

            # ================================================================
            # STEP 1: PRE-FETCH TOKENS (serialized - 1 Chrome at a time)
            # ================================================================
            self.log("")
            self.log("=" * 50)
            self.log(f"🔑 PRE-FETCH {actual_workers} TOKENS...")
            self.log("=" * 50)

            prefetched_tokens = []
            token_lock = threading.Lock()

            # Get first profile for token extraction
            profiles_dir = ROOT_DIR / "chrome_profiles"
            chrome_profiles = []
            if profiles_dir.exists():
                chrome_profiles = [p for p in profiles_dir.iterdir() if p.is_dir()]

            if not chrome_profiles:
                self.log("❌ Không có Chrome profile! Thêm profile trong Cài đặt.", "ERROR")
                return

            profile_path = chrome_profiles[0]  # Use first profile for all token extractions
            self.log(f"   Dùng profile: {profile_path.name}")

            for i in range(actual_workers):
                if self._stop:
                    break

                self.log(f"   [{i+1}/{actual_workers}] Đang lấy token...")

                try:
                    # Create engine to get token
                    engine = SmartEngine()
                    engine.profiles = self.profiles
                    engine.chrome_path = self.chrome_path

                    # Get a new token (opens new project)
                    token_data = engine._prefetch_token_for_worker(str(profile_path), i)

                    if token_data and token_data.get('token'):
                        prefetched_tokens.append(token_data)
                        self.log(f"   [{i+1}/{actual_workers}] ✅ Token OK (project: {token_data.get('project_id', 'N/A')[:8]}...)")
                    else:
                        self.log(f"   [{i+1}/{actual_workers}] ⚠️ Token FAILED", "WARN")
                        prefetched_tokens.append(None)

                except Exception as e:
                    self.log(f"   [{i+1}/{actual_workers}] ❌ Error: {e}", "ERROR")
                    prefetched_tokens.append(None)

            valid_tokens = [t for t in prefetched_tokens if t]
            self.log(f"🔑 Pre-fetched: {len(valid_tokens)}/{actual_workers} tokens OK")

            if not valid_tokens:
                self.log("❌ Không lấy được token nào! Kiểm tra Chrome profile.", "ERROR")
                return

            # ================================================================
            # STEP 2: PARALLEL PROCESSING
            # ================================================================
            self.log("")
            self.log("=" * 50)
            self.log(f"🚀 BẮT ĐẦU XỬ LÝ SONG SONG ({len(valid_tokens)} luồng)")
            self.log("=" * 50)

            # Create work queue
            work_queue = queue.Queue()
            for f in pending_files:
                work_queue.put(f)

            # Results tracking
            results_lock = threading.Lock()
            results = {"success": 0, "failed": 0, "processed": 0}

            def worker_func(worker_id: int, token_data: dict):
                """Worker function - processes voices from queue."""
                while not self._stop:
                    try:
                        file_info = work_queue.get_nowait()
                    except queue.Empty:
                        break

                    voice_path = file_info['voice_path']
                    output_folder = file_info['output_folder']

                    self.root.after(0, lambda wp=voice_path, wid=worker_id:
                        self.log(f"[W{wid}] 🎙️ {wp.name}"))

                    try:
                        output_folder.mkdir(parents=True, exist_ok=True)

                        engine = SmartEngine()

                        # Pass pre-fetched token to engine
                        if token_data:
                            engine._prefetched_token = token_data

                        def log_cb(msg, wid=worker_id):
                            level = "INFO"
                            if "[OK]" in msg or "OK!" in msg:
                                level = "OK"
                            elif "[ERROR]" in msg or "ERROR" in msg:
                                level = "ERROR"
                            elif "[WARN]" in msg:
                                level = "WARN"
                            self.root.after(0, lambda: self.log(f"[W{wid}] {msg}", level))

                        result = engine.run(
                            str(voice_path),
                            output_dir=str(output_folder),
                            callback=log_cb
                        )

                        with results_lock:
                            results["processed"] += 1
                            if result and result.get('success'):
                                results["success"] += 1
                            else:
                                results["failed"] += 1

                            # Update progress
                            progress = (results["processed"] / total) * 100
                            self.root.after(0, lambda p=progress: self.progress_var.set(p))
                            self.root.after(0, lambda: self.progress_label.config(
                                text=f"Xong: {results['processed']}/{total} | OK: {results['success']} | Fail: {results['failed']}"
                            ))

                    except Exception as e:
                        with results_lock:
                            results["processed"] += 1
                            results["failed"] += 1
                        self.root.after(0, lambda err=e, wid=worker_id:
                            self.log(f"[W{wid}] ❌ {err}", "ERROR"))

                    work_queue.task_done()

            # Start workers
            with ThreadPoolExecutor(max_workers=len(valid_tokens)) as executor:
                futures = []
                for i, token_data in enumerate(valid_tokens):
                    future = executor.submit(worker_func, i, token_data)
                    futures.append(future)

                # Wait for all workers to complete
                for future in as_completed(futures):
                    try:
                        future.result()
                    except Exception as e:
                        self.log(f"Worker error: {e}", "ERROR")

            # ================================================================
            # SUMMARY
            # ================================================================
            self.log("")
            self.log("=" * 50)
            self.log("🏁 HOÀN THÀNH BATCH")
            self.log(f"   Tổng: {total} | ✅ OK: {results['success']} | ❌ Fail: {results['failed']}")
            self.log(f"   Luồng: {len(valid_tokens)} workers")
            self.log("=" * 50)

            self.root.after(0, lambda: self.progress_var.set(100))
            self.root.after(0, lambda: self.progress_label.config(text="Hoàn thành!"))

            new_pending = self._count_pending_voices()
            self.root.after(0, lambda: self.input_info_label.config(
                text=f"📂 voice → done | {new_pending} file chờ xử lý"
            ))

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda err=e: self.log(f"Lỗi batch: {err}", "ERROR"))
        finally:
            self._running = False
            self.root.after(0, self._reset_ui)

    def _process_folder(self):
        """Process folder with multiple voice files - PARALLEL with headless Chrome."""
        try:
            from modules.smart_engine import SmartEngine
            from concurrent.futures import ThreadPoolExecutor, as_completed

            folder = Path(self.input_path.get())
            voices = list(folder.glob("*.mp3")) + list(folder.glob("*.wav"))

            if not voices:
                self.root.after(0, lambda: messagebox.showerror("Lỗi", "Không tìm thấy file voice nào!"))
                return

            self.log(f"📁 Tìm thấy {len(voices)} file voice")

            # Get settings
            num_parallel = self._get_parallel_workers()
            use_headless = self._get_headless_setting()

            # Get Chrome profiles from chrome_profiles/ directory (GUI creates profiles here)
            chrome_profiles = []
            profiles_dir = ROOT_DIR / "chrome_profiles"

            # Debug: Log exact path being scanned
            self.log(f"📂 Scan profiles: {profiles_dir.resolve()}")

            if profiles_dir.exists():
                # List all items in directory
                all_items = list(profiles_dir.iterdir())
                self.log(f"   Tìm thấy {len(all_items)} items")

                for item in all_items:
                    if item.is_dir() and not item.name.startswith('.'):
                        chrome_profiles.append(str(item))
                        self.log(f"   ✓ {item.name}")
                    else:
                        self.log(f"   ✗ {item.name} (không phải thư mục)")
            else:
                self.log(f"   ⚠️ Thư mục không tồn tại, tạo mới...")
                profiles_dir.mkdir(exist_ok=True)

            # Fallback: Load from accounts.json (additional profiles)
            try:
                accounts_path = CONFIG_DIR / "accounts.json"
                if accounts_path.exists():
                    with open(accounts_path, 'r', encoding='utf-8') as f:
                        accounts = json.load(f)
                    existing_paths = set(chrome_profiles)
                    for p in accounts.get('chrome_profiles', []):
                        if p and not p.startswith('THAY_BANG') and Path(p).exists():
                            if p not in existing_paths:
                                chrome_profiles.append(p)
                                self.log(f"   + {Path(p).name} (từ accounts.json)")
            except Exception as e:
                self.log(f"Load accounts.json error: {e}", "WARN")

            # Create default profile if none
            if not chrome_profiles:
                default_profile = profiles_dir / "main"
                default_profile.mkdir(exist_ok=True)
                chrome_profiles = [str(default_profile)]
                self.log(f"   → Tạo profile mặc định: main")

            num_profiles = len(chrome_profiles)
            num_parallel = min(num_parallel, num_profiles, len(voices))

            self.log(f"🌐 {num_profiles} profile(s) | Song song: {num_parallel} | Headless: {'ON' if use_headless else 'OFF'}")
            for i, p in enumerate(chrome_profiles[:num_parallel]):
                profile_name = Path(p).name
                self.log(f"   [{i+1}] {profile_name}")

            # Result tracking
            results_lock = threading.Lock()
            total_results = {"success": 0, "failed": 0}
            completed_count = [0]

            def process_voice(voice_path, profile_path, voice_idx, worker_id):
                """Process single voice with dedicated Chrome profile (headless)."""
                try:
                    profile_name = Path(profile_path).name
                    self.root.after(0, lambda: self.log(f"\n[{voice_idx+1}/{len(voices)}] 📄 {voice_path.name} → Worker {worker_id} ({profile_name})"))

                    # Create engine with specific profile
                    engine = SmartEngine(assigned_profile=profile_name)

                    # Force headless mode if enabled
                    engine.use_headless = use_headless

                    def log_cb(msg):
                        self.root.after(0, lambda m=msg, w=worker_id: self.log(f"  [W{w}] {m}"))

                    result = engine.run(str(voice_path), callback=log_cb)

                    with results_lock:
                        if 'error' not in result:
                            total_results["success"] += result.get('success', 0)
                            total_results["failed"] += result.get('failed', 0)
                        completed_count[0] += 1
                        progress = (completed_count[0] / len(voices)) * 100
                        self.root.after(0, lambda p=progress: self.update_progress(p, f"Xong {completed_count[0]}/{len(voices)}"))

                    return result
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    self.root.after(0, lambda err=e: self.log(f"Lỗi: {err}", "ERROR"))
                    return {"error": str(e)}

            # Process voices in parallel with dedicated profiles
            with ThreadPoolExecutor(max_workers=num_parallel) as executor:
                futures = {}
                for i, voice in enumerate(voices):
                    if self._stop:
                        break
                    # Assign worker and profile (round-robin)
                    worker_id = i % num_parallel
                    profile_path = chrome_profiles[worker_id % num_profiles]
                    future = executor.submit(process_voice, voice, profile_path, i, worker_id)
                    futures[future] = (voice, profile_path)

                # Wait for completion
                for future in as_completed(futures):
                    if self._stop:
                        break
                    voice, profile = futures[future]
                    try:
                        future.result()
                    except Exception as e:
                        self.root.after(0, lambda err=e, v=voice.name: self.log(f"❌ {v}: {err}", "ERROR"))

            # Summary
            self.root.after(0, lambda: self.update_progress(100, "Hoàn tất!"))
            self.root.after(0, lambda s=total_results["success"], f=total_results["failed"]:
                self.log(f"\n📊 TỔNG KẾT: {s} ✅ | {f} ❌", "OK"))

            if total_results["failed"] > 0:
                self.root.after(0, lambda s=total_results["success"], f=total_results["failed"]:
                    messagebox.showwarning("Chưa hoàn thành", f"✅ Thành công: {s}\n❌ Thất bại: {f}"))
            else:
                self.root.after(0, lambda s=total_results["success"]:
                    messagebox.showinfo("Hoàn tất!", f"✅ Đã tạo {s} ảnh!"))

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda err=e: self.log(f"Lỗi: {err}", "ERROR"))
        finally:
            self._running = False
            self.root.after(0, self._reset_ui)
    
    def _on_complete(self, results):
        """Handle completion."""
        # Set current project for preview
        path = self.input_path.get()
        if path:
            name = Path(path).stem
            self.current_project_dir = PROJECTS_DIR / name
        
        if 'error' in results:
            err = results['error']
            if err == 'missing_requirements':
                missing = results.get('missing', [])
                messagebox.showerror("Thiếu yêu cầu",
                    "Cần bổ sung:\n\n" + "\n".join(f"• {m}" for m in missing))
            else:
                messagebox.showerror("Lỗi", str(err))
        else:
            success = results.get('success', 0)
            failed = results.get('failed', 0)
            
            self.update_progress(100, "Hoàn tất!")
            
            # Auto refresh preview
            self.refresh_preview()
            
            if failed > 0:
                messagebox.showwarning("Chưa hoàn thành",
                    f"✅ Thành công: {success}\n❌ Thất bại: {failed}\n\nXem log để biết chi tiết.")
            else:
                messagebox.showinfo("Hoàn tất!", f"✅ Đã tạo {success} ảnh!")
    
    def _reset_ui(self):
        """Reset UI after processing."""
        self._running = False
        self._stop_auto_refresh()
        self.start_btn.config(state=tk.NORMAL, bg=self.COLORS['primary'])
        self.stop_btn.config(state=tk.DISABLED)
        # Final refresh
        self.refresh_preview()
    
    # ========== PREVIEW ==========
    
    def on_char_selected(self, event=None):
        """Handle character selection."""
        sel = self.char_combo.get()
        if not sel or not self.current_project_dir:
            return
        
        # Load character image
        img_path = self.current_project_dir / "nv" / f"{sel}.png"
        self.load_image_to_label(img_path, self.char_image_label, (200, 200))
        
        # Load prompt from Excel
        self.char_prompt_text.delete(1.0, tk.END)
        prompt = self.get_prompt_for_id(sel)
        if prompt:
            self.char_prompt_text.insert(tk.END, prompt)
    
    def on_scene_selected(self, event=None):
        """Handle scene selection."""
        sel = self.scene_combo.get()
        if not sel or not self.current_project_dir:
            return
        
        # Load scene image
        img_path = self.current_project_dir / "img" / f"{sel}.png"
        self.load_image_to_label(img_path, self.result_image_label, (300, 200))
        
        # Load prompt
        self.result_prompt_text.delete(1.0, tk.END)
        prompt = self.get_prompt_for_id(sel)
        if prompt:
            self.result_prompt_text.insert(tk.END, prompt)
    
    def load_image_to_label(self, img_path: Path, label: ttk.Label, size: tuple):
        """Load image and display on label."""
        if not HAS_PIL:
            label.config(text="Cần cài PIL:\npip install Pillow")
            return
        
        if not img_path.exists():
            label.config(text=f"Chưa có ảnh\n{img_path.name}")
            return
        
        try:
            img = Image.open(img_path)
            img.thumbnail(size, Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            
            # Keep reference
            label._photo = photo
            label.config(image=photo, text="")
        except Exception as e:
            label.config(text=f"Lỗi: {e}")
    
    def get_prompt_for_id(self, pid: str) -> str:
        """Get prompt for an ID from Excel."""
        if not self.current_project_dir:
            return ""
        
        # Find Excel file
        prompts_dir = self.current_project_dir / "prompts"
        excel_files = list(prompts_dir.glob("*_prompts.xlsx"))
        
        if not excel_files:
            return ""
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_files[0], read_only=True)
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [c.value for c in ws[1]]
                
                # Find columns
                id_col = prompt_col = None
                for i, h in enumerate(headers or []):
                    if h is None:
                        continue
                    h_lower = str(h).lower()
                    if 'id' in h_lower and id_col is None:
                        id_col = i
                    if 'english' in h_lower and 'prompt' in h_lower:
                        prompt_col = i
                    elif h_lower == 'img_prompt' and prompt_col is None:
                        prompt_col = i
                    elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower and 'viet' not in h_lower:
                        prompt_col = i
                
                if id_col is None or prompt_col is None:
                    continue
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > max(id_col, prompt_col):
                        if str(row[id_col]).strip() == pid:
                            return str(row[prompt_col] or "")
            
            wb.close()
        except:
            pass
        
        return ""
    
    def refresh_preview(self):
        """Refresh preview - populate unified tree."""
        # Find project dir from input
        path = self.input_path.get()
        if path:
            name = Path(path).stem
            self.current_project_dir = PROJECTS_DIR / name

        if not self.current_project_dir or not self.current_project_dir.exists():
            self.thumb_progress.config(text="Chưa có project", foreground='gray')
            return

        # Clear tree
        for item in self.main_tree.get_children():
            self.main_tree.delete(item)

        # Get all data from Excel
        all_items = []  # [(id, type, prompt, status), ...]

        prompts_dir = self.current_project_dir / "prompts"
        excel_files = list(prompts_dir.glob("*_prompts.xlsx")) if prompts_dir.exists() else []
        nv_dir = self.current_project_dir / "nv"
        img_dir = self.current_project_dir / "img"

        if excel_files:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_files[0], read_only=True)

                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    headers = [c.value for c in ws[1]]

                    # Find columns
                    id_col = prompt_col = None
                    for i, h in enumerate(headers or []):
                        if h is None:
                            continue
                        h_lower = str(h).lower()
                        if 'id' in h_lower and id_col is None:
                            id_col = i
                        if 'english' in h_lower and 'prompt' in h_lower:
                            prompt_col = i
                        elif h_lower == 'img_prompt' and prompt_col is None:
                            prompt_col = i
                        elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower and 'viet' not in h_lower:
                            prompt_col = i

                    if id_col is None:
                        continue

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) <= id_col:
                            continue
                        pid = str(row[id_col] or "").strip()
                        if not pid:
                            continue

                        prompt = str(row[prompt_col] or "")[:80] + "..." if prompt_col and len(row) > prompt_col else ""
                        # Characters (nv*) and Locations (loc*) are reference images -> save in nv/
                        is_reference = pid.startswith('nv') or pid.startswith('loc')
                        if pid.startswith('nv'):
                            item_type = "NV"
                        elif pid.startswith('loc'):
                            item_type = "LOC"
                        else:
                            item_type = "Scene"

                        # Check status - reference images in nv/, scene images in img/
                        if is_reference:
                            img_path = nv_dir / f"{pid}.png"
                        else:
                            img_path = img_dir / f"{pid}.png"
                        status = "✅" if img_path.exists() else "⏳"

                        all_items.append((pid, item_type, prompt, status))

                wb.close()
            except Exception as e:
                self.log(f"Error reading Excel: {e}", "ERROR")

        # Sort: reference images (nv*, loc*) first, then scenes by ID
        def sort_key(item):
            pid = item[0]
            is_reference = pid.startswith('nv') or pid.startswith('loc')
            try:
                num = int(''.join(filter(str.isdigit, pid)))
            except:
                num = 999
            # Sort order: nv (0), loc (1), scenes (2)
            if pid.startswith('nv'):
                order = 0
            elif pid.startswith('loc'):
                order = 1
            else:
                order = 2
            return (order, num)

        all_items.sort(key=sort_key)

        # Deduplicate by ID (keep first occurrence)
        seen_ids = set()
        unique_items = []
        for item in all_items:
            if item[0] not in seen_ids:
                seen_ids.add(item[0])
                unique_items.append(item)
        all_items = unique_items

        # Populate tree
        for pid, item_type, prompt, status in all_items:
            try:
                # Check if item exists before inserting to avoid TclError
                if not self.main_tree.exists(pid):
                    self.main_tree.insert('', tk.END, iid=pid, values=(pid, item_type, prompt, status))
            except tk.TclError:
                pass  # Skip if item already exists (race condition with auto-refresh)
            except Exception:
                pass  # Skip other errors silently

        # Update progress
        total = len(all_items)
        done = sum(1 for item in all_items if item[3] == "✅")
        color = '#27ae60' if done == total else '#f39c12'
        self.thumb_progress.config(text=f"Tiến độ: {done}/{total} ảnh hoàn thành", foreground=color)

        # Select first item
        if all_items:
            first_id = all_items[0][0]
            self.main_tree.selection_set(first_id)
            self.main_tree.focus(first_id)
            self._on_item_selected(first_id)

        self.log(f"Loaded {len(all_items)} items", "OK")
    
    def update_thumbnails(self, scene_ids: List[str]):
        """Update scene thumbnails with progress status."""
        self.thumb_canvas.delete("all")

        if not self.current_project_dir:
            self.thumb_progress.config(text="")
            return

        img_dir = self.current_project_dir / "img"
        nv_dir = self.current_project_dir / "nv"

        # Count progress
        total_scenes = len(scene_ids)
        done_scenes = sum(1 for sid in scene_ids if (img_dir / f"{sid}.png").exists())

        # Count characters too
        char_pngs = list(nv_dir.glob("*.png")) if nv_dir.exists() else []

        self.thumb_progress.config(
            text=f"Scenes: {done_scenes}/{total_scenes} ✅  |  Nhân vật: {len(char_pngs)}",
            foreground='#27ae60' if done_scenes == total_scenes else '#f39c12'
        )

        if not HAS_PIL:
            return

        x = 5
        self._thumb_photos = []  # Keep references
        self._thumb_scene_ids = scene_ids  # Store for click handling

        for sid in scene_ids[:30]:  # Max 30 thumbnails
            img_path = img_dir / f"{sid}.png"

            if img_path.exists():
                try:
                    img = Image.open(img_path)
                    img.thumbnail((80, 80), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    self._thumb_photos.append(photo)

                    self.thumb_canvas.create_image(x, 5, anchor=tk.NW, image=photo)
                    # Green border for done
                    self.thumb_canvas.create_rectangle(x-1, 4, x+81, 86, outline='#27ae60', width=2)
                except:
                    # Placeholder for failed load
                    self.thumb_canvas.create_rectangle(x, 5, x+80, 85, fill='#bdc3c7', outline='#7f8c8d')
                    self.thumb_canvas.create_text(x+40, 45, text="?", font=('Segoe UI', 16))
            else:
                # Placeholder for pending
                self.thumb_canvas.create_rectangle(x, 5, x+80, 85, fill='#ecf0f1', outline='#f39c12', width=2)
                self.thumb_canvas.create_text(x+40, 40, text="⏳", font=('Segoe UI', 16))

            # Scene ID label
            self.thumb_canvas.create_text(x + 40, 92, text=sid, font=('Segoe UI', 7))
            x += 90

        self.thumb_canvas.configure(scrollregion=(0, 0, x, 105))
    
    def update_prompts_tab(self):
        """Update prompts treeviews."""
        # Clear existing
        for item in self.char_tree.get_children():
            self.char_tree.delete(item)
        for item in self.scene_tree.get_children():
            self.scene_tree.delete(item)
        
        if not self.current_project_dir:
            return
        
        # Find Excel
        prompts_dir = self.current_project_dir / "prompts"
        excel_files = list(prompts_dir.glob("*_prompts.xlsx"))
        
        if not excel_files:
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_files[0], read_only=True)
            
            nv_dir = self.current_project_dir / "nv"
            img_dir = self.current_project_dir / "img"
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [c.value for c in ws[1]]
                
                # Find columns
                id_col = prompt_col = time_col = None
                for i, h in enumerate(headers or []):
                    if h is None:
                        continue
                    h_lower = str(h).lower()
                    if 'id' in h_lower and id_col is None:
                        id_col = i
                    if 'english' in h_lower and 'prompt' in h_lower:
                        prompt_col = i
                    elif h_lower == 'img_prompt' and prompt_col is None:
                        prompt_col = i
                    elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower and 'viet' not in h_lower:
                        prompt_col = i
                    if 'time' in h_lower or 'start' in h_lower:
                        time_col = i
                
                if id_col is None or prompt_col is None:
                    continue
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= max(id_col, prompt_col):
                        continue
                    
                    pid = str(row[id_col] or "").strip()
                    prompt = str(row[prompt_col] or "")[:60] + "..."
                    time_str = str(row[time_col] or "") if time_col else ""
                    
                    if not pid:
                        continue

                    # Check status - reference images (nv*, loc*) in nv/, scenes in img/
                    is_reference = pid.startswith('nv') or pid.startswith('loc')
                    try:
                        if is_reference:
                            img_path = nv_dir / f"{pid}.png"
                            status = "✅" if img_path.exists() else "⏳"
                            self.char_tree.insert('', tk.END, values=(pid, prompt, status))
                        else:
                            img_path = img_dir / f"{pid}.png"
                            status = "✅" if img_path.exists() else "⏳"
                            self.scene_tree.insert('', tk.END, values=(pid, time_str, prompt, status))
                    except tk.TclError:
                        pass  # Skip duplicates
            
            wb.close()
        except Exception as e:
            self.log(f"Error loading prompts: {e}", "ERROR")
    
    def get_token_manual(self):
        """Get token manually."""
        if not self.profiles:
            messagebox.showerror("Lỗi", "Chưa có Chrome profile!\n\nThêm vào config/accounts.json")
            return

        self.log("🔑 Đang lấy token thủ công...")

        def worker():
            try:
                from modules.auto_token import ChromeAutoToken

                extractor = ChromeAutoToken(
                    chrome_path=self.chrome_path,
                    profile_path=self.profiles[0]
                )

                def log_cb(msg):
                    self.root.after(0, lambda: self.log(msg))

                token, proj_id, error = extractor.extract_token(callback=log_cb)

                if token:
                    self.root.after(0, lambda: self.log(f"✅ Token: {token[:40]}...", "OK"))
                    self.root.after(0, lambda: messagebox.showinfo("OK", "Đã lấy được token!"))
                else:
                    self.root.after(0, lambda: self.log(f"❌ {error}", "ERROR"))

            except Exception as e:
                self.root.after(0, lambda err=e: self.log(f"Lỗi: {err}", "ERROR"))

        threading.Thread(target=worker, daemon=True).start()

    # ========== PREVIEW ACTIONS ==========

    def save_char_prompt(self):
        """Save edited character prompt to Excel."""
        char_id = self.char_combo.get()
        if not char_id or not self.current_project_dir:
            return

        new_prompt = self.char_prompt_text.get(1.0, tk.END).strip()
        if self._update_prompt_in_excel(char_id, new_prompt):
            self.char_status_label.config(text="✅ Đã lưu", foreground='green')
            self.log(f"Đã lưu prompt cho {char_id}", "OK")
        else:
            self.char_status_label.config(text="❌ Lỗi", foreground='red')

    def save_scene_prompt(self):
        """Save edited scene prompt to Excel."""
        scene_id = self.scene_combo.get()
        if not scene_id or not self.current_project_dir:
            return

        new_prompt = self.result_prompt_text.get(1.0, tk.END).strip()
        if self._update_prompt_in_excel(scene_id, new_prompt):
            self.scene_status_label.config(text="✅ Đã lưu", foreground='green')
            self.log(f"Đã lưu prompt cho {scene_id}", "OK")
        else:
            self.scene_status_label.config(text="❌ Lỗi", foreground='red')

    def _update_prompt_in_excel(self, item_id: str, new_prompt: str) -> bool:
        """Update prompt in Excel file."""
        if not self.current_project_dir:
            return False

        prompts_dir = self.current_project_dir / "prompts"
        excel_files = list(prompts_dir.glob("*_prompts.xlsx"))

        if not excel_files:
            self.log("Không tìm thấy file Excel", "ERROR")
            return False

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_files[0])

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = [c.value for c in ws[1]]

                # Find columns
                id_col = prompt_col = None
                for i, h in enumerate(headers or []):
                    if h is None:
                        continue
                    h_lower = str(h).lower()
                    if 'id' in h_lower and id_col is None:
                        id_col = i + 1  # openpyxl is 1-indexed
                    if 'english' in h_lower and 'prompt' in h_lower:
                        prompt_col = i + 1
                    elif h_lower == 'img_prompt' and prompt_col is None:
                        prompt_col = i + 1
                    elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower and 'viet' not in h_lower:
                        prompt_col = i + 1

                if id_col is None or prompt_col is None:
                    continue

                for row_num in range(2, ws.max_row + 1):
                    cell_id = ws.cell(row=row_num, column=id_col).value
                    if str(cell_id).strip() == item_id:
                        ws.cell(row=row_num, column=prompt_col).value = new_prompt
                        wb.save(excel_files[0])
                        wb.close()
                        return True

            wb.close()
        except Exception as e:
            self.log(f"Error updating Excel: {e}", "ERROR")
            return False

        return False

    def regenerate_char_image(self):
        """Regenerate character image with current prompt."""
        char_id = self.char_combo.get()
        if not char_id or not self.current_project_dir:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn nhân vật cần tạo lại")
            return

        prompt = self.char_prompt_text.get(1.0, tk.END).strip()
        if not prompt:
            messagebox.showwarning("Thiếu prompt", "Prompt không được để trống")
            return

        self._regenerate_single_image(char_id, prompt, is_char=True)

    def regenerate_scene_image(self):
        """Regenerate scene image with current prompt."""
        scene_id = self.scene_combo.get()
        if not scene_id or not self.current_project_dir:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn scene cần tạo lại")
            return

        prompt = self.result_prompt_text.get(1.0, tk.END).strip()
        if not prompt:
            messagebox.showwarning("Thiếu prompt", "Prompt không được để trống")
            return

        self._regenerate_single_image(scene_id, prompt, is_char=False)

    def _regenerate_single_image(self, item_id: str, prompt: str, is_char: bool = False):
        """Regenerate a single image."""
        if not self.profiles:
            messagebox.showerror("Lỗi", "Chưa có Chrome profile!")
            return

        self.log(f"🔄 Đang tạo lại ảnh: {item_id}...")

        def worker():
            try:
                from modules.smart_engine import SmartEngine

                engine = SmartEngine()

                # Get token
                token, proj_id = engine.get_token_for_profile(self.profiles[0])
                if not token:
                    self.root.after(0, lambda: self.log(f"❌ Không lấy được token", "ERROR"))
                    return

                # Generate image
                from modules.flow_image_generator import FlowImageGenerator
                generator = FlowImageGenerator()

                # Determine output path
                if is_char:
                    output_path = self.current_project_dir / "nv" / f"{item_id}.png"
                else:
                    output_path = self.current_project_dir / "img" / f"{item_id}.png"

                # Backup old image
                if output_path.exists():
                    backup = output_path.with_suffix('.bak.png')
                    shutil.copy(output_path, backup)

                success = generator.generate_and_save(
                    prompt=prompt,
                    output_path=str(output_path),
                    token=token,
                    project_id=proj_id
                )

                if success:
                    self.root.after(0, lambda: self.log(f"✅ Đã tạo lại: {item_id}", "OK"))
                    self.root.after(0, self.refresh_preview)
                else:
                    self.root.after(0, lambda: self.log(f"❌ Lỗi tạo ảnh: {item_id}", "ERROR"))

            except Exception as e:
                self.root.after(0, lambda err=e: self.log(f"Lỗi: {err}", "ERROR"))

        threading.Thread(target=worker, daemon=True).start()

    def on_thumb_click(self, event):
        """Handle click on thumbnail canvas."""
        # Calculate which thumbnail was clicked
        x = self.thumb_canvas.canvasx(event.x)
        idx = int(x // 90)

        # Use stored scene IDs
        if hasattr(self, '_thumb_scene_ids') and idx < len(self._thumb_scene_ids):
            self.scene_combo.set(self._thumb_scene_ids[idx])
            self.on_scene_selected()

    # ========== PROMPTS TAB ACTIONS ==========

    def on_char_tree_double_click(self, event):
        """Handle double-click on character tree to edit prompt."""
        item = self.char_tree.selection()
        if not item:
            return

        values = self.char_tree.item(item[0], 'values')
        char_id = values[0]

        # Get full prompt from Excel
        full_prompt = self.get_prompt_for_id(char_id)

        # Open edit dialog
        new_prompt = self._show_prompt_edit_dialog(f"Sửa prompt: {char_id}", full_prompt)

        if new_prompt is not None and new_prompt != full_prompt:
            if self._update_prompt_in_excel(char_id, new_prompt):
                self.log(f"Đã cập nhật prompt: {char_id}", "OK")
                self.update_prompts_tab()

    def on_scene_tree_double_click(self, event):
        """Handle double-click on scene tree to edit prompt."""
        item = self.scene_tree.selection()
        if not item:
            return

        values = self.scene_tree.item(item[0], 'values')
        scene_id = values[0]

        # Get full prompt from Excel
        full_prompt = self.get_prompt_for_id(scene_id)

        # Open edit dialog
        new_prompt = self._show_prompt_edit_dialog(f"Sửa prompt: {scene_id}", full_prompt)

        if new_prompt is not None and new_prompt != full_prompt:
            if self._update_prompt_in_excel(scene_id, new_prompt):
                self.log(f"Đã cập nhật prompt: {scene_id}", "OK")
                self.update_prompts_tab()

    def _show_prompt_edit_dialog(self, title: str, current_prompt: str) -> Optional[str]:
        """Show dialog to edit a prompt."""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("600x300")
        dialog.transient(self.root)
        dialog.grab_set()

        result = [None]

        ttk.Label(dialog, text="Chỉnh sửa prompt:", font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W, padx=10, pady=(10, 5))

        text = tk.Text(dialog, wrap=tk.WORD, font=('Segoe UI', 10), height=10)
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        text.insert(tk.END, current_prompt)

        def on_save():
            result[0] = text.get(1.0, tk.END).strip()
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        btn_row = ttk.Frame(dialog)
        btn_row.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(btn_row, text="💾 Lưu", command=on_save).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="Hủy", command=on_cancel).pack(side=tk.LEFT)

        dialog.wait_window()
        return result[0]

    def regenerate_selected_char(self):
        """Regenerate image for selected character in tree."""
        item = self.char_tree.selection()
        if not item:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn nhân vật trong danh sách")
            return

        values = self.char_tree.item(item[0], 'values')
        char_id = values[0]
        prompt = self.get_prompt_for_id(char_id)

        if prompt:
            self._regenerate_single_image(char_id, prompt, is_char=True)

    def regenerate_selected_scene(self):
        """Regenerate image for selected scene in tree."""
        item = self.scene_tree.selection()
        if not item:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn scene trong danh sách")
            return

        values = self.scene_tree.item(item[0], 'values')
        scene_id = values[0]
        prompt = self.get_prompt_for_id(scene_id)

        if prompt:
            self._regenerate_single_image(scene_id, prompt, is_char=False)

    # ========== UNIFIED TREE HANDLERS ==========

    def on_tree_select(self, event=None):
        """Handle tree item selection."""
        sel = self.main_tree.selection()
        if not sel:
            return

        item_id = sel[0]
        self._on_item_selected(item_id)

    def on_tree_double_click(self, event=None):
        """Handle double-click on tree item to edit prompt."""
        sel = self.main_tree.selection()
        if not sel:
            return

        item_id = sel[0]
        full_prompt = self.get_prompt_for_id(item_id)

        new_prompt = self._show_prompt_edit_dialog(f"Sửa prompt: {item_id}", full_prompt)

        if new_prompt is not None and new_prompt != full_prompt:
            if self._update_prompt_in_excel(item_id, new_prompt):
                self.log(f"Đã cập nhật prompt: {item_id}", "OK")
                self.refresh_preview()

    def _on_item_selected(self, item_id: str):
        """Update detail panel when an item is selected."""
        self._current_item_id = item_id
        # Identify item type: char (nv*), loc (loc*), or scene
        if item_id.startswith('nv'):
            self._current_item_type = "char"
        elif item_id.startswith('loc'):
            self._current_item_type = "loc"
        else:
            self._current_item_type = "scene"

        if not self.current_project_dir:
            return

        # Load prompt
        full_prompt = self.get_prompt_for_id(item_id)
        self.detail_prompt_text.delete(1.0, tk.END)
        self.detail_prompt_text.insert(tk.END, full_prompt)

        # Load images based on item type
        if self._current_item_type in ("char", "loc"):
            # Character/Location: reference images stored in nv/
            label_text = "N/A (nhân vật)" if self._current_item_type == "char" else "N/A (bối cảnh)"
            self.ref_image_label.config(image='', text=label_text)
            img_path = self.current_project_dir / "nv" / f"{item_id}.png"
            self.load_image_to_label(img_path, self.result_image_label, (200, 200))
        else:
            # Scene: reference (character/location) + result image
            # Try to find character reference in prompt
            ref_char = self._find_ref_char_in_prompt(full_prompt)
            if ref_char:
                ref_path = self.current_project_dir / "nv" / f"{ref_char}.png"
                self.load_image_to_label(ref_path, self.ref_image_label, (150, 150))
            else:
                self.ref_image_label.config(image='', text="Không có tham chiếu")

            img_path = self.current_project_dir / "img" / f"{item_id}.png"
            self.load_image_to_label(img_path, self.result_image_label, (200, 200))

        self.detail_status.config(text=f"Đang xem: {item_id}")

    def _find_ref_char_in_prompt(self, prompt: str) -> Optional[str]:
        """Find character reference (nv1, nv2, etc.) in prompt."""
        import re
        match = re.search(r'\bnv(\d+)\b', prompt.lower())
        if match:
            return f"nv{match.group(1)}"
        return None

    def save_current_prompt(self):
        """Save edited prompt from detail panel to Excel."""
        if not self._current_item_id:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn một item trước")
            return

        new_prompt = self.detail_prompt_text.get(1.0, tk.END).strip()
        if self._update_prompt_in_excel(self._current_item_id, new_prompt):
            self.detail_status.config(text="✅ Đã lưu", foreground='green')
            self.log(f"Đã lưu prompt cho {self._current_item_id}", "OK")
            # Update tree display
            self.refresh_preview()
        else:
            self.detail_status.config(text="❌ Lỗi", foreground='red')

    def regenerate_current_image(self):
        """Regenerate image for currently selected item."""
        if not self._current_item_id:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn một item trước")
            return

        prompt = self.detail_prompt_text.get(1.0, tk.END).strip()
        if not prompt:
            messagebox.showwarning("Thiếu prompt", "Prompt không được để trống")
            return

        # Characters and Locations are reference images -> save in nv/
        is_reference = self._current_item_type in ("char", "loc")
        self._regenerate_single_image(self._current_item_id, prompt, is_char=is_reference)

    def regenerate_all_pending(self):
        """Regenerate all pending (not done) images."""
        if not self.current_project_dir:
            return

        # Collect pending items from unified tree
        pending = []

        img_dir = self.current_project_dir / "img"
        nv_dir = self.current_project_dir / "nv"

        for item in self.main_tree.get_children():
            values = self.main_tree.item(item, 'values')
            # values = (id, type, prompt, status)
            if len(values) >= 4 and values[3] == "⏳":  # Status is pending
                item_id = values[0]
                # Characters (NV) and Locations (LOC) are reference images -> save in nv/
                item_type = str(values[1]).upper()
                is_reference = item_type in ("NV", "LOC")
                prompt = self.get_prompt_for_id(item_id)
                if prompt:
                    pending.append((item_id, prompt, is_reference))

        if not pending:
            messagebox.showinfo("Thông báo", "Tất cả ảnh đã hoàn thành!")
            return

        if not messagebox.askyesno("Xác nhận", f"Tạo lại {len(pending)} ảnh chưa xong?"):
            return

        self.log(f"🔄 Bắt đầu tạo lại {len(pending)} ảnh...")

        def worker():
            try:
                from modules.smart_engine import SmartEngine
                from modules.flow_image_generator import FlowImageGenerator

                engine = SmartEngine()
                generator = FlowImageGenerator()

                for i, (item_id, prompt, is_reference) in enumerate(pending):
                    self.root.after(0, lambda id=item_id, n=i+1, t=len(pending):
                        self.log(f"[{n}/{t}] Đang tạo: {id}..."))

                    # Get token
                    profile = self.profiles[i % len(self.profiles)]
                    token, proj_id = engine.get_token_for_profile(profile)

                    if not token:
                        self.root.after(0, lambda id=item_id:
                            self.log(f"❌ Không có token cho {id}", "ERROR"))
                        continue

                    # Output path - reference images (nv*, loc*) in nv/, scenes in img/
                    if is_reference:
                        output_path = nv_dir / f"{item_id}.png"
                    else:
                        output_path = img_dir / f"{item_id}.png"

                    success = generator.generate_and_save(
                        prompt=prompt,
                        output_path=str(output_path),
                        token=token,
                        project_id=proj_id
                    )

                    if success:
                        self.root.after(0, lambda id=item_id:
                            self.log(f"✅ Xong: {id}", "OK"))
                    else:
                        self.root.after(0, lambda id=item_id:
                            self.log(f"❌ Lỗi: {id}", "ERROR"))

                self.root.after(0, lambda: self.log("🎉 Hoàn tất tạo lại ảnh!", "OK"))
                self.root.after(0, self.refresh_preview)

            except Exception as e:
                self.root.after(0, lambda err=e: self.log(f"Lỗi: {err}", "ERROR"))

        threading.Thread(target=worker, daemon=True).start()

    # ========== PROMPTS TEMPLATE EDITOR ==========

    def open_prompts_editor(self):
        """Open prompts template editor dialog."""
        prompts_file = CONFIG_DIR / "prompts.yaml"

        if not prompts_file.exists():
            messagebox.showwarning("Không tìm thấy", "File prompts.yaml không tồn tại")
            return

        win = tk.Toplevel(self.root)
        win.title("📝 Sửa Prompts Template")
        win.geometry("900x700")
        win.transient(self.root)

        # Instructions
        ttk.Label(win, text="Chỉnh sửa các prompt template cho AI. Lưu ý: Thay đổi sẽ ảnh hưởng đến project mới.",
                  foreground='gray').pack(anchor=tk.W, padx=10, pady=(10, 5))

        # Text editor with scrollbar
        editor_frame = ttk.Frame(win)
        editor_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        text = tk.Text(editor_frame, wrap=tk.NONE, font=('Consolas', 10), bg='#1e1e1e', fg='#d4d4d4',
                       insertbackground='white')
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        y_scroll = ttk.Scrollbar(editor_frame, orient=tk.VERTICAL, command=text.yview)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        x_scroll = ttk.Scrollbar(win, orient=tk.HORIZONTAL, command=text.xview)
        x_scroll.pack(fill=tk.X, padx=10)

        text.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        # Load content
        try:
            with open(prompts_file, 'r', encoding='utf-8') as f:
                content = f.read()
            text.insert(tk.END, content)
        except Exception as e:
            text.insert(tk.END, f"# Error loading file: {e}")

        def save_prompts():
            try:
                new_content = text.get(1.0, tk.END)
                with open(prompts_file, 'w', encoding='utf-8') as f:
                    f.write(new_content)

                # Reload prompts
                try:
                    from modules.prompts_loader import reload_prompts
                    reload_prompts()
                except:
                    pass

                messagebox.showinfo("OK", "Đã lưu prompts.yaml")
                self.log("Đã cập nhật prompts template", "OK")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu: {e}")

        def open_in_editor():
            if sys.platform == "win32":
                os.startfile(str(prompts_file))
            else:
                import subprocess
                subprocess.Popen(["xdg-open", str(prompts_file)])

        # Buttons
        btn_row = ttk.Frame(win)
        btn_row.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(btn_row, text="💾 Lưu", command=save_prompts).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="📂 Mở bằng Editor", command=open_in_editor).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="Đóng", command=win.destroy).pack(side=tk.RIGHT)

    # ========== RUN ==========
    
    def run(self):
        """Start application."""
        self.root.mainloop()


# ============================================================================
# MAIN
# ============================================================================

def setup_file_logging():
    """
    Setup logging to file for debugging.
    All output goes to logs/app.log
    """
    log_dir = ROOT_DIR / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / "app.log"

    # Rotate log if too big (> 5MB)
    if log_file.exists() and log_file.stat().st_size > 5 * 1024 * 1024:
        old_log = log_dir / "app.old.log"
        if old_log.exists():
            old_log.unlink()
        log_file.rename(old_log)

    # Open log file
    try:
        log_handle = open(log_file, 'a', encoding='utf-8')

        # Write startup marker
        log_handle.write(f"\n{'='*60}\n")
        log_handle.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] APP STARTED\n")
        log_handle.write(f"{'='*60}\n")
        log_handle.flush()

        return log_handle
    except:
        return None


def fix_stdio(log_handle=None):
    """
    Fix stdout/stderr when running without console (pythonw.exe).
    Redirect to log file if available, otherwise devnull.
    """
    if sys.stdout is None:
        sys.stdout = log_handle if log_handle else open(os.devnull, 'w')
    if sys.stderr is None:
        sys.stderr = log_handle if log_handle else open(os.devnull, 'w')


def main():
    """Entry point."""
    # Setup file logging first
    log_handle = setup_file_logging()

    # Fix stdio (for pythonw.exe)
    fix_stdio(log_handle)

    # Print startup info (goes to console AND log file)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Uni-x Voice to Video...")
    print(f"Python: {sys.executable}")
    print(f"Working dir: {os.getcwd()}")

    try:
        app = UnixVoiceToVideo()
        app.run()
    except Exception as e:
        print(f"[ERROR] App crashed: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
