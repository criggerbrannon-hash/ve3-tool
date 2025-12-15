#!/usr/bin/env python3
"""
VTV Tool - Generate Images với Reference Images
================================================
Workflow đúng:
1. Upload ảnh reference (nv, loc) -> lấy media_name
2. Generate scene với media_name (KHÔNG phải base64)

Usage:
    python vtv_with_references.py <project_path> <bearer_token>

Example:
    python vtv_with_references.py "PROJECTS/1" "ya29.xxx..."
"""

import os
import sys
import json
import time
import base64
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Add parent to path
sys.path.insert(0, str(Path(__file__).parent))

from modules.google_flow_api import (
    GoogleFlowAPI,
    AspectRatio,
    ImageInput,
    ImageInputType,
    GeneratedImage
)

# Video composer (optional)
try:
    from modules.video_composer import VideoComposer, VideoConfig, compose_final_video
    VIDEO_COMPOSER_AVAILABLE = True
except ImportError:
    VIDEO_COMPOSER_AVAILABLE = False


class VTVGenerator:
    """
    VTV Generator với Reference Images support.

    Workflow:
    1. Upload ảnh từ nv/ folder -> lấy media_name
    2. Generate scene images với media_name làm reference
    """

    def __init__(
        self,
        project_path: str,
        bearer_token: str,
        parallel_workers: int = 2,
        delay: float = 2.0,
        verbose: bool = True
    ):
        self.project_path = Path(project_path)
        self.bearer_token = bearer_token
        self.parallel = parallel_workers
        self.delay = delay
        self.verbose = verbose

        # Paths
        self.nv_path = self.project_path / "nv"
        self.img_path = self.project_path / "img"
        self.prompts_path = self.project_path / "prompts"

        # Ensure directories exist
        self.nv_path.mkdir(parents=True, exist_ok=True)
        self.img_path.mkdir(parents=True, exist_ok=True)

        # API client
        self.api = GoogleFlowAPI(
            bearer_token=bearer_token,
            verbose=verbose
        )

        # Reference cache: char_id -> media_name
        # QUAN TRỌNG: Lưu media_name, KHÔNG phải base64!
        self.reference_cache: Dict[str, str] = {}

        # Lock for thread safety
        self._lock = threading.Lock()
        self._stop = False

    def log(self, msg: str, level: str = "INFO"):
        """Log message."""
        if self.verbose:
            ts = datetime.now().strftime("%H:%M:%S")
            prefix = {"OK": "✅", "ERROR": "❌", "WARN": "⚠️", "INFO": "ℹ️"}.get(level, "•")
            print(f"[{ts}] [{level}] {prefix} {msg}")

    # =========================================================================
    # STEP 1: Upload references để lấy media_name
    # =========================================================================

    def upload_reference(self, image_path: Path) -> Optional[str]:
        """
        Upload 1 ảnh reference và trả về media_name.

        QUAN TRỌNG: Trả về media_name (string từ API response),
        KHÔNG phải base64 data!

        Returns:
            media_name string hoặc None nếu fail
        """
        if not image_path.exists():
            self.log(f"Reference not found: {image_path}", "WARN")
            return None

        self.log(f"Uploading reference: {image_path.name}...")

        success, img_input, error = self.api.upload_image(image_path)

        if success and img_input:
            media_name = img_input.name  # Đây là media_name từ API, KHÔNG phải base64
            self.log(f"  -> Got media_name: {media_name[:50]}...", "OK")
            return media_name
        else:
            self.log(f"  -> Upload failed: {error}", "ERROR")
            return None

    def upload_all_references(self) -> int:
        """
        Upload tất cả ảnh trong nv/ folder.

        Returns:
            Số lượng references đã upload thành công
        """
        self.log("=" * 50)
        self.log("UPLOADING REFERENCE IMAGES")
        self.log("=" * 50)

        if not self.nv_path.exists():
            self.log(f"NV folder not found: {self.nv_path}", "WARN")
            return 0

        uploaded = 0

        for img_path in self.nv_path.glob("*.png"):
            ref_id = img_path.stem  # nv1.png -> nv1, loc1.png -> loc1

            # Skip if already in cache
            if ref_id in self.reference_cache:
                self.log(f"  {ref_id}: Already cached, skip")
                uploaded += 1
                continue

            media_name = self.upload_reference(img_path)

            if media_name:
                self.reference_cache[ref_id] = media_name
                uploaded += 1

        self.log(f"Total: {uploaded} references uploaded", "OK")
        return uploaded

    # =========================================================================
    # STEP 2: Generate với references
    # =========================================================================

    def generate_with_references(
        self,
        prompt: str,
        reference_ids: List[str],  # ["nv1", "nv2", "loc1"]
        output_path: Path
    ) -> bool:
        """
        Generate 1 ảnh với references.

        Args:
            prompt: Text prompt
            reference_ids: List of reference IDs (phải đã upload trước)
            output_path: Đường dẫn lưu ảnh output

        Returns:
            True nếu thành công
        """
        # Build image_inputs từ cache
        image_inputs = []

        for ref_id in reference_ids:
            if ref_id in self.reference_cache:
                media_name = self.reference_cache[ref_id]

                # QUAN TRỌNG: Gửi media_name, KHÔNG phải base64!
                image_inputs.append(ImageInput(
                    name=media_name,  # Đây phải là media_name từ upload response
                    input_type=ImageInputType.REFERENCE
                ))
                self.log(f"  -> Using reference: {ref_id}")
            else:
                self.log(f"  -> Reference not in cache: {ref_id}", "WARN")

        if image_inputs:
            self.log(f"  -> Total {len(image_inputs)} references")

        # Generate
        success, images, error = self.api.generate_images(
            prompt=prompt,
            count=1,
            aspect_ratio=AspectRatio.LANDSCAPE,
            image_inputs=image_inputs if image_inputs else None
        )

        if success and images:
            # Download
            downloaded = self.api.download_image(
                images[0],
                output_path.parent,
                output_path.stem
            )

            if downloaded:
                self.log(f"  -> Saved: {downloaded}", "OK")
                return True

        self.log(f"  -> Generate failed: {error}", "ERROR")
        return False

    # =========================================================================
    # MAIN: Load prompts và generate
    # =========================================================================

    def load_prompts(self) -> List[Dict]:
        """Load prompts từ Excel."""
        import openpyxl

        # Find Excel file
        excel_files = list(self.prompts_path.glob("*_prompts.xlsx")) + \
                     list(self.prompts_path.glob("*.xlsx"))

        if not excel_files:
            # Try project root
            excel_files = list(self.project_path.glob("*_prompts.xlsx")) + \
                         list(self.project_path.glob("*.xlsx"))

        if not excel_files:
            self.log("No Excel file found!", "ERROR")
            return []

        excel_path = excel_files[0]
        self.log(f"Loading prompts from: {excel_path}")

        prompts = []
        wb = openpyxl.load_workbook(excel_path, read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1]]

            if not headers:
                continue

            # Find columns
            id_col = prompt_col = ref_col = None

            for i, h in enumerate(headers):
                if not h:
                    continue
                h_lower = str(h).lower()

                if 'id' in h_lower and id_col is None:
                    id_col = i
                if 'english' in h_lower and 'prompt' in h_lower:
                    prompt_col = i
                elif h_lower == 'img_prompt' and prompt_col is None:
                    prompt_col = i
                elif 'prompt' in h_lower and prompt_col is None and 'video' not in h_lower:
                    prompt_col = i
                if h_lower in ['references', 'ref', 'nv', 'characters', 'refs']:
                    ref_col = i

            if id_col is None or prompt_col is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(id_col, prompt_col):
                    continue

                pid = row[id_col]
                prompt = row[prompt_col]

                if not pid or not prompt:
                    continue

                # Parse references
                refs = []
                if ref_col is not None and ref_col < len(row) and row[ref_col]:
                    ref_str = str(row[ref_col])
                    for sep in [',', ';', '|', ' ']:
                        if sep in ref_str:
                            refs = [r.strip() for r in ref_str.split(sep) if r.strip()]
                            break
                    if not refs:
                        refs = [ref_str.strip()]

                # Determine output path
                pid_str = str(pid).strip()
                if pid_str.startswith('nv') or pid_str.startswith('loc'):
                    out_path = self.nv_path / f"{pid_str}.png"
                else:
                    out_path = self.img_path / f"{pid_str}.png"

                prompts.append({
                    'id': pid_str,
                    'prompt': str(prompt),
                    'references': refs,
                    'output_path': out_path,
                    'sheet': sheet_name
                })

        wb.close()
        self.log(f"Loaded {len(prompts)} prompts", "OK")
        return prompts

    def run(self, skip_existing: bool = True) -> Dict:
        """
        Chạy toàn bộ pipeline.

        Returns:
            Dict với kết quả
        """
        self.log("=" * 60)
        self.log("VTV GENERATOR WITH REFERENCES")
        self.log("=" * 60)

        results = {"success": 0, "failed": 0, "skipped": 0}

        # Step 1: Upload all references
        self.upload_all_references()

        if not self.reference_cache:
            self.log("No references uploaded. Continuing without references...", "WARN")

        # Step 2: Load prompts
        prompts = self.load_prompts()

        if not prompts:
            self.log("No prompts to process!", "ERROR")
            return results

        # Filter existing
        if skip_existing:
            prompts = [p for p in prompts if not p['output_path'].exists()]
            self.log(f"After filtering existing: {len(prompts)} to generate")

        if not prompts:
            self.log("All images already exist!", "OK")
            return results

        # Step 3: Generate images
        self.log("=" * 50)
        self.log(f"GENERATING {len(prompts)} IMAGES")
        self.log("=" * 50)

        for i, p in enumerate(prompts):
            if self._stop:
                break

            pid = p['id']
            prompt = p['prompt']
            refs = p['references']
            out_path = p['output_path']

            self.log(f"\n[{i+1}/{len(prompts)}] {pid}")
            self.log(f"  Prompt: {prompt[:60]}...")

            if refs:
                self.log(f"  References: {refs}")

            success = self.generate_with_references(prompt, refs, out_path)

            if success:
                results["success"] += 1
            else:
                results["failed"] += 1

            # Delay
            if i < len(prompts) - 1:
                time.sleep(self.delay)

        # Summary
        self.log("\n" + "=" * 50)
        self.log("SUMMARY")
        self.log("=" * 50)
        self.log(f"Success: {results['success']}", "OK")
        self.log(f"Failed: {results['failed']}", "ERROR" if results['failed'] > 0 else "INFO")

        return results

    def compose_video(
        self,
        voice_path: Optional[str] = None,
        srt_path: Optional[str] = None,
        output_name: str = "final_video.mp4",
        config: Optional['VideoConfig'] = None
    ) -> Optional[str]:
        """
        Ghép video từ ảnh đã generate + voice + phụ đề.

        Args:
            voice_path: Đường dẫn file voice (mp3/wav). Nếu None, tự tìm trong project.
            srt_path: Đường dẫn file SRT. Nếu None, tự tìm trong project.
            output_name: Tên file video output.
            config: VideoConfig (optional).

        Returns:
            Đường dẫn video output nếu thành công.
        """
        if not VIDEO_COMPOSER_AVAILABLE:
            self.log("Video composer not available. Please install FFmpeg.", "ERROR")
            return None

        self.log("=" * 50)
        self.log("COMPOSING FINAL VIDEO")
        self.log("=" * 50)

        # Tìm voice file
        if not voice_path:
            voice_files = list(self.project_path.glob("*.mp3")) + list(self.project_path.glob("*.wav"))
            if voice_files:
                voice_path = str(voice_files[0])
                self.log(f"Found voice: {voice_path}")
            else:
                self.log("No voice file found in project!", "ERROR")
                return None

        # Tìm SRT file
        if not srt_path:
            srt_files = list(self.project_path.glob("*.srt"))
            if srt_files:
                srt_path = str(srt_files[0])
                self.log(f"Found SRT: {srt_path}")
            else:
                self.log("No SRT file found. Video will be without subtitles.", "WARN")

        # Tìm Excel file
        excel_files = list(self.prompts_path.glob("*.xlsx"))
        if not excel_files:
            self.log("No Excel file found!", "ERROR")
            return None

        excel_path = str(excel_files[0])
        output_path = str(self.project_path / output_name)

        self.log(f"Excel: {excel_path}")
        self.log(f"Voice: {voice_path}")
        self.log(f"SRT: {srt_path or 'None'}")
        self.log(f"Output: {output_path}")

        try:
            composer = VideoComposer(config)

            if composer.compose_video(excel_path, voice_path, output_path, srt_path):
                self.log(f"Video created: {output_path}", "OK")
                return output_path
            else:
                self.log("Failed to create video!", "ERROR")
                return None

        except Exception as e:
            self.log(f"Video compose error: {e}", "ERROR")
            return None

    def stop(self):
        """Stop processing."""
        self._stop = True


# =============================================================================
# CLI
# =============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="VTV Generator với Reference Images + Video Composer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Project structure:
    PROJECTS/1/
    ├── nv/           # Ảnh reference (nv1.png, nv2.png, loc1.png, ...)
    ├── img/          # Output images
    ├── prompts/      # Excel file với prompts
    │   └── xxx_prompts.xlsx
    ├── voice.mp3     # Voice file (for video compose)
    └── subtitles.srt # SRT file (for video compose)

Examples:
    # Generate images only
    python vtv_with_references.py PROJECTS/1 "ya29.xxx..."

    # Generate images + compose video
    python vtv_with_references.py PROJECTS/1 "ya29.xxx..." --compose

    # Compose video only (skip image generation)
    python vtv_with_references.py PROJECTS/1 --compose-only
"""
    )

    parser.add_argument("project_path", help="Đường dẫn project")
    parser.add_argument("bearer_token", nargs="?", default="", help="Bearer token (không cần nếu --compose-only)")
    parser.add_argument("--compose", action="store_true", help="Ghép video sau khi generate ảnh")
    parser.add_argument("--compose-only", action="store_true", help="Chỉ ghép video (không generate ảnh)")
    parser.add_argument("--output", default="final_video.mp4", help="Tên file video output")
    parser.add_argument("--workers", type=int, default=2, help="Số workers parallel")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay giữa các request")

    args = parser.parse_args()

    # Kiểm tra args
    if not args.compose_only and not args.bearer_token:
        parser.error("bearer_token is required unless using --compose-only")

    generator = VTVGenerator(
        project_path=args.project_path,
        bearer_token=args.bearer_token,
        parallel_workers=args.workers,
        delay=args.delay,
        verbose=True
    )

    # Flow
    if args.compose_only:
        # Chỉ ghép video
        video_path = generator.compose_video(output_name=args.output)
        sys.exit(0 if video_path else 1)
    else:
        # Generate images
        results = generator.run()

        # Ghép video nếu có flag --compose
        if args.compose and results["success"] > 0:
            generator.compose_video(output_name=args.output)

        sys.exit(0 if results["failed"] == 0 else 1)


if __name__ == "__main__":
    main()
